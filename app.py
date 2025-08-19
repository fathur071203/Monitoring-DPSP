from flask import Flask, render_template, abort, redirect, url_for, request, session, send_file
from datetime import datetime, timedelta
import calendar
import json
import os
import io
import openpyxl # Import library openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key_here' # Ganti dengan kunci rahasia yang kuat di produksi

# --- Konfigurasi Akun Admin (Baru) ---
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "password123" # Ganti dengan password yang lebih kuat di produksi

# Helper function to hash passwords (dummy, for demonstration)
def hash_password(password):
    # Dalam aplikasi nyata, gunakan library hashing yang kuat seperti bcrypt
    # Untuk tujuan demonstrasi, kita menggunakan hash sederhana.
    return str(hash(password))

# Hash password admin saat startup
ADMIN_PASSWORD_HASHED = hash_password(ADMIN_PASSWORD)

# Load data from data.json and other data files
def load_data():
    app_data = {
        "PJP_COMPANIES_DATA": [],
        "SKSP_TRIWULANAN_DATA": {},
        "SKSP_TAHUNAN_DATA": {},
        "FRAUD_REPORTS_DATA": {},
        "LTDBB_REPORTS_DATA": {},
        "P2P_COOPERATION_REPORTS_DATA": {},
        "DTTOT_REPORTS_DATA": {},
        "GANGGUAN_IT_REPORTS_DATA": [],
        "KEUANGAN_TRIWULANAN_DATA": {},
        "KEUANGAN_TAHUNAN_DATA": {}, 
        "PENTEST_REPORTS_DATA": {},
        "AUDIT_SI_REPORTS_DATA": {},
        "APUPPT_REPORTS_DATA": [],
        "MANAGEMENT_REPORTS_DATA": [] # NEW: Tambahkan ini untuk data Laporan Manajemen
    }

    # Load from data.json
    data_file_path = os.path.join(app.root_path, 'templates', 'data.json')
    if not os.path.exists(data_file_path):
        print(f"Error: data.json not found at {data_file_path}")
    else:
        with open(data_file_path, 'r') as f:
            data = json.load(f)
            for key, value in data.items():
                if key in app_data and isinstance(app_data[key], dict) and isinstance(value, dict):
                    app_data[key].update(value)
                elif key in app_data and isinstance(app_data[key], list) and isinstance(value, list):
                    app_data[key].extend(value) 
                else:
                    app_data[key] = value

    # Load from data_gangguanit.json
    gangguan_it_file_path = os.path.join(app.root_path, 'templates', 'data_gangguanit.json')
    if not os.path.exists(gangguan_it_file_path):
        print(f"Error: data_gangguanit.json not found at {gangguan_it_file_path}")
    else:
        with open(gangguan_it_file_path, 'r') as f:
            gangguan_it_data = json.load(f)
            app_data["GANGGUAN_IT_REPORTS_DATA"] = gangguan_it_data.get("GANGGUAN_IT_REPORTS_DATA", [])

    # Load from data_keuangantriwulanan.json
    keuangan_triwulanan_file_path = os.path.join(app.root_path, 'templates', 'data_keuangantriwulanan.json')
    if not os.path.exists(keuangan_triwulanan_file_path):
        print(f"Error: data_keuangantriwulanan.json not found at {keuangan_triwulanan_file_path}")
    else:
        with open(keuangan_triwulanan_file_path, 'r') as f:
            raw_keuangan_triwulanan_data = json.load(f).get("KEUANGAN_TRIWULANAN_DATA", [])
            transformed_keuangan_data = {}
            for report in raw_keuangan_triwulanan_data:
                sandi_pjp = report.get("sandi_pjp")
                tahun_laporan = str(report.get("tahun_laporan"))
                periode_laporan = report.get("periode_laporan")

                if sandi_pjp not in transformed_keuangan_data:
                    transformed_keuangan_data[sandi_pjp] = {}
                if tahun_laporan not in transformed_keuangan_data[sandi_pjp]:
                    transformed_keuangan_data[sandi_pjp][tahun_laporan] = {}
                transformed_keuangan_data[sandi_pjp][tahun_laporan][periode_laporan] = report
            app_data["KEUANGAN_TRIWULANAN_DATA"] = transformed_keuangan_data

    # Load from data_keuangantahunan.json
    keuangan_tahunan_file_path = os.path.join(app.root_path, 'templates', 'data_keuangantahunan.json')
    if not os.path.exists(keuangan_tahunan_file_path):
        print(f"Error: data_keuangantahunan.json not found at {keuangan_tahunan_file_path}")
    else:
        with open(keuangan_tahunan_file_path, 'r') as f:
            raw_keuangan_tahunan_data = json.load(f).get("KEUANGAN_TAHUNAN_DATA", []) 
            transformed_keuangan_tahunan_data = {}
            for report in raw_keuangan_tahunan_data:
                sandi_pjp = report.get("sandi_pjp")
                tahun_raw = report.get("tahun_laporan") 
                
                if sandi_pjp is None:
                    print(f"Warning: Skipping report due to missing 'sandi_pjp' field: {report}")
                    continue
                if tahun_raw is None:
                    print(f"Warning: Skipping report for {sandi_pjp} due to missing 'tahun_laporan' field: {report}")
                    continue
                
                try:
                    tahun_laporan = str(int(tahun_raw)) 
                except (ValueError, TypeError):
                    print(f"Warning: Skipping report for {sandi_pjp} due to invalid 'tahun_laporan' value: {tahun_raw}")
                    continue

                report["total_liabilitas"] = report.get("total_hutang", 0)
                report["liabilitas_lancar"] = report.get("hutang_jangka_pendek", 0)
                report["liabilitas_tidak_lancar"] = report.get("hutang_jangka_panjang", 0)
                
                report["pendapatan"] = report.get("total_pendapatan", 0)
                report["pendapatan_fee"] = report.get("pendapatan_fee", 0)

                report["laba_bersih"] = report.get("laba", 0) - report.get("rugi", 0)
                
                report["aset_tidak_lancar"] = report.get("aset_tetap", 0)
                report["kas_dan_setara_kas"] = report.get("kas_dan_setara_kas", 0)

                report["ekuitas"] = report.get("total_ekuitas", 0)
                report["beban_operasional"] = report.get("beban_operasional", 0)
                report["total_beban"] = report.get("total_beban", 0)

                report["beban_pokok_penjualan"] = 0 
                report["laba_kotor"] = 0 
                report["arus_kas_operasi"] = 0
                report["arus_kas_investasi"] = 0
                report["arus_kas_pendanaan"] = 0
                report["saldo_kas_akhir"] = 0


                if sandi_pjp not in transformed_keuangan_tahunan_data:
                    transformed_keuangan_tahunan_data[sandi_pjp] = {}
                transformed_keuangan_tahunan_data[sandi_pjp][tahun_laporan] = report
            app_data["KEUANGAN_TAHUNAN_DATA"] = transformed_keuangan_tahunan_data

    # Load from data_pentest.json
    pentest_file_path = os.path.join(app.root_path, 'templates', 'data_pentest.json')
    if not os.path.exists(pentest_file_path):
        print(f"Error: data_pentest.json not found at {pentest_file_path}")
    else:
        with open(pentest_file_path, 'r') as f:
            raw_pentest_data = json.load(f).get("PENTEST_REPORTS_DATA", [])
            transformed_pentest_data = {}
            for report in raw_pentest_data:
                sandi_pjp = report.get("sandi_pjp")
                tahun_raw = report.get("tahun_laporan")

                if sandi_pjp is None:
                    print(f"Warning: Skipping pentest report due to missing 'sandi_pjp' field: {report}")
                    continue
                if tahun_raw is None:
                    print(f"Warning: Skipping pentest report for {sandi_pjp} due to missing 'tahun_laporan' field: {report}")
                    continue
                
                try:
                    tahun_laporan = str(int(tahun_raw))
                except (ValueError, TypeError):
                    print(f"Warning: Skipping pentest report for {sandi_pjp} due to invalid 'tahun_laporan' value: {tahun_raw}")
                    continue
                
                report["temuan_low"] = report.get("temuan_low", 0)
                report["temuan_medium"] = report.get("temuan_medium", 0)
                report["temuan_high"] = report.get("temuan_high", 0)
                report["temuan_critical"] = report.get("temuan_critical", 0)
                report["jumlah_temuan"] = report.get("jumlah_temuan", 0)
                report["jumlah_temuan_diselesaikan"] = report.get("jumlah_temuan_diselesaikan", 0)
                report["jumlah_temuan_belum_diselesaikan"] = report.get("jumlah_temuan_belum_diselesaikan", 0)

                if sandi_pjp not in transformed_pentest_data:
                    transformed_pentest_data[sandi_pjp] = {}
                transformed_pentest_data[sandi_pjp][tahun_laporan] = report
            app_data["PENTEST_REPORTS_DATA"] = transformed_pentest_data

    # NEW: Load from data_auditsi.json
    auditsi_file_path = os.path.join(app.root_path, 'templates', 'data_auditsi.json')
    if not os.path.exists(auditsi_file_path):
        print(f"Error: data_auditsi.json not found at {auditsi_file_path}")
    else:
        with open(auditsi_file_path, 'r') as f:
            raw_auditsi_data = json.load(f).get("AUDIT_SI_REPORTS_DATA", [])
            print(f"DEBUG: raw_auditsi_data loaded: {raw_auditsi_data[:2]}...")
            transformed_auditsi_data = {}
            for report in raw_auditsi_data:
                sandi_pjp = report.get("sandi_pjp")
                tahun_raw = report.get("tahun_laporan")

                if sandi_pjp is None:
                    print(f"Warning: Skipping Audit SI report due to missing 'sandi_pjp' field: {report}")
                    continue
                if tahun_raw is None:
                    print(f"Warning: Skipping Audit SI report for {sandi_pjp} due to missing 'tahun_laporan' field: {report}")
                    continue
                
                try:
                    tahun_laporan = str(int(tahun_raw))
                except (ValueError, TypeError):
                    print(f"Warning: Skipping Audit SI report for {sandi_pjp} due to invalid 'tahun_laporan' value: {tahun_raw}")
                    continue
                
                # Ensure all expected fields are present, default to 0 if not
                report["confidentiality"] = report.get("confidentiality", 0)
                report["integrity"] = report.get("integrity", 0)
                report["availability"] = report.get("availability", 0)
                report["authenticity"] = report.get("authenticity", 0)
                report["non_repudiation"] = report.get("non_repudiation", 0)
                report["jumlah_temuan"] = report.get("jumlah_temuan", 0)
                report["jumlah_temuan_diselesaikan"] = report.get("jumlah_temuan_diselesaikan", 0)
                report["jumlah_temuan_belum_diselesaikan"] = report.get("jumlah_temuan_belum_diselesaikan", 0)

                if sandi_pjp not in transformed_auditsi_data:
                    transformed_auditsi_data[sandi_pjp] = {}
                transformed_auditsi_data[sandi_pjp][tahun_laporan] = report
            app_data["AUDIT_SI_REPORTS_DATA"] = transformed_auditsi_data
            print(f"DEBUG: transformed_auditsi_data after processing: {transformed_auditsi_data.keys()}")

    # NEW: Load from data_apuppt.json
    apuppt_file_path = os.path.join(app.root_path, 'templates', 'data_apuppt.json')
    if not os.path.exists(apuppt_file_path):
        print(f"Error: data_apuppt.json not found at {apuppt_file_path}")
    else:
        with open(apuppt_file_path, 'r') as f:
            apuppt_data = json.load(f).get("APUPPT_REPORTS_DATA", [])
            app_data["APUPPT_REPORTS_DATA"] = apuppt_data
            print(f"DEBUG: APUPPT_REPORTS_DATA loaded. Count: {len(apuppt_data)}")

    # NEW: Load from data_manajemen.json
    manajemen_file_path = os.path.join(app.root_path, 'templates', 'data_manajemen.json')
    if not os.path.exists(manajemen_file_path):
        print(f"Error: data_manajemen.json not found at {manajemen_file_path}")
    else:
        with open(manajemen_file_path, 'r') as f:
            manajemen_data = json.load(f).get("MANAGEMENT_REPORTS_DATA", [])
            # Transform manajemen_data into a dictionary for easier lookup by sandi_pjp and tahun_laporan
            transformed_manajemen_data = {}
            for report in manajemen_data:
                sandi_pjp = report.get("sandi_pjp")
                tahun_laporan = str(report.get("tahun_laporan"))
                if sandi_pjp not in transformed_manajemen_data:
                    transformed_manajemen_data[sandi_pjp] = {}
                transformed_manajemen_data[sandi_pjp][tahun_laporan] = report
            app_data["MANAGEMENT_REPORTS_DATA"] = transformed_manajemen_data
            print(f"DEBUG: MANAGEMENT_REPORTS_DATA loaded. Count: {len(manajemen_data)}")


    # Hash passwords for PJP_COMPANIES_DATA after loading if they are placeholders
    for pjp in app_data.get("PJP_COMPANIES_DATA", []):
        if "password_hash" in pjp and pjp["password_hash"] == "b2f6b86f3f0e0c8d7b3a9e4f5a6b7c8d9e0f1a2b3c4d5e6f7a8b9c0d1e2f3a4b":
            pjp_name_parts = pjp["nama"].split(' ')
            if len(pjp_name_parts) > 1:
                dummy_pjp_password = pjp_name_parts[1].lower() + "pass"
            else:
                dummy_pjp_password = pjp["nama"].lower().replace('.', '') + "pass"
            pjp["password_hash"] = hash_password(dummy_pjp_password)
    return app_data

# Global data loaded at app startup
APP_DATA = load_data()
PJP_COMPANIES_DATA = APP_DATA["PJP_COMPANIES_DATA"]
SKSP_TRIWULANAN_DATA = APP_DATA["SKSP_TRIWULANAN_DATA"]
SKSP_TAHUNAN_DATA = APP_DATA["SKSP_TAHUNAN_DATA"]
FRAUD_REPORTS_DATA = APP_DATA["FRAUD_REPORTS_DATA"]
LTDBB_REPORTS_DATA = APP_DATA["LTDBB_REPORTS_DATA"]
P2P_COOPERATION_REPORTS_DATA = APP_DATA["P2P_COOPERATION_REPORTS_DATA"]
DTTOT_REPORTS_DATA = APP_DATA["DTTOT_REPORTS_DATA"]
GANGGUAN_IT_REPORTS_DATA = APP_DATA["GANGGUAN_IT_REPORTS_DATA"]
KEUANGAN_TRIWULANAN_DATA = APP_DATA["KEUANGAN_TRIWULANAN_DATA"]
KEUANGAN_TAHUNAN_DATA = APP_DATA["KEUANGAN_TAHUNAN_DATA"] 
PENTEST_REPORTS_DATA = APP_DATA["PENTEST_REPORTS_DATA"] 
AUDIT_SI_REPORTS_DATA = APP_DATA["AUDIT_SI_REPORTS_DATA"]
APUPPT_REPORTS_DATA = APP_DATA["APUPPT_REPORTS_DATA"]
MANAGEMENT_REPORTS_DATA = APP_DATA["MANAGEMENT_REPORTS_DATA"] # NEW: Tambahkan ini

# Helper function to convert month number to name
def get_month_name(month_num_or_name):
    # Mapping from numeric string to full month name
    month_names_map = {
        "01": "Januari", "02": "Februari", "03": "Maret", "04": "April", "05": "Mei", "06": "Juni",
        "07": "Juli", "08": "Agustus", "09": "September", "10": "Oktober", "11": "November", "12": "Desember",
        "Januari": "Januari", "Februari": "Februari", "Maret": "Maret", "April": "April", "Mei": "Mei", "Juni": "Juni",
        "Juli": "Juli", "Agustus": "Agustus", "September": "September", "Oktober": "Oktober", "November": "November", "Desember": "Desember"
    }
    
    # If input is a numeric string (e.g., "01"), return mapped name
    if isinstance(month_num_or_name, str) and month_num_or_name.isdigit() and len(month_num_or_name) <= 2:
        return month_names_map.get(month_num_or_name, month_num_or_name)
    
    # If it's already a month name (e.g., "Agustus") or "TW1", return as is
    return month_names_map.get(month_num_or_name, month_num_or_name) 

# Helper function to convert month name to number for sorting
def get_month_number(month_name_or_num):
    month_map = {
        "Januari": 1, "Februari": 2, "Maret": 3, "April": 4, "Mei": 5, "Juni": 6,
        "Juli": 7, "Agustus": 8, "September": 9, "Oktober": 10, "November": 11, "Desember": 12,
        "01": 1, "02": 2, "03": 3, "04": 4, "05": 5, "06": 6,
        "07": 7, "08": 8, "09": 9, "10": 10, "11": 11, "12": 12,
        "TW1": 1, "TW2": 4, "TW3": 7, "TW4": 10, # Assign a numeric value for quarterly reports for sorting
        "Q1": 1, "Q2": 4, "Q3": 7, "Q4": 10 # For quarterly financial reports
    }
    return month_map.get(month_name_or_num, 0) # Default to 0 if not found


@app.route('/')
def root():
    # Redirect ke halaman login jika belum login
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    total_pjp = len(PJP_COMPANIES_DATA)
    pjp_uploaded_reports = 0 
    pjp_report_status_distribution = {"Telah Melapor": 0, "Belum Melapor": 0, "Terlambat Melapor": 0}

    reporting_activity = {} 

    pjp_report_counts = {pjp["sandi"]: {"nama": pjp["nama"], "total_reports": 0, "last_report_date": None} for pjp in PJP_COMPANIES_DATA}

    today = datetime.now()

    for pjp_data in PJP_COMPANIES_DATA: 
        pjp_sandi = pjp_data["sandi"] 
        has_reported_this_year = False
        
        current_pjp_fraud_data = FRAUD_REPORTS_DATA.get(pjp_sandi, {})

        for year_str, months_data in current_pjp_fraud_data.items(): 
            year_int = int(year_str) 
            for month_str, report_data in months_data.items():
                report_date_str = report_data.get("Created at")
                if report_date_str:
                    try:
                        report_date = datetime.strptime(report_date_str, '%Y-%m-%d')
                    except ValueError:
                        print(f"Warning: Invalid date format for report {report_data.get('Nomor Surat')} in {pjp_sandi}, skipping.")
                        continue

                    current_last_report_date = pjp_report_counts[pjp_sandi]["last_report_date"]
                    if current_last_report_date is None or report_date > datetime.strptime(current_last_report_date, '%Y-%m-%d'):
                        pjp_report_counts[pjp_sandi]["last_report_date"] = report_date_str

                    if report_date.year == today.year:
                        has_reported_this_year = True

                        month_key = f"{report_date.year}-{report_date.month:02d}"
                        reporting_activity[month_key] = reporting_activity.get(month_key, 0) + 1

                        pjp_report_counts[pjp_sandi]["total_reports"] += 1
        
        reports_this_year = []
        if pjp_sandi in FRAUD_REPORTS_DATA:
            for year_str, months_data in FRAUD_REPORTS_DATA[pjp_sandi].items():
                if int(year_str) == today.year:
                    for month_str, report_data in months_data.items():
                        reports_this_year.append(report_data)
        
        if not reports_this_year:
            pjp_report_status_distribution["Belum Melapor"] += 1
        else:
            has_timely_report = False
            has_late_report = False
            for report_data in reports_this_year:
                report_date_str = report_data.get("Created at")
                if report_date_str:
                    try:
                        report_date = datetime.strptime(report_date_str, '%Y-%m-%d')
                        due_year = report_date.year
                        due_month = report_date.month + 1
                        if due_month > 12:
                            due_month = 1
                            due_year += 1
                        due_date = datetime(due_year, due_month, 5) 
                        
                        if report_date <= due_date:
                            has_timely_report = True
                        else:
                            has_late_report = True
                    except ValueError:
                        pass 

            if has_timely_report:
                pjp_report_status_distribution["Telah Melapor"] += 1
            elif has_late_report:
                pjp_report_status_distribution["Terlambat Melapor"] += 1
            else:
                pjp_report_status_distribution["Belum Melapor"] += 1


    sorted_reporting_activity_keys = sorted(reporting_activity.keys())
    reporting_activity_labels = [f"{get_month_name(key.split('-')[1])} {key.split('-')[0]}" for key in sorted_reporting_activity_keys]
    reporting_activity_data = [reporting_activity[key] for key in sorted_reporting_activity_keys]

    sorted_pjp_by_reports = sorted(pjp_report_counts.values(), key=lambda x: x["total_reports"], reverse=True)
    top_active_pjp = sorted_pjp_by_reports[:5]

    inactive_pjps = []
    for pjp_sandi, data in pjp_report_counts.items():
        last_report_year = None
        if data["last_report_date"]:
            last_report_year = datetime.strptime(data["last_report_date"], '%Y-%m-%d').year
        
        if data["total_reports"] == 0 or (last_report_year and last_report_year < today.year):
            inactive_pjps.append(data)

    top_inactive_pjp = sorted(inactive_pjps, key=lambda x: x["last_report_date"] if x["last_report_date"] else "0000-00-00")[:5]


    overview_data = {
        "total_pjp": total_pjp,
        "pjp_uploaded_reports": pjp_report_status_distribution["Telah Melapor"], 
        "pjp_not_uploaded_reports": pjp_report_status_distribution["Belum Melapor"], 
        "pjp_report_status_distribution": pjp_report_status_distribution,
        "reporting_activity_labels": reporting_activity_labels,
        "reporting_activity_data": reporting_activity_data,
        "top_active_pjp": top_active_pjp,
        "top_inactive_pjp": inactive_pjps, # Changed to show all inactive PJPs
        "recent_activities": [ 
            {"date": "2024-07-15", "description": "PJP 'PT. ABC Remit' mengajukan perpanjangan izin."},
            {"date": "2024-07-10", "description": "PJP 'PT. XYZ E-Wallet' meluncurkan fitur baru."},
            {"date": "2024-07-01", "description": "Audit rutin PJP 'PT. Payment Solutions' selesai."},
        ]
    }

    return render_template('index.html', overview_data=overview_data)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == ADMIN_USERNAME and hash_password(password) == ADMIN_PASSWORD_HASHED:
            session['logged_in'] = True
            session['username'] = ADMIN_USERNAME 
            session['pjp_sandi'] = None 
            session['pjp_name'] = "Admin" 
            return redirect(url_for('root'))
        else:
            return render_template('login.html', error='Invalid credentials')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('username', None) 
    session.pop('pjp_sandi', None) 
    session.pop('pjp_name', None) 
    return redirect(url_for('login'))


@app.route('/analisis_pjp')
def analisis_pjp():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('analisis_pjp.html', pjp_companies=PJP_COMPANIES_DATA)

@app.route('/analisis_pjp/<sandi>')
def show_pjp_detail(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)
    return render_template('pjp_detail.html', pjp_info=pjp_info, sandi_pjp=pjp_info["sandi"], pjp_name=pjp_info["nama"])

@app.route('/analisis_pjp/<sandi>/fraud_report')
def fraud_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_fraud_data_raw = FRAUD_REPORTS_DATA.get(sandi, {})

    all_years_in_data_str = sorted(list(pjp_fraud_data_raw.keys())) 
    all_years_in_data_int = sorted([int(y) for y in all_years_in_data_str]) 

    current_year = datetime.now().year
    min_year_display = min(all_years_in_data_int) if all_years_in_data_int else current_year - 2
    max_year_display = max(current_year, max(all_years_in_data_int) if all_years_in_data_int else current_year)


    monthly_upload_status = {}
    today = datetime.now()

    for year_int in range(min_year_display, max_year_display + 1): 
        year_str_key = str(year_int) 
        for month_num in range(1, 13):
            month_str = f"{month_num:02d}"
            period_key = f"{year_int}-{month_str}" 

            due_year = year_int 
            due_month = month_num + 1 
            if due_month > 12:
                due_month = 1
                due_year += 1
            
            try:
                due_date = datetime(due_year, due_month, 5) 
            except ValueError:
                due_date = datetime(due_year, due_month, calendar.monthrange(due_year, due_month)[1])


            report_entry = pjp_fraud_data_raw.get(year_str_key, {}).get(month_str) 

            status = {
                'year': year_int,
                'month_num': month_num,
                'month_name': get_month_name(month_str),
                'uploaded': False,
                'upload_date': None,
                'days_late': None,
                'due_date': due_date.strftime('%Y-%m-%d') 
            }

            if report_entry:
                status['uploaded'] = True
                status['upload_date'] = report_entry.get('Created at')

                if status['upload_date']:
                    uploaded_dt = datetime.strptime(status['upload_date'], '%Y-%m-%d')
                    if uploaded_dt > due_date:
                        status['days_late'] = (uploaded_dt - due_date).days
                    else:
                        status['days_late'] = 0 

                status['report_data'] = report_entry 
            else:
                report_period_date = datetime(year_int, month_num, 1)
                if today > due_date and report_period_date < datetime(today.year, today.month, 1):
                    status['days_late'] = (today - due_date).days
                else:
                    status['days_late'] = None

            monthly_upload_status[period_key] = status

    all_reports = []
    for year_str, months_data in pjp_fraud_data_raw.items(): 
        for month_str, report in months_data.items():
            report_copy = report.copy()
            report_copy["Tahun Laporan"] = int(year_str) 
            report_copy["Periode Laporan"] = month_str 
            all_reports.append(report_copy)

    all_reports.sort(key=lambda x: (x["Tahun Laporan"], int(x["Periode Laporan"])))

    available_years_for_filter = sorted(list(set([report["Tahun Laporan"] for report in all_reports])))
    if not available_years_for_filter:
        available_years_for_filter = [current_year] 

    min_potential_loss = min([report["Besar Potensi Kerugian"] for report in all_reports]) if all_reports else 0
    max_potential_loss = max([report["Besar Potensi Kerugian"] for report in all_reports]) if all_reports else 100000000 

    total_fraud = sum([report["Jumlah Fraud"] for report in all_reports])
    total_potensi_kerugian = sum([report["Besar Potensi Kerugian"] for report in all_reports])

    bulan_terbanyak_fraud = "N/A"
    if all_reports:
        fraud_counts_per_month = {}
        for report in all_reports:
            key = f"{get_month_name(report['Periode Laporan'])} {report['Tahun Laporan']}"
            fraud_counts_per_month[key] = fraud_counts_per_month.get(key, 0) + report["Jumlah Fraud"]
        if fraud_counts_per_month:
            bulan_terbanyak_fraud = max(fraud_counts_per_month, key=fraud_counts_per_month.get) + \
                                     f" ({fraud_counts_per_month[max(fraud_counts_per_month, key=fraud_counts_per_month.get)]} kasus)"


    rata_rata_kerugian_per_fraud = 0
    total_fraud_incidents = sum([report["Jumlah Fraud"] for report in all_reports])
    if total_fraud_incidents > 0:
        rata_rata_kerugian_per_fraud = total_potensi_kerugian / total_fraud_incidents

    kpis = {
        "total_fraud": total_fraud,
        "total_potensi_kerugian": total_potensi_kerugian,
        "bulan_terbanyak_fraud": bulan_terbanyak_fraud,
        "rata_rata_kerugian_per_fraud": rata_rata_kerugian_per_fraud
    }

    return render_template(
        'fraud_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data=json.dumps(all_reports), 
        available_years=json.dumps(available_years_for_filter), 
        min_potential_loss=min_potential_loss,
        max_potential_loss=max_potential_loss,
        kpis=kpis,
        monthly_upload_status=json.dumps(monthly_upload_status) 
    )

@app.route('/analisis_pjp/<sandi>/fraud_report_detail/<int:tahun>/<periode>')
def fraud_report_detail(sandi, tahun, periode):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = FRAUD_REPORTS_DATA.get(sandi, {}).get(str(tahun), {}).get(periode) 
    if report_data is None:
        abort(404)
    
    report_data["Nama PJP"] = pjp_info["nama"]
    report_data["Sandi PJP"] = pjp_info["sandi"]
    report_data["Nama Periode Laporan"] = get_month_name(periode)
    report_data["Tahun Laporan"] = tahun

    all_fraud_reports_for_pjp = []
    for y_str, months_data in FRAUD_REPORTS_DATA.get(sandi, {}).items(): 
        for m, r in months_data.items():
            report_copy = r.copy()
            report_copy["Tahun Laporan"] = int(y_str) 
            report_copy["Periode Laporan"] = m
            all_fraud_reports_for_pjp.append(report_copy)
    
    all_fraud_reports_for_pjp.sort(key=lambda x: datetime(x["Tahun Laporan"], int(x["Periode Laporan"]), 1))

    comparison_data_5_months = []
    current_report_date = datetime(tahun, int(periode), 1)
    current_report_index = -1
    for i, r in enumerate(all_fraud_reports_for_pjp):
        if datetime(r["Tahun Laporan"], int(r["Periode Laporan"]), 1) == current_report_date:
            current_report_index = i
            break
    
    comparison_data_5_months = []
    if current_report_index != -1:
        start_index = max(0, current_report_index - 4)
        comparison_data_5_months = all_fraud_reports_for_pjp[start_index : current_report_index + 1]
    
    comparison_same_month_yearly = []
    for y_str in sorted([yr for yr in FRAUD_REPORTS_DATA.get(sandi, {}).keys()]): 
        if periode in FRAUD_REPORTS_DATA.get(sandi, {}).get(y_str, {}):
            report_at_same_month = FRAUD_REPORTS_DATA[sandi][y_str][periode].copy()
            report_at_same_month["Tahun Laporan"] = int(y_str) 
            report_at_same_month["Periode Laporan"] = periode
            comparison_same_month_yearly.append(report_at_same_month)


    return render_template(
        'fraud_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report=report_data, 
        comparison_data_5_months=json.dumps(comparison_data_5_months),
        comparison_same_month_yearly=json.dumps(comparison_same_month_yearly)
    )


# New route for LTDBB Report Dashboard and Detail (Unified)
@app.route('/analisis_pjp/<sandi>/ltdbb_report')
def ltdbb_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_ltdbb_data_raw = LTDBB_REPORTS_DATA.get(sandi, {})

    all_years_in_data_str = sorted(list(pjp_ltdbb_data_raw.keys())) 
    all_years_in_data_int = sorted([int(y) for y in all_years_in_data_str]) 

    current_year = datetime.now().year
    min_year_display = min(all_years_in_data_int) if all_years_in_data_int else current_year - 2
    max_year_display = max(current_year, max(all_years_in_data_int) if all_years_in_data_int else current_year) 

    monthly_upload_status_ltdbb = {}
    today = datetime.now()

    for year_int in range(min_year_display, max_year_display + 1): 
        year_str_key = str(year_int) 
        for month_num in range(1, 13):
            month_str = f"{month_num:02d}"
            period_key = f"{year_int}-{month_str}"

            due_year = year_int
            due_month = month_num + 1
            if due_month > 12:
                due_month = 1
                due_year += 1

            try:
                due_date = datetime(due_year, due_month, 5) 
            except ValueError:
                due_date = datetime(due_year, due_month, calendar.monthrange(due_year, due_month)[1])


            report_entry = pjp_ltdbb_data_raw.get(year_str_key, {}).get(month_str) 

            status = {
                'year': year_int,
                'month_num': month_num,
                'month_name': get_month_name(month_str),
                'uploaded': False,
                'upload_date': None,
                'days_late': None,
                'due_date': due_date.strftime('%Y-%m-%d')
            }

            if report_entry:
                status['uploaded'] = True
                status['upload_date'] = report_entry.get('Created at')
                if status['upload_date']:
                    uploaded_dt = datetime.strptime(status['upload_date'], '%Y-%m-%d')
                    if uploaded_dt > due_date:
                        status['days_late'] = (uploaded_dt - due_date).days
                    else:
                        status['days_late'] = 0

                status['report_data'] = report_entry
            else:
                report_period_date = datetime(year_int, month_num, 1)
                if today > due_date and report_period_date < datetime(today.year, today.month, 1):
                    status['days_late'] = (today - due_date).days
                else:
                    status['days_late'] = None

            monthly_upload_status_ltdbb[period_key] = status


    all_ltdbb_reports = []
    for year_str, months_data in pjp_ltdbb_data_raw.items(): 
        for month_str, report in months_data.items():
            report_copy = report.copy()
            report_copy["Tahun Laporan"] = int(year_str)
            report_copy["Periode Laporan"] = month_str
            all_ltdbb_reports.append(report_copy)

    all_ltdbb_reports.sort(key=lambda x: (x["Tahun Laporan"], int(x["Periode Laporan"])))

    available_years_for_filter = sorted(list(set([report["Tahun Laporan"] for report in all_ltdbb_reports])))
    if not available_years_for_filter:
        available_years_for_filter = [current_year]

    initial_dashboard_data_ltdbb = all_ltdbb_reports

    total_outgoing_num = sum([report.get("Number Outgoing Transactions", 0) for report in all_ltdbb_reports])
    total_outgoing_amt = sum([report.get("Amount Outgoing Transactions", 0) for report in all_ltdbb_reports])
    total_incoming_num = sum([report.get("Number Incoming Transactions", 0) for report in all_ltdbb_reports])
    total_incoming_amt = sum([report.get("Amount Incoming Transactions", 0) for report in all_ltdbb_reports])
    total_domestic_num = sum([report.get("Number Domestic Transactions", 0) for report in all_ltdbb_reports])
    total_domestic_amt = sum([report.get("Amount Domestic Transactions", 0) for report in all_ltdbb_reports])

    monthly_outgoing_num = {}
    monthly_outgoing_amt = {}
    monthly_incoming_num = {}
    monthly_incoming_amt = {}
    monthly_domestic_num = {}
    monthly_domestic_amt = {}

    for report in all_ltdbb_reports:
        key = f"{get_month_name(report['Periode Laporan'])} {report['Tahun Laporan']}"
        monthly_outgoing_num[key] = monthly_outgoing_num.get(key, 0) + report.get("Number Outgoing Transactions", 0)
        monthly_outgoing_amt[key] = monthly_outgoing_amt.get(key, 0) + report.get("Amount Outgoing Transactions", 0)
        monthly_incoming_num[key] = monthly_incoming_num.get(key, 0) + report.get("Number Incoming Transactions", 0)
        monthly_incoming_amt[key] = monthly_incoming_amt.get(key, 0) + report.get("Amount Incoming Transactions", 0)
        monthly_domestic_num[key] = monthly_domestic_num.get(key, 0) + report.get("Number Domestic Transactions", 0)
        monthly_domestic_amt[key] = monthly_domestic_amt.get(key, 0) + report.get("Amount Domestic Transactions", 0)

    highest_outgoing_num_month = max(monthly_outgoing_num, key=monthly_outgoing_num.get) if monthly_outgoing_num else "N/A"
    highest_outgoing_amt_month = max(monthly_outgoing_amt, key=monthly_outgoing_amt.get) if monthly_outgoing_amt else "N/A"
    highest_incoming_num_month = max(monthly_incoming_num, key=monthly_incoming_num.get) if monthly_incoming_num else "N/A"
    highest_incoming_amt_month = max(monthly_incoming_amt, key=monthly_incoming_amt.get) if monthly_incoming_amt else "N/A"
    highest_domestic_num_month = max(monthly_domestic_num, key=monthly_domestic_num.get) if monthly_domestic_num else "N/A"
    highest_domestic_amt_month = max(monthly_domestic_amt, key=monthly_domestic_amt.get) if monthly_domestic_amt else "N/A"


    kpis_ltdbb = {
        "total_outgoing_num": total_outgoing_num,
        "total_outgoing_amt": total_outgoing_amt,
        "total_incoming_num": total_incoming_num,
        "total_incoming_amt": total_incoming_amt,
        "total_domestic_num": total_domestic_num,
        "total_domestic_amt": total_domestic_amt,
        "highest_outgoing_num_month": highest_outgoing_num_month,
        "highest_outgoing_amt_month": highest_outgoing_amt_month,
        "highest_incoming_num_month": highest_incoming_num_month,
        "highest_incoming_amt_month": highest_incoming_amt_month,
        "highest_domestic_num_month": highest_domestic_num_month,
        "highest_domestic_amt_month": highest_domestic_amt_month,
    }

    return render_template(
        'ltdbb_report.html', 
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_ltdbb=initial_dashboard_data_ltdbb, 
        available_years=available_years_for_filter, 
        kpis_ltdbb=kpis_ltdbb, 
        monthly_upload_status_ltdbb=monthly_upload_status_ltdbb 
    )

@app.route('/analisis_pjp/<sandi>/ltdbb_report/<int:tahun>/<periode>')
def ltdbb_report_detail(sandi, tahun, periode):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = LTDBB_REPORTS_DATA.get(sandi, {}).get(str(tahun), {}).get(periode) 
    if report_data is None:
        abort(404)
    
    report_data["Nama PJP"] = pjp_info["nama"]
    report_data["Sandi PJP"] = pjp_info["sandi"]
    report_data["Nama Periode Laporan"] = get_month_name(periode)
    report_data["Tahun Laporan"] = tahun

    all_ltdbb_reports_for_pjp = []
    for y_str, months_data in LTDBB_REPORTS_DATA.get(sandi, {}).items():
        for m, r in months_data.items():
            report_copy = r.copy()
            report_copy["Tahun Laporan"] = int(y_str)
            report_copy["Periode Laporan"] = m
            all_ltdbb_reports_for_pjp.append(report_copy)
    
    all_ltdbb_reports_for_pjp.sort(key=lambda x: (x["Tahun Laporan"], int(x["Periode Laporan"])))

    comparison_data_5_months = []
    current_report_date_obj = datetime(tahun, int(periode), 1)
    
    current_report_index = -1
    for i, r in enumerate(all_ltdbb_reports_for_pjp):
        if datetime(r["Tahun Laporan"], int(r["Periode Laporan"]), 1) == current_report_date_obj:
            current_report_index = i
            break

    if current_report_index != -1:
        start_index = max(0, current_report_index - 4)
        comparison_data_5_months = all_ltdbb_reports_for_pjp[start_index : current_report_index + 1]

    comparison_same_month_yearly = {
        "current_year_report": report_data,
        "previous_year_report": None
    }
    previous_year = tahun - 1
    report_last_year = LTDBB_REPORTS_DATA.get(sandi, {}).get(str(previous_year), {}).get(periode)
    
    if report_last_year:
        comparison_same_month_yearly["previous_year_report"] = report_last_year


    return render_template(
        'ltdbb_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report=report_data, 
        comparison_data_5_months=json.dumps(comparison_data_5_months),
        comparison_same_month_yearly=json.dumps(comparison_same_month_yearly)
    )


@app.route('/analisis_laporan')
def analisis_laporan():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('analisis_laporan.html')

@app.route('/analisis_pjp/<sandi>/laporan_sksp_triwulan/<int:tahun>/<periode>')
def show_sksp_triwulan_detail(sandi, tahun, periode):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    sksp_triwulan_report = SKSP_TRIWULANAN_DATA.get(sandi, {}).get(str(tahun), {}).get(periode) 

    if sksp_triwulan_report is None:
        abort(404)

    sksp_tahunan_plan = SKSP_TAHUNAN_DATA.get(sandi, {}).get(str(tahun)) 

    comparison_data = {
        "Pelaksana (Lv4)": {
            "rencana_tw": sksp_triwulan_report.get("Rencana PBK Pelaksana Tw", 0), 
            "realisasi_tw": sksp_triwulan_report.get("Realisasi PBK Pelaksana Tw", 0), 
            "rencana_tahun": sksp_tahunan_plan.get(f"Rencana PBK Pelaksana {periode}", 0) if sksp_tahunan_plan else 0
        },
        "Penyelia (Lv5)": {
            "rencana_tw": sksp_triwulan_report.get("Rencana PBK Penyelia Tw", 0), 
            "realisasi_tw": sksp_triwulan_report.get("Realisasi PBK Penyelia Tw", 0), 
            "rencana_tahun": sksp_tahunan_plan.get(f"Rencana PBK Penyelia {periode}", 0) if sksp_tahunan_plan else 0
        },
        "Pejabat Eksekutif (Lv6)": {
            "rencana_tw": sksp_triwulan_report.get("Rencana PBK Pejabat Eksekutif Tw", 0), 
            "realisasi_tw": sksp_triwulan_report.get("Realisasi PBK Pejabat Eksekutif Tw", 0), 
            "rencana_tahun": sksp_tahunan_plan.get(f"Rencana PBK Pejabat Eksekutif {periode}", 0) if sksp_tahunan_plan else 0
        },
        "Pejabat Direksi SK": {
            "rencana_tw": sksp_triwulan_report.get("Rencana SK Pejabat Direksi Tw", 0), 
            "realisasi_tw": sksp_triwulan_report.get("Realisasi SK Pejabat Direksi Tw", 0), 
            "rencana_tahun": sksp_tahunan_plan.get(f"Rencana SK Pejabat Direksi {periode}", 0) if sksp_tahunan_plan else 0
        }
    }


    return render_template(
        'sksp_triwulan_detail.html',
        pjp_name=pjp_info["nama"], 
        sandi_pjp=pjp_info["sandi"], 
        tahun_laporan=tahun, 
        periode_laporan=periode, 
        report_data=sksp_triwulan_report, 
        comparison_data=json.dumps(comparison_data)
    )

# NEW ROUTE: P2P Cooperation Report Dashboard
@app.route('/analisis_pjp/<sandi>/kerjasamap2p_report')
def kerjasamap2p_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_p2p_data_raw = P2P_COOPERATION_REPORTS_DATA.get(sandi, {})

    all_years_in_data_str = sorted(list(pjp_p2p_data_raw.keys()))
    all_years_in_data_int = sorted([int(y) for y in all_years_in_data_str])

    current_year = datetime.now().year
    min_year_display = min(all_years_in_data_int) if all_years_in_data_int else current_year - 2
    max_year_display = max(current_year, max(all_years_in_data_int) if all_years_in_data_int else current_year)

    monthly_upload_status_p2p = {}
    today = datetime.now()

    all_periods_in_data = set()
    for year_str, months_data in pjp_p2p_data_raw.items():
        for month_key_in_data in months_data.keys(): 
            all_periods_in_data.add((int(year_str), month_key_in_data))

    for year_int in range(min_year_display, max_year_display + 1):
        for month_num in range(1, 13):
            all_periods_in_data.add((year_int, get_month_name(f"{month_num:02d}")))

    sorted_periods = sorted(list(all_periods_in_data), key=lambda x: (x[0], get_month_number(x[1])))


    for year_int, month_key_in_data_original in sorted_periods:
        report_entry = pjp_p2p_data_raw.get(str(year_int), {}).get(month_key_in_data_original)
        
        month_name_for_display = get_month_name(month_key_in_data_original)
        period_key = f"{year_int}-{month_name_for_display}" 

        report_month_num_for_due_date = get_month_number(month_key_in_data_original)
        
        if report_month_num_for_due_date == 0:
            continue

        due_year = year_int
        due_month = report_month_num_for_due_date + 1
        if due_month > 12:
            due_month = 1
            due_year += 1
        try:
            due_date = datetime(due_year, due_month, 5) 
        except ValueError:
            due_date = datetime(due_year, due_month, calendar.monthrange(due_year, due_month)[1])

        status = {
            'year': year_int,
            'month_name': month_name_for_display,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            status['upload_date'] = report_entry.get('Created at')
            if status['upload_date']:
                uploaded_dt = datetime.strptime(status['upload_date'].split('T')[0], '%Y-%m-%d')
                if uploaded_dt > due_date:
                    status['days_late'] = (uploaded_dt - due_date).days
                else:
                    status['days_late'] = 0
            status['report_data'] = report_entry
        else:
            report_period_date = datetime(year_int, report_month_num_for_due_date, 1)
            if today > due_date and report_period_date < datetime(today.year, today.month, 1):
                status['days_late'] = (today - due_date).days
            else:
                status['days_late'] = None
        
        if period_key not in monthly_upload_status_p2p:
            monthly_upload_status_p2p[period_key] = status


    all_p2p_reports = []
    for year_str, months_data in pjp_p2p_data_raw.items():
        for m, report in months_data.items(): 
            report_copy = report.copy()
            report_copy["Tahun Laporan"] = int(year_str)
            report_copy["Periode Laporan"] = get_month_name(m) 
            all_p2p_reports.append(report_copy)

    all_p2p_reports.sort(key=lambda x: (x["Tahun Laporan"], get_month_number(x["Periode Laporan"])))

    available_years_for_filter = sorted(list(set([report["Tahun Laporan"] for report in all_p2p_reports])))
    if not available_years_for_filter:
        available_years_for_filter = [current_year]

    initial_dashboard_data_p2p = all_p2p_reports

    total_cooperation = sum([report.get("Jumlah Perusahaan Kerjasama P2P", 0) for report in all_p2p_reports])
    avg_cooperation_per_report = total_cooperation / len(all_p2p_reports) if len(all_p2p_reports) > 0 else 0

    monthly_new_cooperation = {}
    for report in all_p2p_reports:
        key = f"{get_month_name(report['Periode Laporan'])} {report['Tahun Laporan']}"
        monthly_new_cooperation[key] = monthly_new_cooperation.get(key, 0) + report.get("Jumlah Perusahaan Kerjasama P2P", 0)

    month_most_new_cooperation = max(monthly_new_cooperation, key=monthly_new_cooperation.get) if monthly_new_cooperation else "N/A"

    kpis_p2p = {
        "total_cooperation": total_cooperation,
        "avg_cooperation_per_report": avg_cooperation_per_report,
        "month_most_new_cooperation": month_most_new_cooperation,
    }

    return render_template(
        'kerjasamap2p_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_p2p=initial_dashboard_data_p2p, 
        available_years=available_years_for_filter, 
        kpis=kpis_p2p,
        monthly_upload_status_p2p=monthly_upload_status_p2p 
    )

@app.route('/analisis_pjp/<sandi>/kerjasamap2p_report_detail/<int:tahun>/<periode>')
def kerjasamap2p_report_detail(sandi, tahun, periode):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = None
    
    possible_month_keys = [periode] 
    if periode.isdigit():
        possible_month_keys.append(get_month_name(periode))
    else:
        month_num_str_map = {v: k for k, v in {
            "01": "Januari", "02": "Februari", "03": "Maret", "04": "April", "05": "Mei", "06": "Juni",
            "07": "Juli", "08": "Agustus", "09": "September", "10": "Oktober", "11": "November", "12": "Desember"
        }.items()}
        alt_numeric_periode = month_num_str_map.get(periode)
        if alt_numeric_periode:
            possible_month_keys.append(alt_numeric_periode)

    for key in set(possible_month_keys):
        report_data = P2P_COOPERATION_REPORTS_DATA.get(sandi, {}).get(str(tahun), {}).get(key)
        if report_data:
            break

    if report_data is None:
        abort(404)
    
    report_data["Nama PJP"] = pjp_info["nama"]
    report_data["Sandi PJP"] = pjp_info["sandi"]
    report_data["Nama Periode Laporan"] = get_month_name(periode) 
    report_data["Tahun Laporan"] = tahun

    all_p2p_reports_for_pjp = []
    for y_str, months_data in P2P_COOPERATION_REPORTS_DATA.get(sandi, {}).items():
        for m, r in months_data.items(): 
            report_copy = r.copy()
            report_copy["Tahun Laporan"] = int(y_str)
            report_copy["Periode Laporan"] = get_month_name(m) 
            all_p2p_reports_for_pjp.append(report_copy)
    
    all_p2p_reports_for_pjp.sort(key=lambda x: (x["Tahun Laporan"], get_month_number(x["Periode Laporan"])))

    comparison_data_5_months = []
    current_report_date_obj = datetime(tahun, get_month_number(periode), 1) 
    
    current_report_index = -1
    for i, r in enumerate(all_p2p_reports_for_pjp):
        if datetime(r["Tahun Laporan"], get_month_number(r["Periode Laporan"]), 1) == current_report_date_obj:
            current_report_index = i
            break

    if current_report_index != -1:
        start_index = max(0, current_report_index - 4)
        comparison_data_5_months = all_p2p_reports_for_pjp[start_index : current_report_index + 1]

    comparison_same_month_yearly = {
        "current_year_report": report_data,
        "previous_year_report": None
    }
    previous_year = tahun - 1
    
    for key in set(possible_month_keys):
        report_last_year = P2P_COOPERATION_REPORTS_DATA.get(sandi, {}).get(str(previous_year), {}).get(key)
        if report_last_year:
            comparison_same_month_yearly["previous_year_report"] = report_last_year
            prev_year_report_found = True
            break


    return render_template(
        'kerjasamap2p_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data, 
        comparison_data_5_months=json.dumps(comparison_data_5_months),
        comparison_same_month_yearly=json.dumps(comparison_same_month_yearly)
    )

# NEW ROUTE: DTTOT Report Dashboard
@app.route('/analisis_pjp/<sandi>/dttot_report')
def dttot_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_dttot_data_raw = DTTOT_REPORTS_DATA.get(sandi, {})

    all_years_in_data_str = sorted(list(pjp_dttot_data_raw.keys()))
    all_years_in_data_int = sorted([int(y) for y in all_years_in_data_str])

    current_year = datetime.now().year
    min_year_display = min(all_years_in_data_int) if all_years_in_data_int else current_year - 2
    max_year_display = max(current_year, max(all_years_in_data_int) if all_years_in_data_int else current_year)

    monthly_upload_status_dttot = {}
    today = datetime.now()

    all_periods_in_data = set()
    for year_str, months_data in pjp_dttot_data_raw.items():
        for month_key_in_data in months_data.keys():
            all_periods_in_data.add((int(year_str), month_key_in_data))

    for year_int in range(min_year_display, max_year_display + 1):
        for month_num in range(1, 13):
            all_periods_in_data.add((year_int, get_month_name(f"{month_num:02d}")))

    sorted_periods = sorted(list(all_periods_in_data), key=lambda x: (x[0], get_month_number(x[1])))

    for year_int, month_key_in_data_original in sorted_periods:
        report_entry = pjp_dttot_data_raw.get(str(year_int), {}).get(month_key_in_data_original)
        
        month_name_for_display = get_month_name(month_key_in_data_original)
        period_key = f"{year_int}-{month_name_for_display}"

        report_month_num_for_due_date = get_month_number(month_key_in_data_original)
        
        if report_month_num_for_due_date == 0:
            continue

        due_year = year_int
        due_month = report_month_num_for_due_date + 1
        if due_month > 12:
            due_month = 1
            due_year += 1
        try:
            due_date = datetime(due_year, due_month, 5) 
        except ValueError:
            due_date = datetime(due_year, due_month, calendar.monthrange(due_year, due_month)[1])

        status = {
            'year': year_int,
            'month_name': month_name_for_display,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            status['upload_date'] = report_entry.get('created_at')
            if status['upload_date']:
                uploaded_dt = datetime.strptime(status['upload_date'].split('T')[0], '%Y-%m-%d')
                if uploaded_dt > due_date:
                    status['days_late'] = (uploaded_dt - due_date).days
                else:
                    status['days_late'] = 0
            status['report_data'] = report_entry
        else:
            report_period_date = datetime(year_int, report_month_num_for_due_date, 1)
            if today > due_date and report_period_date < datetime(today.year, today.month, 1):
                status['days_late'] = (today - due_date).days
            else:
                status['days_late'] = None
        
        if period_key not in monthly_upload_status_dttot:
            monthly_upload_status_dttot[period_key] = status


    all_dttot_reports = []
    for year_str, months_data in pjp_dttot_data_raw.items():
        for m, report in months_data.items(): 
            report_copy = report.copy()
            report_copy["tahun_laporan"] = int(year_str)
            report_copy["periode_laporan"] = get_month_name(m)
            all_dttot_reports.append(report_copy)

    all_dttot_reports.sort(key=lambda x: (x["tahun_laporan"], get_month_number(x["periode_laporan"])))

    available_years_for_filter = sorted(list(set([report["tahun_laporan"] for report in all_dttot_reports])))
    if not available_years_for_filter:
        available_years_for_filter = [current_year]

    initial_dashboard_data_dttot = all_dttot_reports

    total_terduga_teroris = sum([report.get("jumlah_terduga_teroris", 0) for report in all_dttot_reports])
    
    org_counts = {}
    for report in all_dttot_reports:
        org = report.get("organisasi_teroris", "Tidak Diketahui")
        org_counts[org] = org_counts.get(org, 0) + 1
    most_common_org = max(org_counts, key=org_counts.get) if org_counts else "N/A"

    kpis_dttot = {
        "total_terduga_teroris": total_terduga_teroris,
        "most_common_org": most_common_org,
    }

    return render_template(
        'dttot_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_dttot=initial_dashboard_data_dttot, 
        available_years=available_years_for_filter, 
        kpis=kpis_dttot, 
        monthly_upload_status_dttot=monthly_upload_status_dttot 
    )

@app.route('/analisis_pjp/<sandi>/dttot_report_detail/<int:tahun>/<periode>')
def dttot_report_detail(sandi, tahun, periode):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = None
    possible_month_keys = [periode]
    if periode.isdigit():
        possible_month_keys.append(get_month_name(periode))
    else:
        month_num_str_map = {v: k for k, v in {
            "01": "Januari", "02": "Februari", "03": "Maret", "04": "April", "05": "Mei", "06": "Juni",
            "07": "Juli", "08": "Agustus", "09": "September", "10": "Oktober", "11": "November", "12": "Desember"
        }.items()}
        alt_numeric_periode = month_num_str_map.get(periode)
        if alt_numeric_periode:
            possible_month_keys.append(alt_numeric_periode)

    for key in set(possible_month_keys):
        report_data = DTTOT_REPORTS_DATA.get(sandi, {}).get(str(tahun), {}).get(key)
        if report_data:
            break

    if report_data is None:
        abort(404)
    
    report_data["nama_pjp"] = pjp_info["nama"]
    report_data["sandi_pjp"] = pjp_info["sandi"]
    report_data["nama_periode_laporan"] = get_month_name(periode)
    report_data["tahun_laporan"] = tahun

    all_dttot_reports_for_pjp = []
    for y_str, months_data in DTTOT_REPORTS_DATA.get(sandi, {}).items():
        for m, r in months_data.items():
            report_copy = r.copy()
            report_copy["tahun_laporan"] = int(y_str)
            report_copy["periode_laporan"] = get_month_name(m)
            all_dttot_reports_for_pjp.append(report_copy)
    
    all_dttot_reports_for_pjp.sort(key=lambda x: (x["tahun_laporan"], get_month_number(x["periode_laporan"]), 1))

    comparison_data_5_months = []
    current_report_date_obj = datetime(tahun, get_month_number(periode), 1)
    
    current_report_index = -1 
    for i, r in enumerate(all_dttot_reports_for_pjp):
        if datetime(r["tahun_laporan"], get_month_number(r["periode_laporan"]), 1) == current_report_date_obj:
            current_report_index = i
            break

    if current_report_index != -1:
        start_index = max(0, current_report_index - 4)
        comparison_data_5_months = all_dttot_reports_for_pjp[start_index : current_report_index + 1]

    comparison_same_month_yearly = {
        "current_year_report": report_data,
        "previous_year_report": None
    }
    previous_year = tahun - 1
    
    for key in set(possible_month_keys):
        report_last_year = DTTOT_REPORTS_DATA.get(sandi, {}).get(str(previous_year), {}).get(key)
        if report_last_year:
            comparison_same_month_yearly["previous_year_report"] = report_last_year
            break


    return render_template(
        'dttot_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data, 
        comparison_data_5_months=comparison_data_5_months, 
        comparison_same_month_yearly=comparison_same_month_yearly, 
    )

# NEW ROUTE: Gangguan IT Report Dashboard
@app.route('/analisis_pjp/<sandi>/gangguanit_report')
def gangguanit_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_gangguan_it_data_raw = [
        report for report in GANGGUAN_IT_REPORTS_DATA
        if report.get("sandi_pjp") == sandi
    ]

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in pjp_gangguan_it_data_raw])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year]

    total_incidents = len(pjp_gangguan_it_data_raw)
    total_potential_loss = sum([report.get("potensi_kerugian", 0) for report in pjp_gangguan_it_data_raw])

    jenis_gangguan_counts = {}
    for report in pjp_gangguan_it_data_raw:
        jenis = report.get("jenis_gangguan", "Tidak Diketahui")
        jenis_gangguan_counts[jenis] = jenis_gangguan_counts.get(jenis, 0) + 1
    most_common_incident_type = max(jenis_gangguan_counts, key=jenis_gangguan_counts.get) if jenis_gangguan_counts else "N/A"

    avg_loss_per_incident = total_potential_loss / total_incidents if total_incidents > 0 else 0

    kpis_gangguan_it = {
        "total_incidents": total_incidents,
        "total_potential_loss": total_potential_loss,
        "most_common_incident_type": most_common_incident_type,
        "avg_loss_per_incident": avg_loss_per_incident
    }

    return render_template(
        'gangguanit_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_gangguanit=pjp_gangguan_it_data_raw, 
        available_years=all_years_in_data, 
        kpis=kpis_gangguan_it
    )

# NEW ROUTE: Gangguan IT Report Detail
@app.route('/analisis_pjp/<sandi>/gangguanit_report_detail/<int:tahun>/<bulan>/<int:tanggal>/<nomor_surat>')
def gangguanit_report_detail(sandi, tahun, bulan, tanggal, nomor_surat):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = None
    for report in GANGGUAN_IT_REPORTS_DATA:
        if (report.get("sandi_pjp") == sandi and
            report.get("tahun_laporan") == tahun and
            report.get("bulan_laporan") == str(bulan).zfill(2) and 
            report.get("tanggal_laporan") == tanggal and
            report.get("nomor_surat") == nomor_surat):
            report_data = report
            break

    if report_data is None:
        abort(404)
    
    report_data["nama_pjp"] = pjp_info["nama"]
    report_data["sandi_pjp"] = pjp_info["sandi"]
    
    all_gangguan_it_reports_for_pjp = [
        r for r in GANGGUAN_IT_REPORTS_DATA
        if r.get("sandi_pjp") == sandi
    ]

    all_gangguan_it_reports_for_pjp.sort(key=lambda x: datetime.strptime(x["waktu_kejadian"].split('T')[0], '%Y-%m-%d'))

    comparison_data_5_incidents = []
    current_incident_datetime = datetime.strptime(report_data["waktu_kejadian"].split('T')[0], '%Y-%m-%d')
    
    current_incident_index = -1 
    for i, r in enumerate(all_gangguan_it_reports_for_pjp):
        if datetime.strptime(r["waktu_kejadian"].split('T')[0], '%Y-%m-%d') == current_incident_datetime and \
           r.get("nomor_surat") == nomor_surat: 
            current_incident_index = i
            break

    if current_incident_index != -1:
        start_index = max(0, current_incident_index - 4)
        comparison_data_5_incidents = all_gangguan_it_reports_for_pjp[start_index : current_incident_index + 1]

    comparison_same_type_yearly = []
    current_jenis_gangguan = report_data.get("jenis_gangguan")
    current_year_report = report_data

    filtered_by_type = [
        r for r in all_gangguan_it_reports_for_pjp
        if r.get("jenis_gangguan") == current_jenis_gangguan
    ]
    
    yearly_summary_for_type = {}
    for r in filtered_by_type: 
        year = r.get("tahun_laporan")
        if year not in yearly_summary_for_type:
            yearly_summary_for_type[year] = {"total_incidents": 0, "total_loss": 0}
        yearly_summary_for_type[year]["total_incidents"] += 1
        yearly_summary_for_type[year]["total_loss"] += r.get("potensi_kerugian", 0)

    for year in sorted(yearly_summary_for_type.keys()):
        comparison_same_type_yearly.append({
            "tahun_laporan": year,
            "total_incidents": yearly_summary_for_type[year]["total_incidents"],
            "total_loss": yearly_summary_for_type[year]["total_loss"]
        })

    return render_template(
        'gangguanit_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data, 
        comparison_data_5_incidents=comparison_data_5_incidents, 
        comparison_same_type_yearly=comparison_same_type_yearly, 
    )

# NEW ROUTE: Keuangan Triwulanan Report Dashboard
@app.route('/analisis_pjp/<sandi>/keuangantriwulanan_report')
def keuangantriwulanan_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_keuangan_triwulanan_data_raw = KEUANGAN_TRIWULANAN_DATA.get(sandi, {})

    all_keuangan_reports = []
    for year_str, periods_data in pjp_keuangan_triwulanan_data_raw.items():
        for period_key, report in periods_data.items():
            report_copy = report.copy()
            report_copy["tahun_laporan"] = int(year_str)
            report_copy["periode_laporan_nama"] = period_key 
            all_keuangan_reports.append(report_copy)

    all_keuangan_reports.sort(key=lambda x: (x["tahun_laporan"], get_month_number(x["periode_laporan_nama"])))

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in all_keuangan_reports])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year]

    latest_report = all_keuangan_reports[-1] if all_keuangan_reports else {}

    kpis_keuangan = {
        "modal_dasar": latest_report.get("modal_dasar", 0),
        "modal_disetor": latest_report.get("total_aset", 0), 
        "total_aset": latest_report.get("total_aset", 0),
        "total_hutang": latest_report.get("total_hutang", 0),
        "total_ekuitas": latest_report.get("total_ekuitas", 0),
        "total_pendapatan": latest_report.get("total_pendapatan", 0),
        "total_beban": latest_report.get("total_beban", 0),
        "laba_rugi_bersih": latest_report.get("laba", 0) - latest_report.get("rugi", 0)
    }

    current_assets = latest_report.get("aset_lancar", 0)
    current_liabilities = latest_report.get("hutang_jangka_pendek", 0)
    total_debt = latest_report.get("total_hutang", 0)
    total_equity = latest_report.get("total_ekuitas", 0)
    total_revenue = kpis_keuangan["total_pendapatan"]
    net_income = kpis_keuangan["laba_rugi_bersih"]

    kpis_keuangan["rasio_lancar"] = (current_assets / current_liabilities) if current_liabilities > 0 else 0
    kpis_keuangan["rasio_hutang_ekuitas"] = (total_debt / total_equity) if total_equity > 0 else 0
    kpis_keuangan["marjin_laba"] = (net_income / total_revenue) if total_revenue > 0 else 0

    monthly_upload_status_keuangan = {} 
    today = datetime.now()
    
    for year_int in all_years_in_data: 
        for quarter_name in ["Q1", "Q2", "Q3", "Q4"]:
            period_key = f"{year_int}-{quarter_name}"
            report_entry = pjp_keuangan_triwulanan_data_raw.get(str(year_int), {}).get(quarter_name)

            due_month_num = get_month_number(quarter_name) + 2 
            due_year = year_int
            if quarter_name == "Q4": 
                due_month_num = 1
                due_year += 1
            
            try:
                due_date = datetime(due_year, due_month_num, 15)
            except ValueError:
                due_date = datetime(due_year, due_month_num, calendar.monthrange(due_year, due_month_num)[1])

            status = {
                'year': year_int,
                'quarter_name': quarter_name,
                'uploaded': False,
                'upload_date': None,
                'days_late': None,
                'due_date': due_date.strftime('%Y-%m-%d')
            }

            if report_entry:
                status['uploaded'] = True
                status['upload_date'] = report_entry.get('created_at')
                if status['upload_date']:
                    uploaded_dt = datetime.strptime(status['upload_date'].split('T')[0], '%Y-%m-%d')
                    if uploaded_dt > due_date:
                        status['days_late'] = (uploaded_dt - due_date).days
                    else:
                        status['days_late'] = 0
                status['report_data'] = report_entry
            else:
                quarter_end_month = get_month_number(quarter_name) + 2 
                quarter_end_day = calendar.monthrange(year_int, quarter_end_month)[1]
                quarter_end_date = datetime(year_int, quarter_end_month, quarter_end_day)

                if today > due_date and quarter_end_date < datetime(today.year, today.month, 1):
                    status['days_late'] = (today - due_date).days
                else:
                    status['days_late'] = None
            
            monthly_upload_status_keuangan[period_key] = status


    return render_template(
        'keuangantriwulanan_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_keuangan=all_keuangan_reports, 
        available_years=all_years_in_data, 
        kpis=kpis_keuangan,
        monthly_upload_status_keuangan=monthly_upload_status_keuangan 
    )

@app.route('/analisis_pjp/<sandi>/keuangantriwulanan_report_detail/<int:tahun>/<periode>')
def keuangantriwulanan_report_detail(sandi, tahun, periode):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = KEUANGAN_TRIWULANAN_DATA.get(sandi, {}).get(str(tahun), {}).get(periode)
    if report_data is None:
        abort(404)
    
    report_data["nama_pjp"] = pjp_info["nama"]
    report_data["sandi_pjp"] = pjp_info["sandi"]
    report_data["nama_periode_laporan"] = periode 
    report_data["tahun_laporan"] = tahun

    all_keuangan_reports_for_pjp = []
    for year_str, periods_data in KEUANGAN_TRIWULANAN_DATA.get(sandi, {}).items():
        for period_key, r in periods_data.items():
            report_copy = r.copy()
            report_copy["tahun_laporan"] = int(year_str)
            report_copy["periode_laporan_nama"] = period_key
            all_keuangan_reports_for_pjp.append(report_copy)
    
    all_keuangan_reports_for_pjp.sort(key=lambda x: (x["tahun_laporan"], get_month_number(x["periode_laporan_nama"])))

    comparison_data_previous_quarter = None
    current_report_index = -1
    for i, r in enumerate(all_keuangan_reports_for_pjp):
        if r["tahun_laporan"] == tahun and r["periode_laporan_nama"] == periode:
            current_report_index = i
            break
    
    if current_report_index > 0:
        comparison_data_previous_quarter = all_keuangan_reports_for_pjp[current_report_index - 1]

    comparison_data_previous_year_same_quarter = None
    previous_year = tahun - 1
    if str(previous_year) in KEUANGAN_TRIWULANAN_DATA.get(sandi, {}) and \
       periode in KEUANGAN_TRIWULANAN_DATA.get(sandi, {}).get(str(previous_year), {}):
        comparison_data_previous_year_same_quarter = KEUANGAN_TRIWULANAN_DATA[sandi][str(previous_year)][periode]


    return render_template(
        'keuangantriwulanan_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data, 
        comparison_data_previous_quarter=comparison_data_previous_quarter, 
        comparison_data_previous_year_same_quarter=comparison_data_previous_year_same_quarter 
    )

# NEW ROUTE: Keuangan Tahunan Report Dashboard
@app.route('/analisis_pjp/<sandi>/keuangantahunan_report')
def keuangantahunan_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_keuangan_tahunan_data_raw = KEUANGAN_TAHUNAN_DATA.get(sandi, {})

    all_keuangan_tahunan_reports = []
    for year_str, report_data in pjp_keuangan_tahunan_data_raw.items():
        report_copy = report_data.copy()
        report_copy["tahun_laporan"] = int(year_str) 
        all_keuangan_tahunan_reports.append(report_copy)

    all_keuangan_tahunan_reports.sort(key=lambda x: x["tahun_laporan"])

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in all_keuangan_tahunan_reports])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year]

    latest_report = all_keuangan_tahunan_reports[-1] if all_keuangan_tahunan_reports else {}

    kpis_keuangan_tahunan = {
        "modal_dasar": latest_report.get("modal_dasar", 0),
        "modal_disetor": latest_report.get("modal_disetor", 0),
        "total_aset": latest_report.get("total_aset", 0),
        "total_liabilitas": latest_report.get("total_liabilitas", 0), 
        "ekuitas": latest_report.get("ekuitas", 0),
        "pendapatan": latest_report.get("pendapatan", 0), 
        "beban_operasional": latest_report.get("beban_operasional", 0),
        "laba_bersih": latest_report.get("laba_bersih", 0) 
    }

    aset_lancar = latest_report.get("aset_lancar", 0)
    liabilitas_lancar = latest_report.get("liabilitas_lancar", 0) 
    total_liabilitas = latest_report.get("total_liabilitas", 0) 
    ekuitas = latest_report.get("ekuitas", 0)
    pendapatan_for_ratio = kpis_keuangan_tahunan["pendapatan"]
    laba_bersih_for_ratio = kpis_keuangan_tahunan["laba_bersih"]
    total_aset = latest_report.get("total_aset", 0)

    kpis_keuangan_tahunan["rasio_lancar"] = (aset_lancar / liabilitas_lancar) if liabilitas_lancar > 0 else 0
    kpis_keuangan_tahunan["rasio_hutang_ekuitas"] = (total_liabilitas / ekuitas) if ekuitas > 0 else 0
    kpis_keuangan_tahunan["marjin_laba_bersih"] = (laba_bersih_for_ratio / pendapatan_for_ratio) if pendapatan_for_ratio > 0 else 0
    kpis_keuangan_tahunan["roa"] = (laba_bersih_for_ratio / total_aset) if total_aset > 0 else 0
    kpis_keuangan_tahunan["roe"] = (laba_bersih_for_ratio / ekuitas) if ekuitas > 0 else 0


    annual_upload_status_keuangantahunan = {}
    today = datetime.now()
    
    for year_int in all_years_in_data:
        period_key = str(year_int)
        report_entry = pjp_keuangan_tahunan_data_raw.get(str(year_int))

        due_date = datetime(year_int + 1, 3, 31) 

        status = {
            'year': year_int,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            upload_date_str = report_entry.get('created_at', report_entry.get('tanggal_surat'))
            if upload_date_str:
                try:
                    uploaded_dt = datetime.fromisoformat(upload_date_str.split('+')[0]) 
                    status['upload_date'] = uploaded_dt.strftime('%Y-%m-%d')
                    if uploaded_dt.date() > due_date.date(): 
                        status['days_late'] = (uploaded_dt.date() - due_date.date()).days
                    else:
                        status['days_late'] = 0
                except ValueError:
                    print(f"Warning: Could not parse upload_date for {sandi} year {year_int}: {upload_date_str}")
            status['report_data'] = report_entry
        else:
            if today.date() > due_date.date() and year_int < today.year: 
                status['days_late'] = (today.date() - due_date.date()).days
            else:
                status['days_late'] = None
        
        annual_upload_status_keuangantahunan[period_key] = status


    return render_template(
        'keuangantahunan_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_keuangantahunan=all_keuangan_tahunan_reports,
        available_years=all_years_in_data,
        kpis=kpis_keuangan_tahunan,
        annual_upload_status_keuangantahunan=annual_upload_status_keuangantahunan
    )

# NEW ROUTE: Pentest Report Dashboard
@app.route('/analisis_pjp/<sandi>/pentest_report')
def pentest_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_pentest_data_raw = PENTEST_REPORTS_DATA.get(sandi, {})

    all_pentest_reports = []
    for year_str, report_data in pjp_pentest_data_raw.items():
        report_copy = report_data.copy()
        report_copy["tahun_laporan"] = int(year_str)
        all_pentest_reports.append(report_copy)
    
    all_pentest_reports.sort(key=lambda x: x["tahun_laporan"])

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in all_pentest_reports])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year]

    latest_report = all_pentest_reports[-1] if all_pentest_reports else {}

    total_findings_overall = sum([r.get("jumlah_temuan", 0) for r in all_pentest_reports])
    total_resolved_overall = sum([r.get("jumlah_temuan_diselesaikan", 0) for r in all_pentest_reports])
    total_critical_overall = sum([r.get("temuan_critical", 0) for r in all_pentest_reports])
    total_high_overall = sum([r.get("temuan_high", 0) for r in all_pentest_reports])
    total_medium_overall = sum([r.get("temuan_medium", 0) for r in all_pentest_reports]) # Added
    total_low_overall = sum([r.get("temuan_low", 0) for r in all_pentest_reports])     # Added
    
    percentage_resolved_overall = (total_resolved_overall / total_findings_overall * 100) if total_findings_overall > 0 else 0

    kpis_pentest = {
        "total_findings": total_findings_overall,
        "total_critical_findings": total_critical_overall,
        "total_high_findings": total_high_overall,
        "total_medium_findings": total_medium_overall, # Added
        "total_low_findings": total_low_overall,     # Added
        "percentage_resolved": percentage_resolved_overall,
        "latest_report_critical": latest_report.get("temuan_critical", 0),
        "latest_report_high": latest_report.get("temuan_high", 0),
        "latest_report_medium": latest_report.get("temuan_medium", 0),
        "latest_report_low": latest_report.get("temuan_low", 0),
        "latest_report_total": latest_report.get("jumlah_temuan", 0),
        "latest_report_resolved": latest_report.get("jumlah_temuan_diselesaikan", 0),
        "latest_report_unresolved": latest_report.get("jumlah_temuan_belum_diselesaikan", 0)
    }

    annual_upload_status_pentest = {}
    today = datetime.now()

    for year_int in all_years_in_data:
        period_key = str(year_int)
        report_entry = pjp_pentest_data_raw.get(str(year_int))

        due_date = datetime(year_int + 1, 3, 31) 

        status = {
            'year': year_int,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            upload_date_str = report_entry.get('created_at', report_entry.get('tanggal_surat'))
            if upload_date_str:
                try:
                    uploaded_dt = datetime.fromisoformat(upload_date_str.split('+')[0]) 
                    status['upload_date'] = uploaded_dt.strftime('%Y-%m-%d')
                    if uploaded_dt.date() > due_date.date(): 
                        status['days_late'] = (uploaded_dt.date() - due_date.date()).days
                    else:
                        status['days_late'] = 0
                except ValueError:
                    print(f"Warning: Could not parse upload_date for pentest {sandi} year {year_int}: {upload_date_str}")
            status['report_data'] = report_entry
        else:
            if today.date() > due_date.date() and year_int < today.year: 
                status['days_late'] = (today.date() - due_date.date()).days
            else:
                status['days_late'] = None
        
        annual_upload_status_pentest[period_key] = status


    return render_template(
        'pentest_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_pentest=all_pentest_reports,
        available_years=all_years_in_data,
        kpis=kpis_pentest,
        annual_upload_status_pentest=annual_upload_status_pentest
    )

# NEW ROUTE: Pentest Report Detail
@app.route('/analisis_pjp/<sandi>/pentest_report_detail/<int:tahun>')
def pentest_report_detail(sandi, tahun):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = PENTEST_REPORTS_DATA.get(sandi, {}).get(str(tahun))
    if report_data is None:
        abort(404)
    
    report_data_copy = report_data.copy() 
    report_data_copy["nama_pjp"] = pjp_info["nama"]
    report_data_copy["sandi_pjp"] = pjp_info["sandi"]

    # Prepare comparison data for previous year
    comparison_data_previous_year = next((r for r in PENTEST_REPORTS_DATA if r.get("sandi_pjp") == sandi and r.get("tahun_laporan") == tahun - 1), None)
    
    # Prepare trend data for 5 years (including current year)
    all_pentest_reports_for_pjp = [
        r.copy() for r in PENTEST_REPORTS_DATA.get(sandi, {}).values()
    ]
    all_pentest_reports_for_pjp.sort(key=lambda x: x["tahun_laporan"])

    trend_data_5_years = []
    current_report_index = -1
    for i, r in enumerate(all_pentest_reports_for_pjp):
        if r["tahun_laporan"] == tahun:
            current_report_index = i
            break
    
    if current_report_index != -1:
        start_index = max(0, current_report_index - 4) 
        trend_data_5_years = all_pentest_reports_for_pjp[start_index : current_report_index + 1]

    return render_template(
        'pentest_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data_copy,
        comparison_data_previous_year=comparison_data_previous_year,
        trend_data_5_years=trend_data_5_years
    )

# NEW ROUTE: Audit SI Report Dashboard
@app.route('/analisis_pjp/<sandi>/auditsi_report')
def auditsi_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    pjp_auditsi_data_raw = AUDIT_SI_REPORTS_DATA.get(sandi, {})

    all_auditsi_reports = []
    for year_str, report_data in pjp_auditsi_data_raw.items():
        report_copy = report_data.copy()
        report_copy["tahun_laporan"] = int(year_str)
        all_auditsi_reports.append(report_copy)
    
    all_auditsi_reports.sort(key=lambda x: x["tahun_laporan"])

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in all_auditsi_reports])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year]

    # Calculate KPIs for Audit SI
    total_reports_count = len(all_auditsi_reports)
    total_findings_overall = sum([r.get("jumlah_temuan", 0) for r in all_auditsi_reports])
    total_resolved_overall = sum([r.get("jumlah_temuan_diselesaikan", 0) for r in all_auditsi_reports])
    
    percentage_resolved_overall = (total_resolved_overall / total_findings_overall * 100) if total_findings_overall > 0 else 0

    # Average scores for 5 pillars
    avg_confidentiality = sum([r.get("confidentiality", 0) for r in all_auditsi_reports]) / total_reports_count if total_reports_count > 0 else 0
    avg_integrity = sum([r.get("integrity", 0) for r in all_auditsi_reports]) / total_reports_count if total_reports_count > 0 else 0
    avg_availability = sum([r.get("availability", 0) for r in all_auditsi_reports]) / total_reports_count if total_reports_count > 0 else 0
    avg_authenticity = sum([r.get("authenticity", 0) for r in all_auditsi_reports]) / total_reports_count if total_reports_count > 0 else 0
    avg_non_repudiation = sum([r.get("non_repudiation", 0) for r in all_auditsi_reports]) / total_reports_count if total_reports_count > 0 else 0

    kpis_auditsi = {
        "total_findings": total_findings_overall,
        "percentage_resolved": percentage_resolved_overall,
        "avg_confidentiality": avg_confidentiality,
        "avg_integrity": avg_integrity,
        "avg_availability": avg_availability,
        "avg_authenticity": avg_authenticity,
        "avg_non_repudiation": avg_non_repudiation
    }

    # Prepare upload status map (annual)
    annual_upload_status_auditsi = {}
    today = datetime.now()

    for year_int in all_years_in_data:
        period_key = str(year_int)
        report_entry = pjp_auditsi_data_raw.get(str(year_int))

        due_date = datetime(year_int + 1, 3, 31) 

        status = {
            'year': year_int,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            upload_date_str = report_entry.get('created_at', report_entry.get('tanggal_surat'))
            if upload_date_str:
                try:
                    uploaded_dt = datetime.fromisoformat(upload_date_str.split('+')[0]) 
                    status['upload_date'] = uploaded_dt.strftime('%Y-%m-%d')
                    if uploaded_dt.date() > due_date.date(): 
                        status['days_late'] = (uploaded_dt.date() - due_date.date()).days
                    else:
                        status['days_late'] = 0
                except ValueError:
                    print(f"Warning: Could not parse upload_date for Audit SI {sandi} year {year_int}: {upload_date_str}")
            status['report_data'] = report_entry
        else:
            if today.date() > due_date.date() and year_int < today.year: 
                status['days_late'] = (today.date() - due_date.date()).days
            else:
                status['days_late'] = None
        
        annual_upload_status_auditsi[period_key] = status

    return render_template(
        'auditsi_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_auditsi=all_auditsi_reports,
        available_years=all_years_in_data,
        kpis=kpis_auditsi,
        annual_upload_status_auditsi=annual_upload_status_auditsi
    )

# NEW ROUTE: Audit SI Report Detail
@app.route('/analisis_pjp/<sandi>/auditsi_report_detail/<int:tahun>')
def auditsi_report_detail(sandi, tahun):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = AUDIT_SI_REPORTS_DATA.get(sandi, {}).get(str(tahun))
    if report_data is None:
        abort(404)
    
    report_data_copy = report_data.copy() 
    report_data_copy["nama_pjp"] = pjp_info["nama"]
    report_data_copy["sandi_pjp"] = pjp_info["sandi"]

    # Prepare comparison data for previous year
    comparison_data_previous_year = next((r for r in AUDIT_SI_REPORTS_DATA if r.get("sandi_pjp") == sandi and r.get("tahun_laporan") == tahun - 1), None)
    
    # Prepare trend data for 5 years (including current year)
    all_auditsi_reports_for_pjp = [
        r.copy() for r in AUDIT_SI_REPORTS_DATA
        if r.get("sandi_pjp") == sandi
    ]
    all_auditsi_reports_for_pjp.sort(key=lambda x: x["tahun_laporan"])

    trend_data_5_years = []
    current_report_index = -1
    for i, r in enumerate(all_auditsi_reports_for_pjp):
        if r["tahun_laporan"] == tahun:
            current_report_index = i
            break
    
    if current_report_index != -1:
        start_index = max(0, current_report_index - 4) 
        trend_data_5_years = all_auditsi_reports_for_pjp[start_index : current_report_index + 1]

    return render_template(
        'auditsi_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data_copy,
        comparison_data_previous_year=comparison_data_previous_year,
        trend_data_5_years=trend_data_5_years
    )

# NEW ROUTE: APUPPT Report Dashboard
@app.route('/analisis_pjp/<sandi>/apuppt_report')
def apuppt_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    # Filter APUPPT data for the specific PJP
    pjp_apuppt_data_raw = [
        report for report in APUPPT_REPORTS_DATA
        if report.get("sandi_pjp") == sandi
    ]

    # Sort reports by year
    pjp_apuppt_data_raw.sort(key=lambda x: x["tahun_laporan"])

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in pjp_apuppt_data_raw])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year] # Ensure there's at least current year if no data

    # Calculate KPIs for APUPPT Dashboard
    total_ltkt = sum([r.get("jumlah_ltkt", 0) for r in pjp_apuppt_data_raw])
    total_ltkm = sum([r.get("jumlah_ltkm", 0) for r in pjp_apuppt_data_raw])
    total_ltkl = sum([r.get("jumlah_ltkl", 0) for r in pjp_apuppt_data_raw])

    total_sipesat_uploaded_count = 0
    total_sipesat_expected_count = 0
    total_dttot_pemblokiran_reported = 0
    total_dttot_pemblokiran_expected = 0
    total_dppspm_pemblokiran_reported = 0
    total_dppspm_pemblokiran_expected = 0

    for report in pjp_apuppt_data_raw:
        sipesat_uploaded_count = 0
        if report.get("lapor_sipesat_tw1"): sipesat_uploaded_count += 1
        if report.get("lapor_sipesat_tw2"): sipesat_uploaded_count += 1
        if report.get("lapor_sipesat_tw3"): sipesat_uploaded_count += 1
        if report.get("lapor_sipesat_tw4"): sipesat_uploaded_count += 1
        
        report["sipesat_uploaded_count"] = sipesat_uploaded_count
        report["sipesat_expected_count"] = 4 # Always 4 quarters
        total_sipesat_uploaded_count += sipesat_uploaded_count
        total_sipesat_expected_count += 4

        total_dttot_pemblokiran_reported += report.get("jumlah_lapor_pemblokiran_dttot", 0)
        total_dttot_pemblokiran_expected += report.get("expected_lapor_pemblokiran_dttot", 0)
        total_dppspm_pemblokiran_reported += report.get("jumlah_lapor_pemblokiran_dppspm", 0)
        total_dppspm_pemblokiran_expected += report.get("expected_lapor_pemblokiran_dppspm", 0)


    sipesat_percentage = (total_sipesat_uploaded_count / total_sipesat_expected_count * 100) if total_sipesat_expected_count > 0 else 0
    dttot_percentage = (total_dttot_pemblokiran_reported / total_dttot_pemblokiran_expected * 100) if total_dttot_pemblokiran_expected > 0 else 0
    dppspm_percentage = (total_dppspm_pemblokiran_reported / total_dppspm_pemblokiran_expected * 100) if total_dppspm_pemblokiran_expected > 0 else 0

    kpis_apuppt = {
        "total_ltkt": total_ltkt,
        "total_ltkm": total_ltkm,
        "total_ltkl": total_ltkl,
        "sipesat_percentage": sipesat_percentage,
        "dttot_percentage": dttot_percentage,
        "dppspm_percentage": dppspm_percentage
    }

    # Prepare annual upload status map
    annual_upload_status_apuppt = {}
    today = datetime.now()

    for year_int in all_years_in_data:
        period_key = str(year_int)
        report_entry = next((r for r in pjp_apuppt_data_raw if r["tahun_laporan"] == year_int), None)

        due_date = datetime(year_int + 1, 3, 31) 

        status = {
            'year': year_int,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            upload_date_str = report_entry.get('created_at', report_entry.get('tanggal_surat'))
            if upload_date_str:
                try:
                    uploaded_dt = datetime.fromisoformat(upload_date_str.split('+')[0]) 
                    status['upload_date'] = uploaded_dt.strftime('%Y-%m-%d')
                    if uploaded_dt.date() > due_date.date(): 
                        status['days_late'] = (uploaded_dt.date() - due_date.date()).days
                    else:
                        status['days_late'] = 0
                except ValueError:
                    print(f"Warning: Could not parse upload_date for APUPPT {sandi} year {year_int}: {upload_date_str}")
            status['report_data'] = report_entry
        else:
            if today.date() > due_date.date() and year_int < today.year: 
                status['days_late'] = (today.date() - due_date.days).days
            else:
                status['days_late'] = None
        
        annual_upload_status_apuppt[period_key] = status

    return render_template(
        'apuppt_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_apuppt=pjp_apuppt_data_raw,
        available_years=all_years_in_data,
        kpis=kpis_apuppt,
        annual_upload_status_apuppt=annual_upload_status_apuppt
    )

# NEW ROUTE: APUPPT Report Detail
@app.route('/analisis_pjp/<sandi>/apuppt_report_detail/<int:tahun>')
def apuppt_report_detail(sandi, tahun):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    # Find the specific report data for the given PJP and year
    report_data = next((r for r in APUPPT_REPORTS_DATA if r.get("sandi_pjp") == sandi and r.get("tahun_laporan") == tahun), None)
    if report_data is None:
        abort(404)
    
    # Add PJP name and sandi to report_data for easier access in template
    report_data_copy = report_data.copy()
    report_data_copy["nama_pjp"] = pjp_info["nama"]
    report_data_copy["sandi_pjp"] = pjp_info["sandi"]

    # Calculate sipesat_uploaded_count for the current report
    sipesat_uploaded_count = 0
    if report_data_copy.get("lapor_sipesat_tw1"): sipesat_uploaded_count += 1
    if report_data_copy.get("lapor_sipesat_tw2"): sipesat_uploaded_count += 1
    if report_data_copy.get("lapor_sipesat_tw3"): sipesat_uploaded_count += 1
    if report_data_copy.get("lapor_sipesat_tw4"): sipesat_uploaded_count += 1
    report_data_copy["sipesat_uploaded_count"] = sipesat_uploaded_count


    # Prepare comparison data for previous year
    comparison_data_previous_year = next((r for r in APUPPT_REPORTS_DATA if r.get("sandi_pjp") == sandi and r.get("tahun_laporan") == tahun - 1), None)
    
    # Add sipesat_uploaded_count for previous year's report if it exists
    if comparison_data_previous_year:
        prev_sipesat_uploaded_count = 0
        if comparison_data_previous_year.get("lapor_sipesat_tw1"): prev_sipesat_uploaded_count += 1
        if comparison_data_previous_year.get("lapor_sipesat_tw2"): prev_sipesat_uploaded_count += 1
        if comparison_data_previous_year.get("lapor_sipesat_tw3"): prev_sipesat_uploaded_count += 1
        if comparison_data_previous_year.get("lapor_sipesat_tw4"): prev_sipesat_uploaded_count += 1
        comparison_data_previous_year["sipesat_uploaded_count"] = prev_sipesat_uploaded_count


    # Prepare trend data for 5 years (including current year)
    all_apuppt_reports_for_pjp = [
        r.copy() for r in APUPPT_REPORTS_DATA
        if r.get("sandi_pjp") == sandi
    ]
    all_apuppt_reports_for_pjp.sort(key=lambda x: x["tahun_laporan"])

    trend_data_5_years = []
    current_report_index = -1
    for i, r in enumerate(all_apuppt_reports_for_pjp):
        if r["tahun_laporan"] == tahun:
            current_report_index = i
            break
    
    if current_report_index != -1:
        start_index = max(0, current_report_index - 4) 
        trend_data_5_years = all_apuppt_reports_for_pjp[start_index : current_report_index + 1]

    return render_template(
        'apuppt_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data_copy,
        comparison_data_previous_year=comparison_data_previous_year,
        trend_data_5_years=trend_data_5_years
    )

# ... (kode app.py lainnya tetap sama)

# NEW ROUTE: Manajemen Report Dashboard
@app.route('/analisis_pjp/<sandi>/manajemen_report')
def manajemen_report(sandi):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    # Filter Management data for the specific PJP
    pjp_manajemen_data_raw = MANAGEMENT_REPORTS_DATA.get(sandi, {})

    all_manajemen_reports = []
    # Iterate through each year for the current PJP
    # Ensure pjp_manajemen_data_raw.items() is safely iterated
    for year_str, report_data in pjp_manajemen_data_raw.items():
        report_copy = report_data.copy() # 'report_data' is already the dictionary for the year
        report_copy["tahun_laporan"] = int(year_str) 
        report_copy["sandi_pjp"] = sandi # Ensure sandi_pjp is added
        report_copy["nama_pjp"] = pjp_info["nama"] # Ensure nama_pjp is added
        all_manajemen_reports.append(report_copy)

    all_manajemen_reports.sort(key=lambda x: x["tahun_laporan"])

    all_years_in_data = sorted(list(set([report["tahun_laporan"] for report in all_manajemen_reports])))
    current_year = datetime.now().year
    if not all_years_in_data:
        all_years_in_data = [current_year] # Ensure there's at least current year if no data

    # Calculate KPIs for Management Report Dashboard
    total_reports_count = len(all_manajemen_reports)
    
    # Example KPIs: Count of reports with specific status as True
    # We'll dynamically count how many reports have each 'status' field as true
    status_fields = [
        "penilaian_perkembangan_industri_sp_status",
        "pengawasan_perkembangan_bisnis_pjp_status",
        "pengawasan_tata_kelola_pjp_status",
        "pandangan_perbaikan_pjp_status",
        "perubahan_komposisi_komisaris_status",
        "implementasi_kebijakan_direksi_status",
        "proses_pencapaian_kinerja_status",
        "perbandingan_target_dan_realisasi_status",
        "kendala_yang_dihadapi_pjp_status",
        "tata_kelola_pengawasan_direksi_komisaris_status",
        "tata_kelola_risiko_sdm_status",
        "tata_kelola_ketersediaan_prosedur_status",
        "tata_kelola_pengendalian_intern_status",
        "asesmen_kualitatif_sdm_status",
        "asesmen_kuantitatif_sdm_status",
        "asesmen_kepemilikan_domestik_asing_status",
        "asesmen_pengendalian_domestik_asing_status",
        "asesmen_pemenuhan_modal_status",
        "asesmen_manajemen_risiko_it_status",
        "pemantauan_kepatuhan_status"
    ]

    kpi_counts = {field: 0 for field in status_fields}
    total_status_fields_checked = 0
    total_true_status_fields = 0

    for report in all_manajemen_reports:
        for field in status_fields:
            if field in report: # Only count if the field exists in the report
                total_status_fields_checked += 1
                if report.get(field) is True:
                    kpi_counts[field] += 1
                    total_true_status_fields += 1
    
    # Calculate average compliance rate across all boolean status fields
    overall_compliance_percentage = (total_true_status_fields / total_status_fields_checked * 100) if total_status_fields_checked > 0 else 0

    # Get latest report for current status of percentages
    latest_report = all_manajemen_reports[-1] if all_manajemen_reports else {}
    
    kpis_manajemen = {
        "total_reports_submitted": total_reports_count,
        "overall_compliance_percentage": overall_compliance_percentage,
        "latest_domestik_ownership": latest_report.get("persentase_kepemilikan_domestik", 0),
        "latest_asing_ownership": latest_report.get("persentase_kepemilikan_asing", 0),
        "latest_domestik_control": latest_report.get("pengendalian_domestik", 0),
        "latest_asing_control": latest_report.get("pengendalian_asing", 0),
        "kpi_status_counts": kpi_counts # Pass counts for individual status fields
    }

    # Prepare annual upload status map (similar to other annual reports)
    annual_upload_status_manajemen = {}
    today = datetime.now()

    for year_int in all_years_in_data:
        period_key = str(year_int)
        report_entry = pjp_manajemen_data_raw.get(str(year_int))

        due_date = datetime(year_int + 1, 3, 31) 

        status = {
            'year': year_int,
            'uploaded': False,
            'upload_date': None,
            'days_late': None,
            'due_date': due_date.strftime('%Y-%m-%d')
        }

        if report_entry:
            status['uploaded'] = True
            upload_date_str = report_entry.get('created_at', report_entry.get('tanggal_surat'))
            if upload_date_str:
                try:
                    uploaded_dt = datetime.fromisoformat(upload_date_str.split('+')[0]) 
                    status['upload_date'] = uploaded_dt.strftime('%Y-%m-%d')
                    if uploaded_dt.date() > due_date.date(): 
                        status['days_late'] = (uploaded_dt.date() - due_date.date()).days
                    else:
                        status['days_late'] = 0
                except ValueError:
                    print(f"Warning: Could not parse upload_date for Management Report {sandi} year {year_int}: {upload_date_str}")
            status['report_data'] = report_entry
        else:
            if today.date() > due_date.date() and year_int < today.year: 
                status['days_late'] = (today.date() - due_date.date()).days
            else:
                status['days_late'] = None
        
        annual_upload_status_manajemen[period_key] = status

    return render_template(
        'manajemen_report.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        initial_dashboard_data_manajemen=all_manajemen_reports,
        available_years=all_years_in_data,
        kpis=kpis_manajemen,
        annual_upload_status_manajemen=annual_upload_status_manajemen
    )


# NEW ROUTE: Manajemen Report Detail
@app.route('/analisis_pjp/<sandi>/manajemen_report_detail/<int:tahun>')
def manajemen_report_detail(sandi, tahun):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == sandi), None)
    if pjp_info is None:
        abort(404)

    report_data = MANAGEMENT_REPORTS_DATA.get(sandi, {}).get(str(tahun))
    if report_data is None:
        abort(404)
    
    report_data_copy = report_data.copy() 
    report_data_copy["nama_pjp"] = pjp_info["nama"]
    report_data_copy["sandi_pjp"] = pjp_info["sandi"]

    # Prepare comparison data for previous year
    comparison_data_previous_year = MANAGEMENT_REPORTS_DATA.get(sandi, {}).get(str(tahun - 1))
    
    # Prepare trend data for 5 years (including current year)
    all_manajemen_reports_for_pjp = [
        r.copy() for r in MANAGEMENT_REPORTS_DATA.get(sandi, {}).values()
    ]
    all_manajemen_reports_for_pjp.sort(key=lambda x: x["tahun_laporan"])

    trend_data_5_years = []
    current_report_index = -1
    for i, r in enumerate(all_manajemen_reports_for_pjp):
        if r["tahun_laporan"] == tahun:
            current_report_index = i
            break
    
    if current_report_index != -1:
        start_index = max(0, current_report_index - 4) 
        trend_data_5_years = all_manajemen_reports_for_pjp[start_index : current_report_index + 1]

    return render_template(
        'manajemen_report_detail.html',
        sandi_pjp=sandi,
        pjp_name=pjp_info["nama"],
        report_data=report_data_copy,
        comparison_data_previous_year=comparison_data_previous_year,
        trend_data_5_years=trend_data_5_years
    )

# NEW ROUTE: Global Fraud Analysis Page
@app.route('/analisis_fraud_global')
def analisis_fraud_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_fraud_reports = []
    # Iterate through each PJP's fraud data
    for pjp_sandi, pjp_years_data in FRAUD_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        for year_str, months_data in pjp_years_data.items():
            # Iterate through each month for the current year
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["Sandi PJP"] = pjp_sandi
                report_copy["Nama PJP"] = pjp_name
                report_copy["Tahun Laporan"] = int(year_str)
                report_copy["Periode Luler"] = month_str # Store original month string for filtering
                report_copy["Periode Laporan Nama"] = get_month_name(month_str) # For display
                all_fraud_reports.append(report_copy)

    # Sort all reports by year and then by month
    all_fraud_reports.sort(key=lambda x: (x["Tahun Laporan"], get_month_number(x["Periode Luler"])))

    # Get unique periods for filtering (e.g., "2023-01", "2023-02", etc.)
    unique_periods = sorted(list(set([f"{r['Tahun Laporan']}-{r['Periode Luler']}" for r in all_fraud_reports])))
    
    # Transform unique_periods for display in dropdown
    display_periods = [
        {"value": period, "text": f"{get_month_name(period.split('-')[1])} {period.split('-')[0]}"}
        for period in unique_periods
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_fraud_global = sum([report["Jumlah Fraud"] for report in all_fraud_reports])
    total_potensi_kerugian_global = sum([report["Besar Potensi Kerugian"] for report in all_fraud_reports])
    
    avg_fraud_per_report_global = total_fraud_global / len(all_fraud_reports) if all_fraud_reports else 0
    avg_loss_per_report_global = total_potensi_kerugian_global / len(all_fraud_reports) if all_fraud_reports else 0

    kpis_global = {
        "total_fraud_global": total_fraud_global,
        "total_potensi_kerugian_global": total_potensi_kerugian_global,
        "avg_fraud_per_report_global": avg_fraud_per_report_global,
        "avg_loss_per_report_global": avg_loss_per_report_global
    }

    # Prepare data for trend charts (monthly and yearly for all data)
    monthly_fraud_trends = {} # { "YYYY-MM": { "jumlah_fraud": N, "potensi_kerugian": M } }
    yearly_fraud_trends = {}  # { "YYYY": { "jumlah_fraud": N, "potensi_kerugian": M } }

    for report in all_fraud_reports:
        year = report["Tahun Laporan"]
        month = report["Periode Luler"] # e.g., "01", "02"
        jumlah_fraud = report["Jumlah Fraud"]
        potensi_kerugian = report["Besar Potensi Kerugian"]

        # Monthly trends
        month_key = f"{year}-{month}"
        if month_key not in monthly_fraud_trends:
            monthly_fraud_trends[month_key] = {"jumlah_fraud": 0, "potensi_kerugian": 0}
        monthly_fraud_trends[month_key]["jumlah_fraud"] += jumlah_fraud
        monthly_fraud_trends[month_key]["potensi_kerugian"] += potensi_kerugian

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_fraud_trends:
            yearly_fraud_trends[year_key] = {"jumlah_fraud": 0, "potensi_kerugian": 0}
        yearly_fraud_trends[year_key]["jumlah_fraud"] += jumlah_fraud
        yearly_fraud_trends[year_key]["potensi_kerugian"] += potensi_kerugian
    
    # Sort monthly trends for chart labels
    sorted_monthly_keys = sorted(monthly_fraud_trends.keys())
    monthly_trend_labels = [f"{get_month_name(k.split('-')[1])} {k.split('-')[0]}" for k in sorted_monthly_keys]
    monthly_trend_jumlah_fraud = [monthly_fraud_trends[k]["jumlah_fraud"] for k in sorted_monthly_keys]
    monthly_trend_potensi_kerugian = [monthly_fraud_trends[k]["potensi_kerugian"] for k in sorted_monthly_keys]

    # Sort yearly trends for chart labels
    sorted_yearly_keys = sorted(yearly_fraud_trends.keys())
    yearly_trend_labels = [k for k in sorted_yearly_keys]
    yearly_trend_jumlah_fraud = [yearly_fraud_trends[k]["jumlah_fraud"] for k in sorted_yearly_keys]
    yearly_trend_potensi_kerugian = [yearly_fraud_trends[k]["potensi_kerugian"] for k in sorted_yearly_keys]

    trend_chart_data = {
        "monthly": {
            "labels": monthly_trend_labels,
            "jumlah_fraud": monthly_trend_jumlah_fraud,
            "potensi_kerugian": monthly_trend_potensi_kerugian
        },
        "yearly": {
            "labels": yearly_trend_labels,
            "jumlah_fraud": yearly_trend_jumlah_fraud,
            "potensi_kerugian": yearly_trend_potensi_kerugian
        }
    }

    # Prepare data for top/bottom fraud types and PJPs (for initial load, based on all data)
    initial_top_bottom_data = get_top_bottom_analysis_fraud(all_fraud_reports)


    return render_template(
        'analisis_fraud_global.html',
        all_fraud_reports=all_fraud_reports, # Pass the Python list directly
        unique_periods=display_periods,
        kpis_global=kpis_global,
        trend_chart_data=trend_chart_data, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data # Pass initial top/bottom data
    )

# Helper function to get top/bottom fraud types and PJPs for Fraud
def get_top_bottom_analysis_fraud(reports):
    pjp_aggregated = {}        # { "PJP_Name": { "total_jumlah_fraud": N, "total_potensi_kerugian": M } }

    for report in reports:
        pjp_name = report.get("Nama PJP", "Tidak Diketahui")
        jumlah_fraud = report.get("Jumlah Fraud", 0)
        potensi_kerugian = report.get("Besar Potensi Kerugian", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_jumlah_fraud": 0, "total_potensi_kerugian": 0}
        pjp_aggregated[pjp_name]["total_jumlah_fraud"] += jumlah_fraud
        pjp_aggregated[pjp_name]["total_potensi_kerugian"] += potensi_kerugian

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by count
    top_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_jumlah_fraud"], reverse=True)[:5]
    bottom_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_jumlah_fraud"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by loss
    top_5_pjps_by_loss = sorted(pjps_list, key=lambda item: item[1]["total_potensi_kerugian"], reverse=True)[:5]
    bottom_5_pjps_by_loss = sorted(pjps_list, key=lambda item: item[1]["total_potensi_kerugian"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_count = list(reports) # Ensure it's a copy
    reports_by_loss = list(reports) # Ensure it's a copy

    # Sort by Jumlah Fraud
    top_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("Jumlah Fraud", 0), reverse=True)[:5]
    bottom_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("Jumlah Fraud", 0))[:5]

    # Sort by Potensi Kerugian
    top_5_reports_by_loss = sorted(reports_by_loss, key=lambda a: a.get("Besar Potensi Kerugian", 0), reverse=True)[:5]
    bottom_5_reports_by_loss = sorted(reports_by_loss, key=lambda a: a.get("Besar Potensi Kerugian", 0))[:5]

    return {
        "top_5_reports_by_count": top_5_reports_by_count,
        "bottom_5_reports_by_count": bottom_5_reports_by_count,
        "top_5_reports_by_loss": top_5_reports_by_loss,
        "bottom_5_reports_by_loss": bottom_5_reports_by_loss,
        "top_5_pjps_by_count": top_5_pjps_by_count,
        "bottom_5_pjps_by_count": bottom_5_pjps_by_count,
        "top_5_pjps_by_loss": top_5_pjps_by_loss,
        "bottom_5_pjps_by_loss": bottom_5_pjps_by_loss,
    }


# NEW ROUTE: Global DTTOT Analysis Page
@app.route('/analisis_dttot_global')
def analisis_dttot_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_dttot_reports = []
    # Iterate through each PJP's DTTOT data
    for pjp_sandi, pjp_years_data in DTTOT_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        for year_str, months_data in pjp_years_data.items():
            # Iterate through each month for the current year
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                report_copy["periode_luler"] = month_str # Store original month string for filtering
                report_copy["periode_laporan_nama"] = get_month_name(month_str) # For display
                all_dttot_reports.append(report_copy)

    # Sort all reports by year and then by month
    all_dttot_reports.sort(key=lambda x: (x["tahun_laporan"], get_month_number(x["periode_luler"])))

    # Get unique periods for filtering (e.g., "2023-01", "2023-02", etc.)
    unique_periods = sorted(list(set([f"{r['tahun_laporan']}-{r['periode_luler']}" for r in all_dttot_reports])))
    
    # Transform unique_periods for display in dropdown
    display_periods = [
        {"value": period, "text": f"{get_month_name(period.split('-')[1])} {period.split('-')[0]}"}
        for period in unique_periods
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_dttot_global = sum([report["jumlah_terduga_teroris"] for report in all_dttot_reports])
    
    org_counts_global = {}
    for report in all_dttot_reports:
        org = report.get("organisasi_teroris", "Tidak Diketahui")
        org_counts_global[org] = org_counts_global.get(org, 0) + 1
    most_common_org_global = max(org_counts_global, key=org_counts_global.get) if org_counts_global else "N/A"

    kpis_global_dttot = {
        "total_dttot_global": total_dttot_global,
        "most_common_org_global": most_common_org_global,
    }

    # Prepare data for trend charts (monthly and yearly for all data)
    monthly_dttot_trends = {} # { "YYYY-MM": { "jumlah_terduga_teroris": N } }
    yearly_dttot_trends = {}  # { "YYYY": { "jumlah_terduga_teroris": N } }

    for report in all_dttot_reports:
        year = report["tahun_laporan"]
        month = report["periode_luler"] # e.g., "01", "02"
        jumlah_terduga_teroris = report["jumlah_terduga_teroris"]

        # Monthly trends
        month_key = f"{year}-{month}"
        if month_key not in monthly_dttot_trends:
            monthly_dttot_trends[month_key] = {"jumlah_terduga_teroris": 0}
        monthly_dttot_trends[month_key]["jumlah_terduga_teroris"] += jumlah_terduga_teroris

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_dttot_trends:
            yearly_dttot_trends[year_key] = {"jumlah_terduga_teroris": 0}
        yearly_dttot_trends[year_key]["jumlah_terduga_teroris"] += jumlah_terduga_teroris
    
    # Sort monthly trends for chart labels
    sorted_monthly_keys_dttot = sorted(monthly_dttot_trends.keys())
    monthly_trend_labels_dttot = [f"{get_month_name(k.split('-')[1])} {k.split('-')[0]}" for k in sorted_monthly_keys_dttot]
    monthly_trend_jumlah_dttot = [monthly_dttot_trends[k]["jumlah_terduga_teroris"] for k in sorted_monthly_keys_dttot]

    # Sort yearly trends for chart labels
    sorted_yearly_keys_dttot = sorted(yearly_dttot_trends.keys())
    yearly_trend_labels_dttot = [k for k in sorted_yearly_keys_dttot]
    yearly_trend_jumlah_dttot = [yearly_dttot_trends[k]["jumlah_terduga_teroris"] for k in sorted_yearly_keys_dttot]

    trend_chart_data_dttot = {
        "monthly": {
            "labels": monthly_trend_labels_dttot,
            "jumlah_terduga_teroris": monthly_trend_jumlah_dttot,
        },
        "yearly": {
            "labels": yearly_trend_labels_dttot,
            "jumlah_terduga_teroris": yearly_trend_jumlah_dttot,
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_dttot = get_top_bottom_analysis_dttot(all_dttot_reports)


    return render_template(
        'analisis_dttot_global.html',
        all_dttot_reports=all_dttot_reports, # Pass the Python list directly
        unique_periods=display_periods,
        kpis_global=kpis_global_dttot,
        trend_chart_data=trend_chart_data_dttot, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_dttot # Pass initial top/bottom data
    )

# Helper function to get top/bottom reports and PJPs for DTTOT
def get_top_bottom_analysis_dttot(reports):
    pjp_aggregated = {}        # { "PJP_Name": { "total_jumlah_dttot": N } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        jumlah_terduga_teroris = report.get("jumlah_terduga_teroris", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_jumlah_dttot": 0}
        pjp_aggregated[pjp_name]["total_jumlah_dttot"] += jumlah_terduga_teroris

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by count
    top_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_jumlah_dttot"], reverse=True)[:5]
    bottom_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_jumlah_dttot"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_count = list(reports) # Ensure it's a copy

    # Sort by Jumlah Terduga Teroris
    top_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("jumlah_terduga_teroris", 0), reverse=True)[:5]
    bottom_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("jumlah_terduga_teroris", 0))[:5]

    return {
        "top_5_reports_by_count": top_5_reports_by_count,
        "bottom_5_reports_by_count": bottom_5_reports_by_count,
        "top_5_pjps_by_count": top_5_pjps_by_count,
        "bottom_5_pjps_by_count": bottom_5_pjps_by_count,
    }


# NEW ROUTE: Export Fraud Data to XLSX
@app.route('/export_fraud_xlsx')
def export_fraud_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_month = request.args.get('startMonth')
    start_year = request.args.get('startYear')
    end_month = request.args.get('endMonth')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_fraud_global route
    all_fraud_reports_raw = []
    for pjp_sandi, pjp_years_data in FRAUD_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        for year_str, months_data in pjp_years_data.items():
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["Sandi PJP"] = pjp_sandi
                report_copy["Nama PJP"] = pjp_name
                report_copy["Tahun Laporan"] = int(year_str)
                report_copy["Periode Luler"] = month_str # Use original month string for filtering
                report_copy["Periode Laporan Nama"] = get_month_name(month_str) # For display
                all_fraud_reports_raw.append(report_copy)
    
    if start_year and end_year and start_month and end_month:
        start_period_value = int(start_year + start_month)
        end_period_value = int(end_year + end_month)
        filtered_reports = [
            report for report in all_fraud_reports_raw
            if int(str(report["Tahun Laporan"]) + str(report["Periode Luler"])) >= start_period_value and \
               int(str(report["Tahun Laporan"]) + str(report["Periode Luler"])) <= end_period_value
        ]
    else:
        filtered_reports = all_fraud_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Fraud Global"

    # Define headers
    headers = [
        "Tahun", "Periode", "Nama PJP", "Nomor Surat", "Jumlah Fraud",
        "Besar Potensi Kerugian", "Keterangan Fraud", "Keterangan Tindak Lanjut", "Tanggal Dibuat"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            report.get("Tahun Laporan"),
            report.get("Periode Laporan Nama"),
            report.get("Nama PJP"),
            report.get("Nomor Surat"),
            report.get("Jumlah Fraud"),
            report.get("Besar Potensi Kerugian"),
            report.get("Keterangan Fraud"),
            report.get("Keterangan Tindak Lanjut"),
            report.get("Created at")
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter # Get the column name (e.g., 'A', 'B')
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Fraud_Global"]
    if start_year and start_month and end_year and end_month:
        start_month_name = get_month_name(start_month)
        end_month_name = get_month_name(end_month)
        filename_parts.append(f"{start_month_name}_{start_year}_to_{end_month_name}_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# NEW ROUTE: Export DTTOT Data to XLSX
@app.route('/export_dttot_xlsx')
def export_dttot_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_month = request.args.get('startMonth')
    start_year = request.args.get('startYear')
    end_month = request.args.get('endMonth')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_dttot_global route
    all_dttot_reports_raw = []
    for pjp_sandi, pjp_years_data in DTTOT_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        for year_str, months_data in pjp_years_data.items():
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                report_copy["periode_luler"] = month_str # Use original month string for filtering
                report_copy["periode_laporan_nama"] = get_month_name(month_str) # For display
                all_dttot_reports_raw.append(report_copy)
    
    if start_year and end_year and start_month and end_month:
        start_period_value = int(start_year + start_month)
        end_period_value = int(end_year + end_month)
        filtered_reports = [
            report for report in all_dttot_reports_raw
            if int(str(report["tahun_laporan"]) + str(report["periode_luler"])) >= start_period_value and \
               int(str(report["tahun_laporan"]) + str(report["periode_luler"])) <= end_period_value
        ]
    else:
        filtered_reports = all_dttot_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan DTTOT Global"

    # Define headers
    headers = [
        "Tahun", "Periode", "Nama PJP", "Sandi PJP", "Nomor Surat Kepolisian", "Tanggal Surat",
        "Jumlah Terduga Teroris", "Organisasi Teroris", "Keterangan", "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            report.get("tahun_laporan"),
            report.get("periode_laporan_nama"),
            report.get("nama_pjp"),
            report.get("sandi_pjp"),
            report.get("nomor_surat_kepolisian"),
            report.get("tanggal_surat"),
            report.get("jumlah_terduga_teroris"),
            report.get("organisasi_teroris"),
            report.get("keterangan"),
            report.get("created_at")
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter # Get the column name (e.g., 'A', 'B')
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_DTTOT_Global"]
    if start_year and start_month and end_year and end_month:
        start_month_name = get_month_name(start_month)
        end_month_name = get_month_name(end_month)
        filename_parts.append(f"{start_month_name}_{start_year}_to_{end_month_name}_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# app.py additions (only new routes and helper functions)

# Helper function to get top/bottom reports and PJPs for LTDBB
def get_top_bottom_analysis_ltdbb(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_outgoing_num": N, "total_outgoing_amt": M, ... } }

    for report in reports:
        pjp_name = report.get("Nama PJP", "Tidak Diketahui")
        num_outgoing = report.get("Number Outgoing Transactions", 0)
        amt_outgoing = report.get("Amount Outgoing Transactions", 0)
        num_incoming = report.get("Number Incoming Transactions", 0)
        amt_incoming = report.get("Amount Incoming Transactions", 0)
        num_domestic = report.get("Number Domestic Transactions", 0)
        amt_domestic = report.get("Amount Domestic Transactions", 0)

        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {
                "total_outgoing_num": 0,
                "total_outgoing_amt": 0,
                "total_incoming_num": 0,
                "total_incoming_amt": 0,
                "total_domestic_num": 0,
                "total_domestic_amt": 0
            }
        pjp_aggregated[pjp_name]["total_outgoing_num"] += num_outgoing
        pjp_aggregated[pjp_name]["total_outgoing_amt"] += amt_outgoing
        pjp_aggregated[pjp_name]["total_incoming_num"] += num_incoming
        pjp_aggregated[pjp_name]["total_incoming_amt"] += amt_incoming
        pjp_aggregated[pjp_name]["total_domestic_num"] += num_domestic
        pjp_aggregated[pjp_name]["total_domestic_amt"] += amt_domestic

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by outgoing number
    top_5_pjps_by_outgoing_num = sorted(pjps_list, key=lambda item: item[1]["total_outgoing_num"], reverse=True)[:5]
    bottom_5_pjps_by_outgoing_num = sorted(pjps_list, key=lambda item: item[1]["total_outgoing_num"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by outgoing amount
    top_5_pjps_by_outgoing_amt = sorted(pjps_list, key=lambda item: item[1]["total_outgoing_amt"], reverse=True)[:5]
    bottom_5_pjps_by_outgoing_amt = sorted(pjps_list, key=lambda item: item[1]["total_outgoing_amt"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by incoming number
    top_5_pjps_by_incoming_num = sorted(pjps_list, key=lambda item: item[1]["total_incoming_num"], reverse=True)[:5]
    bottom_5_pjps_by_incoming_num = sorted(pjps_list, key=lambda item: item[1]["total_incoming_num"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by incoming amount
    top_5_pjps_by_incoming_amt = sorted(pjps_list, key=lambda item: item[1]["total_incoming_amt"], reverse=True)[:5]
    bottom_5_pjps_by_incoming_amt = sorted(pjps_list, key=lambda item: item[1]["total_incoming_amt"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by domestic number
    top_5_pjps_by_domestic_num = sorted(pjps_list, key=lambda item: item[1]["total_domestic_num"], reverse=True)[:5]
    bottom_5_pjps_by_domestic_num = sorted(pjps_list, key=lambda item: item[1]["total_domestic_num"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by domestic amount
    top_5_pjps_by_domestic_amt = sorted(pjps_list, key=lambda item: item[1]["total_domestic_amt"], reverse=True)[:5]
    bottom_5_pjps_by_domestic_amt = sorted(pjps_list, key=lambda item: item[1]["total_domestic_amt"])[:5]


    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_outgoing_num = list(reports)
    reports_by_outgoing_amt = list(reports)
    reports_by_incoming_num = list(reports)
    reports_by_incoming_amt = list(reports)
    reports_by_domestic_num = list(reports)
    reports_by_domestic_amt = list(reports)


    # Sort by Number Outgoing Transactions
    top_5_reports_by_outgoing_num = sorted(reports_by_outgoing_num, key=lambda a: a.get("Number Outgoing Transactions", 0), reverse=True)[:5]
    bottom_5_reports_by_outgoing_num = sorted(reports_by_outgoing_num, key=lambda a: a.get("Number Outgoing Transactions", 0))[:5]

    # Sort by Amount Outgoing Transactions
    top_5_reports_by_outgoing_amt = sorted(reports_by_outgoing_amt, key=lambda a: a.get("Amount Outgoing Transactions", 0), reverse=True)[:5]
    bottom_5_reports_by_outgoing_amt = sorted(reports_by_outgoing_amt, key=lambda a: a.get("Amount Outgoing Transactions", 0))[:5]

    # Sort by Number Incoming Transactions
    top_5_reports_by_incoming_num = sorted(reports_by_incoming_num, key=lambda a: a.get("Number Incoming Transactions", 0), reverse=True)[:5]
    bottom_5_reports_by_incoming_num = sorted(reports_by_incoming_num, key=lambda a: a.get("Number Incoming Transactions", 0))[:5]

    # Sort by Amount Incoming Transactions
    top_5_reports_by_incoming_amt = sorted(reports_by_incoming_amt, key=lambda a: a.get("Amount Incoming Transactions", 0), reverse=True)[:5]
    bottom_5_reports_by_incoming_amt = sorted(reports_by_incoming_amt, key=lambda a: a.get("Amount Incoming Transactions", 0))[:5]

    # Sort by Number Domestic Transactions
    top_5_reports_by_domestic_num = sorted(reports_by_domestic_num, key=lambda a: a.get("Number Domestic Transactions", 0), reverse=True)[:5]
    bottom_5_reports_by_domestic_num = sorted(reports_by_domestic_num, key=lambda a: a.get("Number Domestic Transactions", 0))[:5]

    # Sort by Amount Domestic Transactions
    top_5_reports_by_domestic_amt = sorted(reports_by_domestic_amt, key=lambda a: a.get("Amount Domestic Transactions", 0), reverse=True)[:5]
    bottom_5_reports_by_domestic_amt = sorted(reports_by_domestic_amt, key=lambda a: a.get("Amount Domestic Transactions", 0))[:5]


    return {
        "top_5_reports_by_outgoing_num": top_5_reports_by_outgoing_num,
        "bottom_5_reports_by_outgoing_num": bottom_5_reports_by_outgoing_num,
        "top_5_reports_by_outgoing_amt": top_5_reports_by_outgoing_amt,
        "bottom_5_reports_by_outgoing_amt": bottom_5_reports_by_outgoing_amt,
        "top_5_reports_by_incoming_num": top_5_reports_by_incoming_num,
        "bottom_5_reports_by_incoming_num": bottom_5_reports_by_incoming_num,
        "top_5_reports_by_incoming_amt": top_5_reports_by_incoming_amt,
        "bottom_5_reports_by_incoming_amt": bottom_5_reports_by_incoming_amt,
        "top_5_reports_by_domestic_num": top_5_reports_by_domestic_num,
        "bottom_5_reports_by_domestic_num": bottom_5_reports_by_domestic_num,
        "top_5_reports_by_domestic_amt": top_5_reports_by_domestic_amt,
        "bottom_5_reports_by_domestic_amt": bottom_5_reports_by_domestic_amt,
        "top_5_pjps_by_outgoing_num": top_5_pjps_by_outgoing_num,
        "bottom_5_pjps_by_outgoing_num": bottom_5_pjps_by_outgoing_num,
        "top_5_pjps_by_outgoing_amt": top_5_pjps_by_outgoing_amt,
        "bottom_5_pjps_by_outgoing_amt": bottom_5_pjps_by_outgoing_amt,
        "top_5_pjps_by_incoming_num": top_5_pjps_by_incoming_num,
        "bottom_5_pjps_by_incoming_num": bottom_5_pjps_by_incoming_num,
        "top_5_pjps_by_incoming_amt": top_5_pjps_by_incoming_amt,
        "bottom_5_pjps_by_incoming_amt": bottom_5_pjps_by_incoming_amt,
        "top_5_pjps_by_domestic_num": top_5_pjps_by_domestic_num,
        "bottom_5_pjps_by_domestic_num": bottom_5_pjps_by_domestic_num,
        "top_5_pjps_by_domestic_amt": top_5_pjps_by_domestic_amt,
        "bottom_5_pjps_by_domestic_amt": bottom_5_pjps_by_domestic_amt,
    }

# NEW ROUTE: Global LTDBB Analysis Page
@app.route('/analisis_ltdbb_global')
def analisis_ltdbb_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_ltdbb_reports = []
    # Iterate through each PJP's LTDBB data
    for pjp_sandi, pjp_years_data in LTDBB_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        for year_str, months_data in pjp_years_data.items():
            # Iterate through each month for the current year
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["Sandi PJP"] = pjp_sandi
                report_copy["Nama PJP"] = pjp_name
                report_copy["Tahun Laporan"] = int(year_str)
                report_copy["Periode Luler"] = month_str # Store original month string for filtering
                report_copy["Periode Laporan Nama"] = get_month_name(month_str) # For display
                all_ltdbb_reports.append(report_copy)

    # Sort all reports by year and then by month
    all_ltdbb_reports.sort(key=lambda x: (x["Tahun Laporan"], get_month_number(x["Periode Luler"])))

    # Get unique periods for filtering (e.g., "2023-01", "2023-02", etc.)
    unique_periods = sorted(list(set([f"{r['Tahun Laporan']}-{r['Periode Luler']}" for r in all_ltdbb_reports])))
    
    # Transform unique_periods for display in dropdown
    display_periods = [
        {"value": period, "text": f"{get_month_name(period.split('-')[1])} {period.split('-')[0]}"}
        for period in unique_periods
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_outgoing_num_global = sum([report.get("Number Outgoing Transactions", 0) for report in all_ltdbb_reports])
    total_outgoing_amt_global = sum([report.get("Amount Outgoing Transactions", 0) for report in all_ltdbb_reports])
    total_incoming_num_global = sum([report.get("Number Incoming Transactions", 0) for report in all_ltdbb_reports])
    total_incoming_amt_global = sum([report.get("Amount Incoming Transactions", 0) for report in all_ltdbb_reports])
    total_domestic_num_global = sum([report.get("Number Domestic Transactions", 0) for report in all_ltdbb_reports])
    total_domestic_amt_global = sum([report.get("Amount Domestic Transactions", 0) for report in all_ltdbb_reports])
    
    avg_outgoing_num_per_report_global = total_outgoing_num_global / len(all_ltdbb_reports) if all_ltdbb_reports else 0
    avg_outgoing_amt_per_report_global = total_outgoing_amt_global / len(all_ltdbb_reports) if all_ltdbb_reports else 0

    kpis_global_ltdbb = {
        "total_outgoing_num_global": total_outgoing_num_global,
        "total_outgoing_amt_global": total_outgoing_amt_global,
        "total_incoming_num_global": total_incoming_num_global,
        "total_incoming_amt_global": total_incoming_amt_global,
        "total_domestic_num_global": total_domestic_num_global,
        "total_domestic_amt_global": total_domestic_amt_global,
        "avg_outgoing_num_per_report_global": avg_outgoing_num_per_report_global,
        "avg_outgoing_amt_per_report_global": avg_outgoing_amt_per_report_global,
    }

    # Prepare data for trend charts (monthly and yearly for all data)
    monthly_ltdbb_trends = {} # { "YYYY-MM": { "num_outgoing": N, "amt_outgoing": M, "num_incoming": N, "amt_incoming": M, "num_domestic": N, "amt_domestic": M } }
    yearly_ltdbb_trends = {}  # { "YYYY": { "num_outgoing": N, "amt_outgoing": M, "num_incoming": N, "amt_incoming": M, "num_domestic": N, "amt_domestic": M } }

    for report in all_ltdbb_reports:
        year = report["Tahun Laporan"]
        month = report["Periode Luler"] # e.g., "01", "02"
        num_outgoing = report.get("Number Outgoing Transactions", 0)
        amt_outgoing = report.get("Amount Outgoing Transactions", 0)
        num_incoming = report.get("Number Incoming Transactions", 0)
        amt_incoming = report.get("Amount Incoming Transactions", 0)
        num_domestic = report.get("Number Domestic Transactions", 0)
        amt_domestic = report.get("Amount Domestic Transactions", 0)

        # Monthly trends
        month_key = f"{year}-{month}"
        if month_key not in monthly_ltdbb_trends:
            monthly_ltdbb_trends[month_key] = {"num_outgoing": 0, "amt_outgoing": 0, "num_incoming": 0, "amt_incoming": 0, "num_domestic": 0, "amt_domestic": 0} # Added new keys
        monthly_ltdbb_trends[month_key]["num_outgoing"] += num_outgoing
        monthly_ltdbb_trends[month_key]["amt_outgoing"] += amt_outgoing
        monthly_ltdbb_trends[month_key]["num_incoming"] += num_incoming # Added
        monthly_ltdbb_trends[month_key]["amt_incoming"] += amt_incoming # Added
        monthly_ltdbb_trends[month_key]["num_domestic"] += num_domestic # Added
        monthly_ltdbb_trends[month_key]["amt_domestic"] += amt_domestic # Added

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_ltdbb_trends:
            yearly_ltdbb_trends[year_key] = {"num_outgoing": 0, "amt_outgoing": 0, "num_incoming": 0, "amt_incoming": 0, "num_domestic": 0, "amt_domestic": 0} # Added new keys
        yearly_ltdbb_trends[year_key]["num_outgoing"] += num_outgoing
        yearly_ltdbb_trends[year_key]["amt_outgoing"] += amt_outgoing
        yearly_ltdbb_trends[year_key]["num_incoming"] += num_incoming # Added
        yearly_ltdbb_trends[year_key]["amt_incoming"] += amt_incoming # Added
        yearly_ltdbb_trends[year_key]["num_domestic"] += num_domestic # Added
        yearly_ltdbb_trends[year_key]["amt_domestic"] += amt_domestic # Added
    
    # Sort monthly trends for chart labels
    sorted_monthly_keys_ltdbb = sorted(monthly_ltdbb_trends.keys())
    monthly_trend_labels_ltdbb = [f"{get_month_name(k.split('-')[1])} {k.split('-')[0]}" for k in sorted_monthly_keys_ltdbb]
    monthly_trend_num_outgoing = [monthly_ltdbb_trends[k]["num_outgoing"] for k in sorted_monthly_keys_ltdbb]
    monthly_trend_amt_outgoing = [monthly_ltdbb_trends[k]["amt_outgoing"] for k in sorted_monthly_keys_ltdbb]
    monthly_trend_num_incoming = [monthly_ltdbb_trends[k]["num_incoming"] for k in sorted_monthly_keys_ltdbb] # Added
    monthly_trend_amt_incoming = [monthly_ltdbb_trends[k]["amt_incoming"] for k in sorted_monthly_keys_ltdbb] # Added
    monthly_trend_num_domestic = [monthly_ltdbb_trends[k]["num_domestic"] for k in sorted_monthly_keys_ltdbb] # Added
    monthly_trend_amt_domestic = [monthly_ltdbb_trends[k]["amt_domestic"] for k in sorted_monthly_keys_ltdbb] # Added

    # Sort yearly trends for chart labels
    sorted_yearly_keys_ltdbb = sorted(yearly_ltdbb_trends.keys())
    yearly_trend_labels_ltdbb = [k for k in sorted_yearly_keys_ltdbb]
    yearly_trend_num_outgoing = [yearly_ltdbb_trends[k]["num_outgoing"] for k in sorted_yearly_keys_ltdbb]
    yearly_trend_amt_outgoing = [yearly_ltdbb_trends[k]["amt_outgoing"] for k in sorted_yearly_keys_ltdbb]
    yearly_trend_num_incoming = [yearly_ltdbb_trends[k]["num_incoming"] for k in sorted_yearly_keys_ltdbb] # Added
    yearly_trend_amt_incoming = [yearly_ltdbb_trends[k]["amt_incoming"] for k in sorted_yearly_keys_ltdbb] # Added
    yearly_trend_num_domestic = [yearly_ltdbb_trends[k]["num_domestic"] for k in sorted_yearly_keys_ltdbb] # Added
    yearly_trend_amt_domestic = [yearly_ltdbb_trends[k]["amt_domestic"] for k in sorted_yearly_keys_ltdbb] # Added


    trend_chart_data_ltdbb = {
        "monthly": {
            "labels": monthly_trend_labels_ltdbb,
            "num_outgoing": monthly_trend_num_outgoing,
            "amt_outgoing": monthly_trend_amt_outgoing,
            "num_incoming": monthly_trend_num_incoming, # Added
            "amt_incoming": monthly_trend_amt_incoming, # Added
            "num_domestic": monthly_trend_num_domestic, # Added
            "amt_domestic": monthly_trend_amt_domestic # Added
        },
        "yearly": {
            "labels": yearly_trend_labels_ltdbb,
            "num_outgoing": yearly_trend_num_outgoing,
            "amt_outgoing": yearly_trend_amt_outgoing,
            "num_incoming": yearly_trend_num_incoming, # Added
            "amt_incoming": yearly_trend_amt_incoming, # Added
            "num_domestic": yearly_trend_num_domestic, # Added
            "amt_domestic": yearly_trend_amt_domestic # Added
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_ltdbb = get_top_bottom_analysis_ltdbb(all_ltdbb_reports)


    return render_template(
        'analisis_ltdbb_global.html',
        all_ltdbb_reports=all_ltdbb_reports, # Pass the Python list directly
        unique_periods=display_periods,
        kpis_global=kpis_global_ltdbb,
        trend_chart_data=trend_chart_data_ltdbb, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_ltdbb # Pass initial top/bottom data
    )

# NEW ROUTE: Export LTDBB Data to XLSX
@app.route('/export_ltdbb_xlsx')
def export_ltdbb_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_month = request.args.get('startMonth')
    start_year = request.args.get('startYear')
    end_month = request.args.get('endMonth')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_ltdbb_global route
    all_ltdbb_reports_raw = []
    for pjp_sandi, pjp_years_data in LTDBB_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        for year_str, months_data in pjp_years_data.items():
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["Sandi PJP"] = pjp_sandi
                report_copy["Nama PJP"] = pjp_name
                report_copy["Tahun Laporan"] = int(year_str)
                report_copy["Periode Luler"] = month_str # Use original month string for filtering
                report_copy["Periode Laporan Nama"] = get_month_name(month_str) # For display
                all_ltdbb_reports_raw.append(report_copy)
    
    if start_year and end_year and start_month and end_month:
        start_period_value = int(start_year + start_month)
        end_period_value = int(end_year + end_month)
        filtered_reports = [
            report for report in all_ltdbb_reports_raw
            if int(str(report["Tahun Laporan"]) + str(report["Periode Luler"])) >= start_period_value and \
               int(str(report["Tahun Laporan"]) + str(report["Periode Luler"])) <= end_period_value
        ]
    else:
        filtered_reports = all_ltdbb_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan LTDBB Global"

    # Define headers
    headers = [
        "Tahun", "Periode", "Nama PJP", "Sandi PJP",
        "Number Outgoing Transactions", "Amount Outgoing Transactions",
        "Number Incoming Transactions", "Amount Incoming Transactions",
        "Number Domestic Transactions", "Amount Domestic Transactions",
        "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            report.get("Tahun Laporan"),
            report.get("Periode Laporan Nama"),
            report.get("Nama PJP"),
            report.get("Sandi PJP"),
            report.get("Number Outgoing Transactions"),
            report.get("Amount Outgoing Transactions"),
            report.get("Number Incoming Transactions"),
            report.get("Amount Incoming Transactions"),
            report.get("Number Domestic Transactions"),
            report.get("Amount Domestic Transactions"),
            report.get("Created at")
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter # Get the column name (e.g., 'A', 'B')
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_LTDBB_Global"]
    if start_year and start_month and end_year and end_month:
        start_month_name = get_month_name(start_month)
        end_month_name = get_month_name(end_month)
        filename_parts.append(f"{start_month_name}_{start_year}_to_{end_month_name}_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# app.py additions (only new routes and helper functions)

# ... (kode app.py lainnya tetap sama)

# Helper function to get top/bottom reports and PJPs for P2P Cooperation
def get_top_bottom_analysis_p2p(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_cooperation": N } }

    for report in reports:
        pjp_name = report.get("Nama PJP", "Tidak Diketahui")
        jumlah_kerjasama = report.get("Jumlah Perusahaan Kerjasama P2P", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_cooperation": 0}
        pjp_aggregated[pjp_name]["total_cooperation"] += jumlah_kerjasama

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by cooperation count
    top_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_cooperation"], reverse=True)[:5]
    bottom_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_cooperation"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_count = list(reports)

    # Sort by Jumlah Perusahaan Kerjasama P2P
    top_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("Jumlah Perusahaan Kerjasama P2P", 0), reverse=True)[:5]
    bottom_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("Jumlah Perusahaan Kerjasama P2P", 0))[:5]

    return {
        "top_5_reports_by_count": top_5_reports_by_count,
        "bottom_5_reports_by_count": bottom_5_reports_by_count,
        "top_5_pjps_by_count": top_5_pjps_by_count,
        "bottom_5_pjps_by_count": bottom_5_pjps_by_count,
    }

# NEW ROUTE: Global P2P Cooperation Analysis Page
@app.route('/analisis_kerjasamap2p_global')
def analisis_kerjasamap2p_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_p2p_reports = []
    # Iterate through each PJP's P2P Cooperation data
    for pjp_sandi, pjp_years_data in P2P_COOPERATION_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        for year_str, months_data in pjp_years_data.items():
            # Iterate through each month for the current year
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["Sandi PJP"] = pjp_sandi
                report_copy["Nama PJP"] = pjp_name
                report_copy["Tahun Laporan"] = int(year_str)
                report_copy["Periode Luler"] = month_str # Store original month string for filtering
                report_copy["Periode Laporan Nama"] = get_month_name(month_str) # For display
                all_p2p_reports.append(report_copy)

    # Sort all reports by year and then by month
    all_p2p_reports.sort(key=lambda x: (x["Tahun Laporan"], get_month_number(x["Periode Luler"])))

    # Get unique periods for filtering (e.g., "2023-01", "2023-02", etc.)
    unique_periods = sorted(list(set([f"{r['Tahun Laporan']}-{r['Periode Luler']}" for r in all_p2p_reports])))
    
    # Transform unique_periods for display in dropdown
    display_periods = [
        {"value": period, "text": f"{get_month_name(period.split('-')[1])} {period.split('-')[0]}"}
        for period in unique_periods
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_cooperation_global = sum([report.get("Jumlah Perusahaan Kerjasama P2P", 0) for report in all_p2p_reports])
    total_reports_global = len(all_p2p_reports)
    avg_cooperation_per_report_global = total_cooperation_global / total_reports_global if total_reports_global > 0 else 0

    kpis_global_p2p = {
        "total_cooperation_global": total_cooperation_global,
        "total_reports_global": total_reports_global,
        "avg_cooperation_per_report_global": avg_cooperation_per_report_global,
    }

    # Prepare data for trend charts (monthly and yearly for all data)
    monthly_p2p_trends = {} # { "YYYY-MM": { "total_cooperation": N } }
    yearly_p2p_trends = {}  # { "YYYY": { "total_cooperation": N } }

    for report in all_p2p_reports:
        year = report["Tahun Laporan"]
        month = report["Periode Luler"] # e.g., "01", "02"
        jumlah_kerjasama = report.get("Jumlah Perusahaan Kerjasama P2P", 0)

        # Monthly trends
        month_key = f"{year}-{month}"
        if month_key not in monthly_p2p_trends:
            monthly_p2p_trends[month_key] = {"total_cooperation": 0}
        monthly_p2p_trends[month_key]["total_cooperation"] += jumlah_kerjasama

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_p2p_trends:
            yearly_p2p_trends[year_key] = {"total_cooperation": 0}
        yearly_p2p_trends[year_key]["total_cooperation"] += jumlah_kerjasama
    
    # Sort monthly trends for chart labels
    sorted_monthly_keys_p2p = sorted(monthly_p2p_trends.keys())
    monthly_trend_labels_p2p = [f"{get_month_name(k.split('-')[1])} {k.split('-')[0]}" for k in sorted_monthly_keys_p2p]
    monthly_trend_total_cooperation = [monthly_p2p_trends[k]["total_cooperation"] for k in sorted_monthly_keys_p2p]

    # Sort yearly trends for chart labels
    sorted_yearly_keys_p2p = sorted(yearly_p2p_trends.keys())
    yearly_trend_labels_p2p = [k for k in sorted_yearly_keys_p2p]
    yearly_trend_total_cooperation = [yearly_p2p_trends[k]["total_cooperation"] for k in sorted_yearly_keys_p2p]

    trend_chart_data_p2p = {
        "monthly": {
            "labels": monthly_trend_labels_p2p,
            "total_cooperation": monthly_trend_total_cooperation,
        },
        "yearly": {
            "labels": yearly_trend_labels_p2p,
            "total_cooperation": yearly_trend_total_cooperation,
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_p2p = get_top_bottom_analysis_p2p(all_p2p_reports)


    return render_template(
        'analisis_kerjasamap2p_global.html',
        all_p2p_reports=all_p2p_reports, # Pass the Python list directly
        unique_periods=display_periods,
        kpis_global=kpis_global_p2p,
        trend_chart_data=trend_chart_data_p2p, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_p2p # Pass initial top/bottom data
    )

# ... (kode app.py lainnya tetap sama)

# NEW ROUTE: Export P2P Cooperation Data to XLSX
@app.route('/export_kerjasamap2p_xlsx')
def export_kerjasamap2p_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_month = request.args.get('startMonth')
    start_year = request.args.get('startYear')
    end_month = request.args.get('endMonth')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_kerjasamap2p_global route
    all_p2p_reports_raw = []
    for pjp_sandi, pjp_years_data in P2P_COOPERATION_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        for year_str, months_data in pjp_years_data.items():
            for month_str, report in months_data.items():
                report_copy = report.copy()
                report_copy["Sandi PJP"] = pjp_sandi
                report_copy["Nama PJP"] = pjp_name
                report_copy["Tahun Laporan"] = int(year_str)
                report_copy["Periode Luler"] = month_str # Store original month string for filtering
                report_copy["Periode Laporan Nama"] = get_month_name(month_str) # For display
                all_p2p_reports_raw.append(report_copy)
    
    if start_year and end_year and start_month and end_month:
        # Ensure month values are two digits for consistent parsing
        start_period_value = int(start_year + str(get_month_number(start_month)).zfill(2))
        end_period_value = int(end_year + str(get_month_number(end_month)).zfill(2))
        
        filtered_reports = [
            report for report in all_p2p_reports_raw
            if int(str(report.get("Tahun Laporan", 0)) + str(get_month_number(report.get("Periode Luler", "01"))).zfill(2)) >= start_period_value and \
               int(str(report.get("Tahun Laporan", 0)) + str(get_month_number(report.get("Periode Luler", "01"))).zfill(2)) <= end_period_value
        ]
    else:
        filtered_reports = all_p2p_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet for main reports
    workbook = openpyxl.Workbook()
    main_sheet = workbook.active
    main_sheet.title = "Laporan Kerjasama P2P Global"

    # Define headers for main sheet
    main_headers = [
        "Tahun", "Periode", "Nama PJP", "Sandi PJP",
        "Jumlah Perusahaan Kerjasama P2P", "Keterangan Kerjasama", "Created At"
    ]
    main_sheet.append(main_headers)

    # Add data rows to main sheet
    for report in filtered_reports:
        # Access properties using .get() with a default value to prevent KeyError
        # Convert all values to string to ensure openpyxl handles them gracefully
        row_data = [
            str(report.get("Tahun Laporan", "")),
            str(report.get("Periode Laporan Nama", "")),
            str(report.get("Nama PJP", "")),
            str(report.get("Sandi PJP", "")),
            report.get("Jumlah Perusahaan Kerjasama P2P", 0), # Keep as number for Excel
            str(report.get("Keterangan Kerjasama", "")),
            str(report.get("Created at", ""))
        ]
        main_sheet.append(row_data)

    # Adjust column widths for main sheet
    for column in main_sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[column_name].width = adjusted_width

    # Create a new sheet for "Perusahaan Kerjasama P2P" details
    p2p_details_sheet = workbook.create_sheet(title="Detail Perusahaan Kerjasama")
    p2p_details_headers = [
        "Tahun Laporan Induk", "Periode Laporan Induk", "Nama PJP Induk", "Nomor Surat Induk",
        "Nama Perusahaan Kerjasama", "Peran PJP", "Tanggal Mulai Kerjasama",
        "Tanggal Akhir Kerjasama", "Keterangan Detail"
    ]
    p2p_details_sheet.append(p2p_details_headers)

    # Add data to the new sheet
    for report in filtered_reports:
        # Check if 'Perusahaan Kerjasama P2P' key exists and is a list
        if "Perusahaan Kerjasama P2P" in report and isinstance(report["Perusahaan Kerjasama P2P"], list):
            for company in report["Perusahaan Kerjasama P2P"]:
                detail_row_data = [
                    str(report.get("Tahun Laporan", "")),
                    str(report.get("Periode Laporan Nama", "")),
                    str(report.get("Nama PJP", "")),
                    str(report.get("Nomor Surat", "")), # Assuming Nomor Surat is available in the main report
                    str(company.get("Nama Perusahaan Kerjasama", "")),
                    str(company.get("Peran PJP", "")),
                    str(company.get("Tanggal Mulai Kerjasama", "")),
                    str(company.get("Tanggal Akhir Kerjasama", "")),
                    str(company.get("Keterangan", ""))
                ]
                p2p_details_sheet.append(detail_row_data)
    
    # Adjust column widths for the new sheet
    for column in p2p_details_sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        p2p_details_sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Kerjasama_P2P_Global"]
    if start_year and start_month and end_year and end_month:
        start_month_name = get_month_name(start_month)
        end_month_name = get_month_name(end_month)
        filename_parts.append(f"{start_month_name}_{start_year}_to_{end_month_name}_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# app.py additions (only new routes and helper functions)

# Helper function to get top/bottom reports and PJPs for Gangguan IT
def get_top_bottom_analysis_gangguanit(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_incidents": N, "total_potential_loss": M } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        jumlah_insiden = report.get("jumlah_insiden", 0)
        potensi_kerugian = report.get("potensi_kerugian", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_incidents": 0, "total_potential_loss": 0}
        pjp_aggregated[pjp_name]["total_incidents"] += jumlah_insiden
        pjp_aggregated[pjp_name]["total_potential_loss"] += potensi_kerugian

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by incident count
    top_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_incidents"], reverse=True)[:5]
    bottom_5_pjps_by_count = sorted(pjps_list, key=lambda item: item[1]["total_incidents"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by potential loss
    top_5_pjps_by_loss = sorted(pjps_list, key=lambda item: item[1]["total_potential_loss"], reverse=True)[:5]
    bottom_5_pjps_by_loss = sorted(pjps_list, key=lambda item: item[1]["total_potential_loss"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_count = list(reports)
    reports_by_loss = list(reports)

    # Sort by jumlah_insiden
    top_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("jumlah_insiden", 0), reverse=True)[:5]
    bottom_5_reports_by_count = sorted(reports_by_count, key=lambda a: a.get("jumlah_insiden", 0))[:5]

    # Sort by potensi_kerugian
    top_5_reports_by_loss = sorted(reports_by_loss, key=lambda a: a.get("potensi_kerugian", 0), reverse=True)[:5]
    bottom_5_reports_by_loss = sorted(reports_by_loss, key=lambda a: a.get("potensi_kerugian", 0))[:5]

    return {
        "top_5_reports_by_count": top_5_reports_by_count,
        "bottom_5_reports_by_count": bottom_5_reports_by_count,
        "top_5_reports_by_loss": top_5_reports_by_loss,
        "bottom_5_reports_by_loss": bottom_5_reports_by_loss,
        "top_5_pjps_by_count": top_5_pjps_by_count,
        "bottom_5_pjps_by_count": bottom_5_pjps_by_count,
        "top_5_pjps_by_loss": top_5_pjps_by_loss,
        "bottom_5_pjps_by_loss": bottom_5_pjps_by_loss,
    }

# NEW ROUTE: Global Gangguan IT Analysis Page
@app.route('/analisis_gangguanit_global')
def analisis_gangguanit_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_gangguanit_reports = []
    # Iterate through each Gangguan IT data
    for report in GANGGUAN_IT_REPORTS_DATA:
        report_copy = report.copy()
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == report_copy.get("sandi_pjp")), None)
        report_copy["nama_pjp"] = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({report_copy.get('sandi_pjp')})"
        
        # Ensure 'periode_luler' is available for consistency with other reports, derive from month_laporan
        report_copy["periode_luler"] = str(report_copy.get("bulan_laporan", "01")).zfill(2)
        report_copy["periode_laporan_nama"] = get_month_name(report_copy["periode_luler"])
        
        all_gangguanit_reports.append(report_copy)

    # Sort all reports by year and then by month
    all_gangguanit_reports.sort(key=lambda x: (x.get("tahun_laporan", 0), get_month_number(x.get("periode_luler", "01"))))

    # Get unique periods for filtering (e.g., "2023-01", "2023-02", etc.)
    unique_periods = sorted(list(set([f"{r.get('tahun_laporan', '')}-{r.get('periode_luler', '')}" for r in all_gangguanit_reports])))
    
    # Transform unique_periods for display in dropdown
    display_periods = [
        {"value": period, "text": f"{get_month_name(period.split('-')[1])} {period.split('-')[0]}"}
        for period in unique_periods
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_incidents_global = sum([report.get("jumlah_insiden", 0) for report in all_gangguanit_reports])
    total_potential_loss_global = sum([report.get("potensi_kerugian", 0) for report in all_gangguanit_reports])
    total_reports_global = len(all_gangguanit_reports)
    
    avg_incidents_per_report_global = total_incidents_global / total_reports_global if total_reports_global > 0 else 0
    avg_loss_per_incident_global = total_potential_loss_global / total_incidents_global if total_incidents_global > 0 else 0

    incident_type_counts_global = {}
    for report in all_gangguanit_reports:
        incident_type = report.get("jenis_gangguan", "Tidak Diketahui")
        incident_type_counts_global[incident_type] = incident_type_counts_global.get(incident_type, 0) + 1
    most_common_incident_type_global = max(incident_type_counts_global, key=incident_type_counts_global.get) if incident_type_counts_global else "N/A"


    kpis_global_gangguanit = {
        "total_incidents_global": total_incidents_global,
        "total_potential_loss_global": total_potential_loss_global,
        "total_reports_global": total_reports_global,
        "avg_incidents_per_report_global": avg_incidents_per_report_global,
        "avg_loss_per_incident_global": avg_loss_per_incident_global,
        "most_common_incident_type_global": most_common_incident_type_global,
    }

    # Prepare data for trend charts (monthly and yearly for all data)
    monthly_gangguanit_trends = {} # { "YYYY-MM": { "jumlah_insiden": N, "potensi_kerugian": M } }
    yearly_gangguanit_trends = {}  # { "YYYY": { "jumlah_insiden": N, "potensi_kerugian": M } }

    for report in all_gangguanit_reports:
        year = report.get("tahun_laporan", "")
        month = report.get("periode_luler", "") 
        jumlah_insiden = report.get("jumlah_insiden", 0)
        potensi_kerugian = report.get("potensi_kerugian", 0)

        # Monthly trends
        month_key = f"{year}-{month}"
        if month_key not in monthly_gangguanit_trends:
            monthly_gangguanit_trends[month_key] = {"jumlah_insiden": 0, "potensi_kerugian": 0}
        monthly_gangguanit_trends[month_key]["jumlah_insiden"] += jumlah_insiden
        monthly_gangguanit_trends[month_key]["potensi_kerugian"] += potensi_kerugian

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_gangguanit_trends:
            yearly_gangguanit_trends[year_key] = {"jumlah_insiden": 0, "potensi_kerugian": 0}
        yearly_gangguanit_trends[year_key]["jumlah_insiden"] += jumlah_insiden
        yearly_gangguanit_trends[year_key]["potensi_kerugian"] += potensi_kerugian
    
    # Sort monthly trends for chart labels
    sorted_monthly_keys_gangguanit = sorted(monthly_gangguanit_trends.keys())
    monthly_trend_labels_gangguanit = [f"{get_month_name(k.split('-')[1])} {k.split('-')[0]}" for k in sorted_monthly_keys_gangguanit]
    monthly_trend_jumlah_insiden = [monthly_gangguanit_trends[k]["jumlah_insiden"] for k in sorted_monthly_keys_gangguanit]
    monthly_trend_potensi_kerugian = [monthly_gangguanit_trends[k]["potensi_kerugian"] for k in sorted_monthly_keys_gangguanit]

    # Sort yearly trends for chart labels
    sorted_yearly_keys_gangguanit = sorted(yearly_gangguanit_trends.keys())
    yearly_trend_labels_gangguanit = [k for k in sorted_yearly_keys_gangguanit]
    yearly_trend_jumlah_insiden = [yearly_gangguanit_trends[k]["jumlah_insiden"] for k in sorted_yearly_keys_gangguanit]
    yearly_trend_potensi_kerugian = [yearly_gangguanit_trends[k]["potensi_kerugian"] for k in sorted_yearly_keys_gangguanit]

    trend_chart_data_gangguanit = {
        "monthly": {
            "labels": monthly_trend_labels_gangguanit,
            "jumlah_insiden": monthly_trend_jumlah_insiden,
            "potensi_kerugian": monthly_trend_potensi_kerugian
        },
        "yearly": {
            "labels": yearly_trend_labels_gangguanit,
            "jumlah_insiden": yearly_trend_jumlah_insiden,
            "potensi_kerugian": yearly_trend_potensi_kerugian
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_gangguanit = get_top_bottom_analysis_gangguanit(all_gangguanit_reports)


    return render_template(
        'analisis_gangguanit_global.html',
        all_gangguanit_reports=all_gangguanit_reports, # Pass the Python list directly
        unique_periods=display_periods,
        kpis_global=kpis_global_gangguanit,
        trend_chart_data=trend_chart_data_gangguanit, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_gangguanit # Pass initial top/bottom data
    )

# NEW ROUTE: Export Gangguan IT Data to XLSX
@app.route('/export_gangguanit_xlsx')
def export_gangguanit_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_month = request.args.get('startMonth')
    start_year = request.args.get('startYear')
    end_month = request.args.get('endMonth')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_gangguanit_global route
    all_gangguanit_reports_raw = []
    for report in GANGGUAN_IT_REPORTS_DATA:
        report_copy = report.copy()
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == report_copy.get("sandi_pjp")), None)
        report_copy["nama_pjp"] = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({report_copy.get('sandi_pjp')})"
        report_copy["periode_luler"] = str(report_copy.get("bulan_laporan", "01")).zfill(2)
        report_copy["periode_laporan_nama"] = get_month_name(report_copy["periode_luler"])
        all_gangguanit_reports_raw.append(report_copy)
    
    if start_year and end_year and start_month and end_month:
        start_period_value = int(start_year + str(get_month_number(start_month)).zfill(2))
        end_period_value = int(end_year + str(get_month_number(end_month)).zfill(2))
        
        filtered_reports = [
            report for report in all_gangguanit_reports_raw
            if int(str(report.get("tahun_laporan", 0)) + str(get_month_number(report.get("periode_luler", "01"))).zfill(2)) >= start_period_value and \
               int(str(report.get("tahun_laporan", 0)) + str(get_month_number(report.get("periode_luler", "01"))).zfill(2)) <= end_period_value
        ]
    else:
        filtered_reports = all_gangguanit_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Gangguan IT Global"

    # Define headers
    headers = [
        "Tahun Laporan", "Periode Laporan", "Nama PJP", "Sandi PJP", "Nomor Surat",
        "Tanggal Kejadian", "Waktu Kejadian", "Jenis Gangguan", "Keterangan", "Potensi Kerugian",
        "Tanggal Dibuat"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("periode_laporan_nama", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            str(report.get("nomor_surat", "")),
            str(report.get("tanggal_kejadian", "")),
            str(report.get("waktu_kejadian", "")),
            str(report.get("jenis_gangguan", "")),
            str(report.get("keterangan", "")),
            report.get("potensi_kerugian", 0),
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Gangguan_IT_Global"]
    if start_year and start_month and end_year and end_month:
        start_month_name = get_month_name(start_month)
        end_month_name = get_month_name(end_month)
        filename_parts.append(f"{start_month_name}_{start_year}_to_{end_month_name}_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# app.py additions (only new routes and helper functions)
# app.py additions (only new routes and helper functions)

# Helper function to get top/bottom reports and PJPs for Keuangan Triwulanan
def get_top_bottom_analysis_keuangantriwulanan(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_aset": N, "laba_bersih": M } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        total_aset = report.get("total_aset", 0)
        laba_bersih = report.get("laba", 0) - report.get("rugi", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_aset": 0, "laba_bersih": 0}
        pjp_aggregated[pjp_name]["total_aset"] += total_aset
        pjp_aggregated[pjp_name]["laba_bersih"] += laba_bersih

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by total_aset
    top_5_pjps_by_aset = sorted(pjps_list, key=lambda item: item[1]["total_aset"], reverse=True)[:5]
    bottom_5_pjps_by_aset = sorted(pjps_list, key=lambda item: item[1]["total_aset"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by laba_bersih
    top_5_pjps_by_laba_bersih = sorted(pjps_list, key=lambda item: item[1]["laba_bersih"], reverse=True)[:5]
    bottom_5_pjps_by_laba_bersih = sorted(pjps_list, key=lambda item: item[1]["laba_bersih"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_aset = list(reports)
    reports_by_laba_bersih = list(reports)

    # Sort by total_aset
    top_5_reports_by_aset = sorted(reports_by_aset, key=lambda a: a.get("total_aset", 0), reverse=True)[:5]
    bottom_5_reports_by_aset = sorted(reports_by_aset, key=lambda a: a.get("total_aset", 0))[:5]

    # Sort by laba_bersih
    top_5_reports_by_laba_bersih = sorted(reports_by_laba_bersih, key=lambda a: a.get("laba", 0) - a.get("rugi", 0), reverse=True)[:5]
    bottom_5_reports_by_laba_bersih = sorted(reports_by_laba_bersih, key=lambda a: a.get("laba", 0) - a.get("rugi", 0))[:5]

    return {
        "top_5_reports_by_aset": top_5_reports_by_aset,
        "bottom_5_reports_by_aset": bottom_5_reports_by_aset,
        "top_5_reports_by_laba_bersih": top_5_reports_by_laba_bersih,
        "bottom_5_reports_by_laba_bersih": bottom_5_reports_by_laba_bersih,
        "top_5_pjps_by_aset": top_5_pjps_by_aset,
        "bottom_5_pjps_by_aset": bottom_5_pjps_by_aset,
        "top_5_pjps_by_laba_bersih": top_5_pjps_by_laba_bersih,
        "bottom_5_pjps_by_laba_bersih": bottom_5_pjps_by_laba_bersih,
    }

# NEW ROUTE: Global Keuangan Triwulanan Analysis Page
@app.route('/analisis_keuangantriwulanan_global')
def analisis_keuangantriwulanan_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_keuangan_triwulanan_reports = []
    # Iterate through each PJP's Keuangan Triwulanan data
    for pjp_sandi, pjp_years_data in KEUANGAN_TRIWULANAN_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        for year_str, periods_data in pjp_years_data.items():
            # Iterate through each quarter for the current year
            for period_key, report in periods_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                report_copy["periode_luler"] = period_key # Store original quarter key for filtering
                report_copy["periode_laporan_nama"] = period_key # For display (Q1, Q2, etc.)
                all_keuangan_triwulanan_reports.append(report_copy)

    # Sort all reports by year and then by quarter
    all_keuangan_triwulanan_reports.sort(key=lambda x: (x.get("tahun_laporan", 0), get_month_number(x.get("periode_luler", "Q1"))))

    # Get unique periods for filtering (e.g., "2023-Q1", "2023-Q2", etc.)
    unique_periods = sorted(list(set([f"{r.get('tahun_laporan', '')}-{r.get('periode_luler', '')}" for r in all_keuangan_triwulanan_reports])))
    
    # Transform unique_periods for display in dropdown
    display_periods = [
        {"value": period, "text": f"{period.split('-')[1]} {period.split('-')[0]}"} # Display as Qx YYYY
        for period in unique_periods
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_aset_global = sum([report.get("total_aset", 0) for report in all_keuangan_triwulanan_reports])
    total_ekuitas_global = sum([report.get("total_ekuitas", 0) for report in all_keuangan_triwulanan_reports])
    total_pendapatan_global = sum([report.get("total_pendapatan", 0) for report in all_keuangan_triwulanan_reports])
    total_laba_bersih_global = sum([report.get("laba", 0) - report.get("rugi", 0) for report in all_keuangan_triwulanan_reports])
    total_reports_global = len(all_keuangan_triwulanan_reports)

    avg_aset_per_report_global = total_aset_global / total_reports_global if total_reports_global > 0 else 0
    avg_laba_bersih_per_report_global = total_laba_bersih_global / total_reports_global if total_reports_global > 0 else 0


    kpis_global_keuangantriwulanan = {
        "total_aset_global": total_aset_global,
        "total_ekuitas_global": total_ekuitas_global,
        "total_pendapatan_global": total_pendapatan_global,
        "total_laba_bersih_global": total_laba_bersih_global,
        "total_reports_global": total_reports_global,
        "avg_aset_per_report_global": avg_aset_per_report_global,
        "avg_laba_bersih_per_report_global": avg_laba_bersih_per_report_global,
    }

    # Prepare data for trend charts (monthly/quarterly and yearly for all data)
    monthly_keuangantriwulanan_trends = {} # { "YYYY-Qx": { "total_aset": N, "laba_bersih": M } }
    yearly_keuangantriwulanan_trends = {}  # { "YYYY": { "total_aset": N, "laba_bersih": M } }

    for report in all_keuangan_triwulanan_reports:
        year = report.get("tahun_laporan", "")
        quarter = report.get("periode_luler", "") 
        total_aset = report.get("total_aset", 0)
        laba_bersih = report.get("laba", 0) - report.get("rugi", 0)

        # Monthly/Quarterly trends
        quarter_key = f"{year}-{quarter}"
        if quarter_key not in monthly_keuangantriwulanan_trends:
            monthly_keuangantriwulanan_trends[quarter_key] = {"total_aset": 0, "laba_bersih": 0}
        monthly_keuangantriwulanan_trends[quarter_key]["total_aset"] += total_aset
        monthly_keuangantriwulanan_trends[quarter_key]["laba_bersih"] += laba_bersih

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_keuangantriwulanan_trends:
            yearly_keuangantriwulanan_trends[year_key] = {"total_aset": 0, "laba_bersih": 0}
        yearly_keuangantriwulanan_trends[year_key]["total_aset"] += total_aset
        yearly_keuangantriwulanan_trends[year_key]["laba_bersih"] += laba_bersih
    
    # Sort monthly/quarterly trends for chart labels
    sorted_monthly_keys_keuangantriwulanan = sorted(monthly_keuangantriwulanan_trends.keys(), key=lambda k: (int(k.split('-')[0]), get_month_number(k.split('-')[1])))
    monthly_trend_labels_keuangantriwulanan = [k for k in sorted_monthly_keys_keuangantriwulanan]
    monthly_trend_total_aset = [monthly_keuangantriwulanan_trends[k]["total_aset"] for k in sorted_monthly_keys_keuangantriwulanan]
    monthly_trend_laba_bersih = [monthly_keuangantriwulanan_trends[k]["laba_bersih"] for k in sorted_monthly_keys_keuangantriwulanan]

    # Sort yearly trends for chart labels
    sorted_yearly_keys_keuangantriwulanan = sorted(yearly_keuangantriwulanan_trends.keys())
    yearly_trend_labels_keuangantriwulanan = [k for k in sorted_yearly_keys_keuangantriwulanan]
    yearly_trend_total_aset = [yearly_keuangantriwulanan_trends[k]["total_aset"] for k in sorted_yearly_keys_keuangantriwulanan]
    yearly_trend_laba_bersih = [yearly_keuangantriwulanan_trends[k]["laba_bersih"] for k in sorted_yearly_keys_keuangantriwulanan]

    trend_chart_data_keuangantriwulanan = {
        "monthly": {
            "labels": monthly_trend_labels_keuangantriwulanan,
            "total_aset": monthly_trend_total_aset,
            "laba_bersih": monthly_trend_laba_bersih
        },
        "yearly": {
            "labels": yearly_trend_labels_keuangantriwulanan,
            "total_aset": yearly_trend_total_aset,
            "laba_bersih": yearly_trend_laba_bersih
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_keuangantriwulanan = get_top_bottom_analysis_keuangantriwulanan(all_keuangan_triwulanan_reports)


    return render_template(
        'analisis_keuangantriwulanan_global.html',
        all_keuangan_triwulanan_reports=all_keuangan_triwulanan_reports, # Pass the Python list directly
        unique_periods=display_periods,
        kpis_global=kpis_global_keuangantriwulanan,
        trend_chart_data=trend_chart_data_keuangantriwulanan, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_keuangantriwulanan # Pass initial top/bottom data
    )

# NEW ROUTE: Export Keuangan Triwulanan Data to XLSX
@app.route('/export_keuangantriwulanan_xlsx')
def export_keuangantriwulanan_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_month = request.args.get('startMonth')
    start_year = request.args.get('startYear')
    end_month = request.args.get('endMonth')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_keuangantriwulanan_global route
    all_keuangan_triwulanan_reports_raw = []
    for pjp_sandi, pjp_years_data in KEUANGAN_TRIWULANAN_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        for year_str, periods_data in pjp_years_data.items():
            for period_key, report in periods_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                report_copy["periode_luler"] = period_key
                report_copy["periode_laporan_nama"] = period_key
                all_keuangan_triwulanan_reports_raw.append(report_copy)
    
    if start_year and end_year and start_month and end_month:
        # For quarterly reports, month represents the quarter (e.g., 'Q1' -> 1, 'Q2' -> 4)
        start_period_value = int(start_year + str(get_month_number(start_month)).zfill(2))
        end_period_value = int(end_year + str(get_month_number(end_month)).zfill(2))
        
        filtered_reports = [
            report for report in all_keuangan_triwulanan_reports_raw
            if int(str(report.get("tahun_laporan", 0)) + str(get_month_number(report.get("periode_luler", "Q1"))).zfill(2)) >= start_period_value and \
               int(str(report.get("tahun_laporan", 0)) + str(get_month_number(report.get("periode_luler", "Q1"))).zfill(2)) <= end_period_value
        ]
    else:
        filtered_reports = all_keuangan_triwulanan_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Keuangan Triwulanan Global"

    # Define headers (adjust based on your actual data fields)
    headers = [
        "Tahun Laporan", "Periode Laporan", "Nama PJP", "Sandi PJP",
        "Modal Dasar", "Modal Disetor", "Total Aset", "Aset Lancar", "Aset Tidak Lancar",
        "Total Hutang", "Hutang Jangka Pendek", "Hutang Jangka Panjang",
        "Total Ekuitas", "Total Pendapatan", "Pendapatan Fee", "Beban Operasional", "Total Beban", "Laba", "Rugi",
        "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("periode_laporan_nama", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            report.get("modal_dasar", 0),
            report.get("modal_disetor", 0),
            report.get("total_aset", 0),
            report.get("aset_lancar", 0),
            report.get("aset_tidak_lancar", 0),
            report.get("total_hutang", 0),
            report.get("hutang_jangka_pendek", 0),
            report.get("hutang_jangka_panjang", 0),
            report.get("total_ekuitas", 0),
            report.get("total_pendapatan", 0),
            report.get("pendapatan_fee", 0),
            report.get("beban_operasional", 0),
            report.get("total_beban", 0),
            report.get("laba", 0),
            report.get("rugi", 0),
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Keuangan_Triwulanan_Global"]
    if start_year and start_month and end_year and end_month:
        start_month_name = get_month_name(start_month)
        end_month_name = get_month_name(end_month)
        filename_parts.append(f"{start_month_name}_{start_year}_to_{end_month_name}_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# app.py additions (only new routes and helper functions)

# Helper function to get top/bottom reports and PJPs for Keuangan Tahunan
def get_top_bottom_analysis_keuangantahunan(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_aset": N, "laba_bersih": M } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        total_aset = report.get("total_aset", 0)
        laba_bersih = report.get("laba", 0) - report.get("rugi", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_aset": 0, "laba_bersih": 0}
        pjp_aggregated[pjp_name]["total_aset"] += total_aset
        pjp_aggregated[pjp_name]["laba_bersih"] += laba_bersih

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by total_aset
    top_5_pjps_by_aset = sorted(pjps_list, key=lambda item: item[1]["total_aset"], reverse=True)[:5]
    bottom_5_pjps_by_aset = sorted(pjps_list, key=lambda item: item[1]["total_aset"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by laba_bersih
    top_5_pjps_by_laba_bersih = sorted(pjps_list, key=lambda item: item[1]["laba_bersih"], reverse=True)[:5]
    bottom_5_pjps_by_laba_bersih = sorted(pjps_list, key=lambda item: item[1]["laba_bersih"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_aset = list(reports)
    reports_by_laba_bersih = list(reports)

    # Sort by total_aset
    top_5_reports_by_aset = sorted(reports_by_aset, key=lambda a: a.get("total_aset", 0), reverse=True)[:5]
    bottom_5_reports_by_aset = sorted(reports_by_aset, key=lambda a: a.get("total_aset", 0))[:5]

    # Sort by laba_bersih
    top_5_reports_by_laba_bersih = sorted(reports_by_laba_bersih, key=lambda a: a.get("laba", 0) - a.get("rugi", 0), reverse=True)[:5]
    bottom_5_reports_by_laba_bersih = sorted(reports_by_laba_bersih, key=lambda a: a.get("laba", 0) - a.get("rugi", 0))[:5]

    return {
        "top_5_reports_by_aset": top_5_reports_by_aset,
        "bottom_5_reports_by_aset": bottom_5_reports_by_aset,
        "top_5_reports_by_laba_bersih": top_5_reports_by_laba_bersih,
        "bottom_5_reports_by_laba_bersih": bottom_5_reports_by_laba_bersih,
        "top_5_pjps_by_aset": top_5_pjps_by_aset,
        "bottom_5_pjps_by_aset": bottom_5_pjps_by_aset,
        "top_5_pjps_by_laba_bersih": top_5_pjps_by_laba_bersih,
        "bottom_5_pjps_by_laba_bersih": bottom_5_pjps_by_laba_bersih,
    }

# NEW ROUTE: Global Keuangan Tahunan Analysis Page
@app.route('/analisis_keuangantahunan_global')
def analisis_keuangantahunan_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_keuangan_tahunan_reports = []
    # Iterate through each PJP's Keuangan Tahunan data
    for pjp_sandi, pjp_years_data in KEUANGAN_TAHUNAN_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        report = pjp_years_data # For annual data, pjp_years_data directly holds the report for that year
        if report: # Check if report exists for the year
            for year_str, report_data in report.items(): # Iterate over the single year entry
                report_copy = report_data.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                # For annual reports, periode_luler and periode_laporan_nama can just be the year
                report_copy["periode_luler"] = year_str 
                report_copy["periode_laporan_nama"] = year_str
                all_keuangan_tahunan_reports.append(report_copy)

    # Sort all reports by year
    all_keuangan_tahunan_reports.sort(key=lambda x: x.get("tahun_laporan", 0))

    # Get unique years for filtering
    unique_years = sorted(list(set([r.get('tahun_laporan', '') for r in all_keuangan_tahunan_reports])))
    
    # Transform unique_years for display in dropdown
    display_years = [
        {"value": str(year), "text": str(year)}
        for year in unique_years
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_aset_global = sum([report.get("total_aset", 0) for report in all_keuangan_tahunan_reports])
    total_ekuitas_global = sum([report.get("total_ekuitas", 0) for report in all_keuangan_tahunan_reports])
    total_pendapatan_global = sum([report.get("pendapatan", 0) for report in all_keuangan_tahunan_reports])
    total_laba_bersih_global = sum([report.get("laba_bersih", 0) for report in all_keuangan_tahunan_reports])
    total_reports_global = len(all_keuangan_tahunan_reports)

    avg_aset_per_report_global = total_aset_global / total_reports_global if total_reports_global > 0 else 0
    avg_laba_bersih_per_report_global = total_laba_bersih_global / total_reports_global if total_reports_global > 0 else 0


    kpis_global_keuangantahunan = {
        "total_aset_global": total_aset_global,
        "total_ekuitas_global": total_ekuitas_global,
        "total_pendapatan_global": total_pendapatan_global,
        "total_laba_bersih_global": total_laba_bersih_global,
        "total_reports_global": total_reports_global,
        "avg_aset_per_report_global": avg_aset_per_report_global,
        "avg_laba_bersih_per_report_global": avg_laba_bersih_per_report_global,
    }

    # Prepare data for trend charts (yearly only for annual reports)
    yearly_keuangantahunan_trends = {}  # { "YYYY": { "total_aset": N, "laba_bersih": M } }

    for report in all_keuangan_tahunan_reports:
        year = report.get("tahun_laporan", "")
        total_aset = report.get("total_aset", 0)
        laba_bersih = report.get("laba_bersih", 0)

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_keuangantahunan_trends:
            yearly_keuangantahunan_trends[year_key] = {"total_aset": 0, "laba_bersih": 0}
        yearly_keuangantahunan_trends[year_key]["total_aset"] += total_aset
        yearly_keuangantahunan_trends[year_key]["laba_bersih"] += laba_bersih
    
    # Sort yearly trends for chart labels
    sorted_yearly_keys_keuangantahunan = sorted(yearly_keuangantahunan_trends.keys())
    yearly_trend_labels_keuangantahunan = [k for k in sorted_yearly_keys_keuangantahunan]
    yearly_trend_total_aset = [yearly_keuangantahunan_trends[k]["total_aset"] for k in sorted_yearly_keys_keuangantahunan]
    yearly_trend_laba_bersih = [yearly_keuangantahunan_trends[k]["laba_bersih"] for k in sorted_yearly_keys_keuangantahunan]

    trend_chart_data_keuangantahunan = {
        "yearly": {
            "labels": yearly_trend_labels_keuangantahunan,
            "total_aset": yearly_trend_total_aset,
            "laba_bersih": yearly_trend_laba_bersih
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_keuangantahunan = get_top_bottom_analysis_keuangantahunan(all_keuangan_tahunan_reports)


    return render_template(
        'analisis_keuangantahunan_global.html',
        all_keuangan_tahunan_reports=all_keuangan_tahunan_reports, # Pass the Python list directly
        unique_years=display_years, # Pass unique years for filter
        kpis_global=kpis_global_keuangantahunan,
        trend_chart_data=trend_chart_data_keuangantahunan, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_keuangantahunan # Pass initial top/bottom data
    )

# NEW ROUTE: Export Keuangan Tahunan Data to XLSX
@app.route('/export_keuangantahunan_xlsx')
def export_keuangantahunan_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_year = request.args.get('startYear')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_keuangantahunan_global route
    all_keuangan_tahunan_reports_raw = []
    for pjp_sandi, pjp_years_data in KEUANGAN_TAHUNAN_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        report_data = pjp_years_data # For annual data, pjp_years_data directly holds the report for that year
        if report_data:
            for year_str, report in report_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                all_keuangan_tahunan_reports_raw.append(report_copy)
    
    if start_year and end_year:
        start_year_int = int(start_year)
        end_year_int = int(end_year)
        
        filtered_reports = [
            report for report in all_keuangan_tahunan_reports_raw
            if report.get("tahun_laporan", 0) >= start_year_int and \
               report.get("tahun_laporan", 0) <= end_year_int
        ]
    else:
        filtered_reports = all_keuangan_tahunan_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Keuangan Tahunan Global"

    # Define headers (adjust based on your actual data fields)
    headers = [
        "Tahun Laporan", "Nama PJP", "Sandi PJP",
        "Modal Dasar", "Modal Disetor", "Total Aset", "Aset Lancar", "Aset Tidak Lancar",
        "Total Hutang", "Hutang Jangka Pendek", "Hutang Jangka Panjang",
        "Total Ekuitas", "Total Pendapatan", "Pendapatan Fee", "Beban Operasional", "Total Beban", "Laba", "Rugi",
        "Laba Bersih", "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            report.get("modal_dasar", 0),
            report.get("modal_disetor", 0),
            report.get("total_aset", 0),
            report.get("aset_lancar", 0),
            report.get("aset_tidak_lancar", 0),
            report.get("total_hutang", 0),
            report.get("hutang_jangka_pendek", 0),
            report.get("hutang_jangka_panjang", 0),
            report.get("total_ekuitas", 0),
            report.get("total_pendapatan", 0),
            report.get("pendapatan_fee", 0),
            report.get("beban_operasional", 0),
            report.get("total_beban", 0),
            report.get("laba", 0),
            report.get("rugi", 0),
            (report.get("laba", 0) - report.get("rugi", 0)), # Laba Bersih
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Keuangan_Tahunan_Global"]
    if start_year and end_year:
        filename_parts.append(f"{start_year}_to_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ... (kode app.py lainnya tetap sama)

# Helper function to get top/bottom reports and PJPs for Pentest
def get_top_bottom_analysis_pentest(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_findings": N, "total_critical": M } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        jumlah_temuan = report.get("jumlah_temuan", 0)
        temuan_critical = report.get("temuan_critical", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_findings": 0, "total_critical": 0}
        pjp_aggregated[pjp_name]["total_findings"] += jumlah_temuan
        pjp_aggregated[pjp_name]["total_critical"] += temuan_critical

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by total findings
    top_5_pjps_by_findings = sorted(pjps_list, key=lambda item: item[1]["total_findings"], reverse=True)[:5]
    bottom_5_pjps_by_findings = sorted(pjps_list, key=lambda item: item[1]["total_findings"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by critical findings
    top_5_pjps_by_critical = sorted(pjps_list, key=lambda item: item[1]["total_critical"], reverse=True)[:5]
    bottom_5_pjps_by_critical = sorted(pjps_list, key=lambda item: item[1]["total_critical"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_findings = list(reports)
    reports_by_critical = list(reports)

    # Sort by jumlah_temuan
    top_5_reports_by_findings = sorted(reports_by_findings, key=lambda a: a.get("jumlah_temuan", 0), reverse=True)[:5]
    bottom_5_reports_by_findings = sorted(reports_by_findings, key=lambda a: a.get("jumlah_temuan", 0))[:5]

    # Sort by temuan_critical
    top_5_reports_by_critical = sorted(reports_by_critical, key=lambda a: a.get("temuan_critical", 0), reverse=True)[:5]
    bottom_5_reports_by_critical = sorted(reports_by_critical, key=lambda a: a.get("temuan_critical", 0))[:5]

    return {
        "top_5_reports_by_findings": top_5_reports_by_findings,
        "bottom_5_reports_by_findings": bottom_5_reports_by_findings,
        "top_5_reports_by_critical": top_5_reports_by_critical,
        "bottom_5_reports_by_critical": bottom_5_reports_by_critical,
        "top_5_pjps_by_findings": top_5_pjps_by_findings,
        "bottom_5_pjps_by_findings": bottom_5_pjps_by_findings,
        "top_5_pjps_by_critical": top_5_pjps_by_critical,
        "bottom_5_pjps_by_critical": bottom_5_pjps_by_critical,
    }

# NEW ROUTE: Global Pentest Analysis Page
@app.route('/analisis_pentest_global')
def analisis_pentest_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_pentest_reports = []
    # Iterate through each PJP's Pentest data
    for pjp_sandi, pjp_years_data in PENTEST_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        report = pjp_years_data # For annual data, pjp_years_data directly holds the report for that year
        if report: # Check if report exists for the year
            for year_str, report_data in report.items(): # Iterate over the single year entry
                report_copy = report_data.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                # For annual reports, periode_luler and periode_laporan_nama can just be the year
                report_copy["periode_luler"] = year_str 
                report_copy["periode_laporan_nama"] = year_str
                all_pentest_reports.append(report_copy)

    # Sort all reports by year
    all_pentest_reports.sort(key=lambda x: x.get("tahun_laporan", 0))

    # Get unique years for filtering
    unique_years = sorted(list(set([r.get('tahun_laporan', '') for r in all_pentest_reports])))
    
    # Transform unique_years for display in dropdown
    display_years = [
        {"value": str(year), "text": str(year)}
        for year in unique_years
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_findings_global = sum([report.get("jumlah_temuan", 0) for report in all_pentest_reports])
    total_resolved_global = sum([report.get("jumlah_temuan_diselesaikan", 0) for report in all_pentest_reports])
    total_critical_global = sum([report.get("temuan_critical", 0) for report in all_pentest_reports])
    total_high_global = sum([report.get("temuan_high", 0) for report in all_pentest_reports])
    total_reports_global = len(all_pentest_reports)

    percentage_resolved_global = (total_resolved_global / total_findings_global * 100) if total_findings_global > 0 else 0
    avg_findings_per_report_global = total_findings_global / total_reports_global if total_reports_global > 0 else 0


    kpis_global_pentest = {
        "total_findings_global": total_findings_global,
        "total_resolved_global": total_resolved_global,
        "total_critical_global": total_critical_global,
        "total_high_global": total_high_global,
        "total_reports_global": total_reports_global,
        "percentage_resolved_global": percentage_resolved_global,
        "avg_findings_per_report_global": avg_findings_per_report_global,
    }

    # Prepare data for trend charts (yearly only for annual reports)
    yearly_pentest_trends = {}  # { "YYYY": { "total_findings": N, "critical_findings": M, ... } }

    for report in all_pentest_reports:
        year = report.get("tahun_laporan", "")
        jumlah_temuan = report.get("jumlah_temuan", 0)
        temuan_critical = report.get("temuan_critical", 0)
        temuan_high = report.get("temuan_high", 0)
        temuan_medium = report.get("temuan_medium", 0)
        temuan_low = report.get("temuan_low", 0)
        jumlah_temuan_diselesaikan = report.get("jumlah_temuan_diselesaikan", 0)

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_pentest_trends:
            yearly_pentest_trends[year_key] = {
                "total_findings": 0,
                "critical_findings": 0,
                "high_findings": 0,
                "medium_findings": 0,
                "low_findings": 0,
                "resolved_findings": 0
            }
        yearly_pentest_trends[year_key]["total_findings"] += jumlah_temuan
        yearly_pentest_trends[year_key]["critical_findings"] += temuan_critical
        yearly_pentest_trends[year_key]["high_findings"] += temuan_high
        yearly_pentest_trends[year_key]["medium_findings"] += temuan_medium
        yearly_pentest_trends[year_key]["low_findings"] += temuan_low
        yearly_pentest_trends[year_key]["resolved_findings"] += jumlah_temuan_diselesaikan
    
    # Sort yearly trends for chart labels
    sorted_yearly_keys_pentest = sorted(yearly_pentest_trends.keys())
    yearly_trend_labels_pentest = [k for k in sorted_yearly_keys_pentest]
    yearly_trend_total_findings = [yearly_pentest_trends[k]["total_findings"] for k in sorted_yearly_keys_pentest]
    yearly_trend_critical_findings = [yearly_pentest_trends[k]["critical_findings"] for k in sorted_yearly_keys_pentest]
    yearly_trend_high_findings = [yearly_pentest_trends[k]["high_findings"] for k in sorted_yearly_keys_pentest]
    yearly_trend_medium_findings = [yearly_pentest_trends[k]["medium_findings"] for k in sorted_yearly_keys_pentest]
    yearly_trend_low_findings = [yearly_pentest_trends[k]["low_findings"] for k in sorted_yearly_keys_pentest]
    yearly_trend_resolved_findings = [yearly_pentest_trends[k]["resolved_findings"] for k in sorted_yearly_keys_pentest]


    trend_chart_data_pentest = {
        "yearly": {
            "labels": yearly_trend_labels_pentest,
            "total_findings": yearly_trend_total_findings,
            "critical_findings": yearly_trend_critical_findings,
            "high_findings": yearly_trend_high_findings,
            "medium_findings": yearly_trend_medium_findings,
            "low_findings": yearly_trend_low_findings,
            "resolved_findings": yearly_trend_resolved_findings,
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_pentest = get_top_bottom_analysis_pentest(all_pentest_reports)


    return render_template(
        'analisis_pentest_global.html',
        all_pentest_reports=all_pentest_reports, # Pass the Python list directly
        unique_years=display_years, # Pass unique years for filter
        kpis_global=kpis_global_pentest,
        trend_chart_data=trend_chart_data_pentest, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_pentest # Pass initial top/bottom data
    )

# NEW ROUTE: Export Pentest Data to XLSX
@app.route('/export_pentest_xlsx')
def export_pentest_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_year = request.args.get('startYear')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_pentest_global route
    all_pentest_reports_raw = []
    for pjp_sandi, pjp_years_data in PENTEST_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        report_data = pjp_years_data
        if report_data:
            for year_str, report in report_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                all_pentest_reports_raw.append(report_copy)
    
    if start_year and end_year:
        start_year_int = int(start_year)
        end_year_int = int(end_year)
        
        filtered_reports = [
            report for report in all_pentest_reports_raw
            if report.get("tahun_laporan", 0) >= start_year_int and \
               report.get("tahun_laporan", 0) <= end_year_int
        ]
    else:
        filtered_reports = all_pentest_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Pentest Global"

    # Define headers (adjust based on your actual data fields)
    headers = [
        "Tahun Laporan", "Nama PJP", "Sandi PJP", "Nomor Surat", "Tanggal Surat",
        "Jumlah Temuan", "Temuan Critical", "Temuan High", "Temuan Medium", "Temuan Low",
        "Jumlah Temuan Diselesaikan", "Jumlah Temuan Belum Diselesaikan", "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            str(report.get("nomor_surat", "")),
            str(report.get("tanggal_surat", "")),
            report.get("jumlah_temuan", 0),
            report.get("temuan_critical", 0),
            report.get("temuan_high", 0),
            report.get("temuan_medium", 0),
            report.get("temuan_low", 0),
            report.get("jumlah_temuan_diselesaikan", 0),
            report.get("jumlah_temuan_belum_diselesaikan", 0),
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Pentest_Global"]
    if start_year and end_year:
        filename_parts.append(f"{start_year}_to_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# ... (kode app.py lainnya tetap sama)

# Helper function to get top/bottom reports and PJPs for Audit SI
def get_top_bottom_analysis_auditsi(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_findings": N, "total_critical": M, "total_high": L } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        jumlah_temuan = report.get("jumlah_temuan", 0)
        temuan_critical = report.get("temuan_critical", 0)
        temuan_high = report.get("temuan_high", 0)

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_findings": 0, "total_critical": 0, "total_high": 0}
        pjp_aggregated[pjp_name]["total_findings"] += jumlah_temuan
        pjp_aggregated[pjp_name]["total_critical"] += temuan_critical
        pjp_aggregated[pjp_name]["total_high"] += temuan_high

    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by total findings
    top_5_pjps_by_findings = sorted(pjps_list, key=lambda item: item[1]["total_findings"], reverse=True)[:5]
    bottom_5_pjps_by_findings = sorted(pjps_list, key=lambda item: item[1]["total_findings"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by critical findings
    top_5_pjps_by_critical = sorted(pjps_list, key=lambda item: item[1]["total_critical"], reverse=True)[:5]
    bottom_5_pjps_by_critical = sorted(pjps_list, key=lambda item: item[1]["total_critical"])[:5]

    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_findings = list(reports)
    reports_by_critical = list(reports)

    # Sort by jumlah_temuan
    top_5_reports_by_findings = sorted(reports_by_findings, key=lambda a: a.get("jumlah_temuan", 0), reverse=True)[:5]
    bottom_5_reports_by_findings = sorted(reports_by_findings, key=lambda a: a.get("jumlah_temuan", 0))[:5]

    # Sort by temuan_critical
    top_5_reports_by_critical = sorted(reports_by_critical, key=lambda a: a.get("temuan_critical", 0), reverse=True)[:5]
    bottom_5_reports_by_critical = sorted(reports_by_critical, key=lambda a: a.get("temuan_critical", 0))[:5]

    return {
        "top_5_reports_by_findings": top_5_reports_by_findings,
        "bottom_5_reports_by_findings": bottom_5_reports_by_findings,
        "top_5_reports_by_critical": top_5_reports_by_critical,
        "bottom_5_reports_by_critical": bottom_5_reports_by_critical,
        "top_5_pjps_by_findings": top_5_pjps_by_findings,
        "bottom_5_pjps_by_findings": bottom_5_pjps_by_findings,
        "top_5_pjps_by_critical": top_5_pjps_by_critical,
        "bottom_5_pjps_by_critical": bottom_5_pjps_by_critical,
    }

# NEW ROUTE: Global Audit SI Analysis Page
@app.route('/analisis_auditsi_global')
def analisis_auditsi_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_auditsi_reports = []
    # Iterate through each PJP's Audit SI data
    for pjp_sandi, pjp_years_data in AUDIT_SI_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        report = pjp_years_data # For annual data, pjp_years_data directly holds the report for that year
        if report: # Check if report exists for the year
            for year_str, report_data in report.items(): # Iterate over the single year entry
                report_copy = report_data.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                all_auditsi_reports.append(report_copy)

    # Sort all reports by year
    all_auditsi_reports.sort(key=lambda x: x.get("tahun_laporan", 0))

    # Get unique years for filtering
    unique_years = sorted(list(set([r.get('tahun_laporan', '') for r in all_auditsi_reports])))
    
    # Transform unique_years for display in dropdown
    display_years = [
        {"value": str(year), "text": str(year)}
        for year in unique_years
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_findings_global = sum([report.get("jumlah_temuan", 0) for report in all_auditsi_reports])
    total_resolved_global = sum([report.get("jumlah_temuan_diselesaikan", 0) for report in all_auditsi_reports])
    total_critical_global = sum([report.get("temuan_critical", 0) for report in all_auditsi_reports])
    total_high_global = sum([report.get("temuan_high", 0) for report in all_auditsi_reports])
    total_reports_global = len(all_auditsi_reports)

    percentage_resolved_global = (total_resolved_global / total_findings_global * 100) if total_findings_global > 0 else 0
    avg_findings_per_report_global = total_findings_global / total_reports_global if total_reports_global > 0 else 0

    avg_confidentiality_global = sum([report.get("confidentiality", 0) for report in all_auditsi_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_integrity_global = sum([report.get("integrity", 0) for report in all_auditsi_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_availability_global = sum([report.get("availability", 0) for report in all_auditsi_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_authenticity_global = sum([report.get("authenticity", 0) for report in all_auditsi_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_non_repudiation_global = sum([report.get("non_repudiation", 0) for report in all_auditsi_reports]) / total_reports_global if total_reports_global > 0 else 0


    kpis_global_auditsi = {
        "total_findings_global": total_findings_global,
        "total_resolved_global": total_resolved_global,
        "total_critical_global": total_critical_global,
        "total_high_global": total_high_global,
        "total_reports_global": total_reports_global,
        "percentage_resolved_global": percentage_resolved_global,
        "avg_findings_per_report_global": avg_findings_per_report_global,
        "avg_confidentiality_global": avg_confidentiality_global,
        "avg_integrity_global": avg_integrity_global,
        "avg_availability_global": avg_availability_global,
        "avg_authenticity_global": avg_authenticity_global,
        "avg_non_repudiation_global": avg_non_repudiation_global,
    }

    # Prepare data for trend charts (yearly only for annual reports)
    yearly_auditsi_trends = {}  # { "YYYY": { "total_findings": N, "critical_findings": M, ... } }

    for report in all_auditsi_reports:
        year = report.get("tahun_laporan", "")
        jumlah_temuan = report.get("jumlah_temuan", 0)
        temuan_critical = report.get("temuan_critical", 0)
        temuan_high = report.get("temuan_high", 0)
        temuan_medium = report.get("temuan_medium", 0)
        temuan_low = report.get("temuan_low", 0)
        jumlah_temuan_diselesaikan = report.get("jumlah_temuan_diselesaikan", 0)
        
        confidentiality = report.get("confidentiality", 0)
        integrity = report.get("integrity", 0)
        availability = report.get("availability", 0)
        authenticity = report.get("authenticity", 0)
        non_repudiation = report.get("non_repudiation", 0)


        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_auditsi_trends:
            yearly_auditsi_trends[year_key] = {
                "total_findings": 0,
                "critical_findings": 0,
                "high_findings": 0,
                "medium_findings": 0,
                "low_findings": 0,
                "resolved_findings": 0,
                "confidentiality_sum": 0, # Sum for average calculation
                "integrity_sum": 0,
                "availability_sum": 0,
                "authenticity_sum": 0,
                "non_repudiation_sum": 0,
                "report_count": 0 # Count reports for average calculation
            }
        yearly_auditsi_trends[year_key]["total_findings"] += jumlah_temuan
        yearly_auditsi_trends[year_key]["critical_findings"] += temuan_critical
        yearly_auditsi_trends[year_key]["high_findings"] += temuan_high
        yearly_auditsi_trends[year_key]["medium_findings"] += temuan_medium
        yearly_auditsi_trends[year_key]["low_findings"] += temuan_low
        yearly_auditsi_trends[year_key]["resolved_findings"] += jumlah_temuan_diselesaikan
        yearly_auditsi_trends[year_key]["confidentiality_sum"] += confidentiality
        yearly_auditsi_trends[year_key]["integrity_sum"] += integrity
        yearly_auditsi_trends[year_key]["availability_sum"] += availability
        yearly_auditsi_trends[year_key]["authenticity_sum"] += authenticity
        yearly_auditsi_trends[year_key]["non_repudiation_sum"] += non_repudiation
        yearly_auditsi_trends[year_key]["report_count"] += 1

    # Sort yearly trends for chart labels and calculate averages
    sorted_yearly_keys_auditsi = sorted(yearly_auditsi_trends.keys())
    yearly_trend_labels_auditsi = [k for k in sorted_yearly_keys_auditsi]
    yearly_trend_total_findings = [yearly_auditsi_trends[k]["total_findings"] for k in sorted_yearly_keys_auditsi]
    yearly_trend_critical_findings = [yearly_auditsi_trends[k]["critical_findings"] for k in sorted_yearly_keys_auditsi]
    yearly_trend_high_findings = [yearly_auditsi_trends[k]["high_findings"] for k in sorted_yearly_keys_auditsi]
    yearly_trend_resolved_findings = [yearly_auditsi_trends[k]["resolved_findings"] for k in sorted_yearly_keys_auditsi]
    
    yearly_trend_avg_confidentiality = [yearly_auditsi_trends[k]["confidentiality_sum"] / yearly_auditsi_trends[k]["report_count"] if yearly_auditsi_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_auditsi]
    yearly_trend_avg_integrity = [yearly_auditsi_trends[k]["integrity_sum"] / yearly_auditsi_trends[k]["report_count"] if yearly_auditsi_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_auditsi]
    yearly_trend_avg_availability = [yearly_auditsi_trends[k]["availability_sum"] / yearly_auditsi_trends[k]["report_count"] if yearly_auditsi_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_auditsi]
    yearly_trend_avg_authenticity = [yearly_auditsi_trends[k]["authenticity_sum"] / yearly_auditsi_trends[k]["report_count"] if yearly_auditsi_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_auditsi]
    yearly_trend_avg_non_repudiation = [yearly_auditsi_trends[k]["non_repudiation_sum"] / yearly_auditsi_trends[k]["report_count"] if yearly_auditsi_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_auditsi]


    trend_chart_data_auditsi = {
        "yearly": {
            "labels": yearly_trend_labels_auditsi,
            "total_findings": yearly_trend_total_findings,
            "critical_findings": yearly_trend_critical_findings,
            "high_findings": yearly_trend_high_findings,
            "resolved_findings": yearly_trend_resolved_findings,
            "avg_confidentiality": yearly_trend_avg_confidentiality,
            "avg_integrity": yearly_trend_avg_integrity,
            "avg_availability": yearly_trend_avg_availability,
            "avg_authenticity": yearly_trend_avg_authenticity,
            "avg_non_repudiation": yearly_trend_avg_non_repudiation,
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_auditsi = get_top_bottom_analysis_auditsi(all_auditsi_reports)


    return render_template(
        'analisis_auditsi_global.html',
        all_auditsi_reports=all_auditsi_reports, # Pass the Python list directly
        unique_years=display_years, # Pass unique years for filter
        kpis_global=kpis_global_auditsi,
        trend_chart_data=trend_chart_data_auditsi, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_auditsi # Pass initial top/bottom data
    )

# NEW ROUTE: Export Audit SI Data to XLSX
@app.route('/export_auditsi_xlsx')
def export_auditsi_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_year = request.args.get('startYear')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_auditsi_global route
    all_auditsi_reports_raw = []
    for pjp_sandi, pjp_years_data in AUDIT_SI_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        report_data = pjp_years_data
        if report_data:
            for year_str, report in report_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                all_auditsi_reports_raw.append(report_copy)
    
    if start_year and end_year:
        start_year_int = int(start_year)
        end_year_int = int(end_year)
        
        filtered_reports = [
            report for report in all_auditsi_reports_raw
            if report.get("tahun_laporan", 0) >= start_year_int and \
               report.get("tahun_laporan", 0) <= end_year_int
        ]
    else:
        filtered_reports = all_auditsi_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Audit SI Global"

    # Define headers (adjust based on your actual data fields)
    headers = [
        "Tahun Laporan", "Nama PJP", "Sandi PJP", "Nomor Surat", "Tanggal Surat",
        "Confidentiality", "Integrity", "Availability", "Authenticity", "Non-Repudiation",
        "Jumlah Temuan", "Jumlah Temuan Diselesaikan", "Jumlah Temuan Belum Diselesaikan",
        "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            str(report.get("nomor_surat", "")),
            str(report.get("tanggal_surat", "")),
            report.get("confidentiality", 0),
            report.get("integrity", 0),
            report.get("availability", 0),
            report.get("authenticity", 0),
            report.get("non_repudiation", 0),
            report.get("jumlah_temuan", 0),
            report.get("jumlah_temuan_diselesaikan", 0),
            report.get("jumlah_temuan_belum_diselesaikan", 0),
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Audit_SI_Global"]
    if start_year and end_year:
        filename_parts.append(f"{start_year}_to_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ... (kode app.py lainnya tetap sama)

# Helper function to get top/bottom reports and PJPs for APUPPT
def get_top_bottom_analysis_apuppt(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_ltkt": N, "sipesat_uploaded_count_sum": M, "report_count": R } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        jumlah_ltkt = report.get("jumlah_ltkt", 0)
        sipesat_uploaded_count = report.get("sipesat_uploaded_count", 0) # Derived in route

        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {"total_ltkt": 0, "sipesat_uploaded_count_sum": 0, "report_count": 0}
        pjp_aggregated[pjp_name]["total_ltkt"] += jumlah_ltkt
        pjp_aggregated[pjp_name]["sipesat_uploaded_count_sum"] += sipesat_uploaded_count
        pjp_aggregated[pjp_name]["report_count"] += 1
        pjp_aggregated[pjp_name]["avg_sipesat_compliance"] = (pjp_aggregated[pjp_name]["sipesat_uploaded_count_sum"] / (pjp_aggregated[pjp_name]["report_count"] * 4) * 100) if pjp_aggregated[pjp_name]["report_count"] > 0 else 0


    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by total LTKT
    top_5_pjps_by_ltkt = sorted(pjps_list, key=lambda item: item[1]["total_ltkt"], reverse=True)[:5]
    bottom_5_pjps_by_ltkt = sorted(pjps_list, key=lambda item: item[1]["total_ltkt"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by average Sipesat compliance (highest is best)
    top_5_pjps_by_sipesat_compliance = sorted(pjps_list, key=lambda item: item[1]["avg_sipesat_compliance"], reverse=True)[:5]
    bottom_5_pjps_by_sipesat_compliance = sorted(pjps_list, key=lambda item: item[1]["avg_sipesat_compliance"])[:5]


    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_ltkt = list(reports)
    reports_by_sipesat_compliance = list(reports)

    # Sort by jumlah_ltkt
    top_5_reports_by_ltkt = sorted(reports_by_ltkt, key=lambda a: a.get("jumlah_ltkt", 0), reverse=True)[:5]
    bottom_5_reports_by_ltkt = sorted(reports_by_ltkt, key=lambda a: a.get("jumlah_ltkt", 0))[:5]

    # Sort by sipesat_uploaded_count (highest is best)
    top_5_reports_by_sipesat_compliance = sorted(reports_by_sipesat_compliance, key=lambda a: a.get("sipesat_uploaded_count", 0), reverse=True)[:5]
    bottom_5_reports_by_sipesat_compliance = sorted(reports_by_sipesat_compliance, key=lambda a: a.get("sipesat_uploaded_count", 0))[:5]

    return {
        "top_5_reports_by_ltkt": top_5_reports_by_ltkt,
        "bottom_5_reports_by_ltkt": bottom_5_reports_by_ltkt,
        "top_5_reports_by_sipesat_compliance": top_5_reports_by_sipesat_compliance,
        "bottom_5_reports_by_sipesat_compliance": bottom_5_reports_by_sipesat_compliance,
        "top_5_pjps_by_ltkt": top_5_pjps_by_ltkt,
        "bottom_5_pjps_by_ltkt": bottom_5_pjps_by_ltkt,
        "top_5_pjps_by_sipesat_compliance": top_5_pjps_by_sipesat_compliance,
        "bottom_5_pjps_by_sipesat_compliance": bottom_5_pjps_by_sipesat_compliance,
    }

# NEW ROUTE: Global APUPPT Analysis Page
@app.route('/analisis_apuppt_global')
def analisis_apuppt_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_apuppt_reports = []
    # Iterate through each APUPPT data
    for report in APUPPT_REPORTS_DATA:
        report_copy = report.copy()
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == report_copy.get("sandi_pjp")), None)
        report_copy["nama_pjp"] = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({report_copy.get('sandi_pjp')})"
        
        # Derive sipesat_uploaded_count
        sipesat_uploaded_count = 0
        if report_copy.get("lapor_sipesat_tw1"): sipesat_uploaded_count += 1
        if report_copy.get("lapor_sipesat_tw2"): sipesat_uploaded_count += 1
        if report_copy.get("lapor_sipesat_tw3"): sipesat_uploaded_count += 1
        if report_copy.get("lapor_sipesat_tw4"): sipesat_uploaded_count += 1
        report_copy["sipesat_uploaded_count"] = sipesat_uploaded_count
        
        # For annual reports, periode_luler and periode_laporan_nama can just be the year
        report_copy["periode_luler"] = str(report_copy.get("tahun_laporan", ""))
        report_copy["periode_laporan_nama"] = str(report_copy.get("tahun_laporan", ""))
        
        all_apuppt_reports.append(report_copy)

    # Sort all reports by year
    all_apuppt_reports.sort(key=lambda x: x.get("tahun_laporan", 0))

    # Get unique years for filtering
    unique_years = sorted(list(set([r.get('tahun_laporan', '') for r in all_apuppt_reports])))
    
    # Transform unique_years for display in dropdown
    display_years = [
        {"value": str(year), "text": str(year)}
        for year in unique_years
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_ltkt_global = sum([report.get("jumlah_ltkt", 0) for report in all_apuppt_reports])
    total_ltkm_global = sum([report.get("jumlah_ltkm", 0) for report in all_apuppt_reports])
    total_ltkl_global = sum([report.get("jumlah_ltkl", 0) for report in all_apuppt_reports])
    total_sipesat_uploaded_count_global = sum([report.get("sipesat_uploaded_count", 0) for report in all_apuppt_reports])
    total_sipesat_expected_count_global = len(all_apuppt_reports) * 4 # 4 quarters per report
    total_reports_global = len(all_apuppt_reports)

    sipesat_percentage_global = (total_sipesat_uploaded_count_global / total_sipesat_expected_count_global * 100) if total_sipesat_expected_count_global > 0 else 0
    avg_ltkt_per_report_global = total_ltkt_global / total_reports_global if total_reports_global > 0 else 0


    kpis_global_apuppt = {
        "total_ltkt_global": total_ltkt_global,
        "total_ltkm_global": total_ltkm_global,
        "total_ltkl_global": total_ltkl_global,
        "total_reports_global": total_reports_global,
        "sipesat_percentage_global": sipesat_percentage_global,
        "avg_ltkt_per_report_global": avg_ltkt_per_report_global,
    }

    # Prepare data for trend charts (yearly only for annual reports)
    yearly_apuppt_trends = {}  # { "YYYY": { "total_ltkt": N, "sipesat_uploaded_count": M } }

    for report in all_apuppt_reports:
        year = report.get("tahun_laporan", "")
        jumlah_ltkt = report.get("jumlah_ltkt", 0)
        sipesat_uploaded_count = report.get("sipesat_uploaded_count", 0)

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_apuppt_trends:
            yearly_apuppt_trends[year_key] = {
                "total_ltkt": 0,
                "sipesat_uploaded_count": 0,
                "report_count": 0
            }
        yearly_apuppt_trends[year_key]["total_ltkt"] += jumlah_ltkt
        yearly_apuppt_trends[year_key]["sipesat_uploaded_count"] += sipesat_uploaded_count
        yearly_apuppt_trends[year_key]["report_count"] += 1

    # Sort yearly trends for chart labels and calculate averages
    sorted_yearly_keys_apuppt = sorted(yearly_apuppt_trends.keys())
    yearly_trend_labels_apuppt = [k for k in sorted_yearly_keys_apuppt]
    yearly_trend_total_ltkt = [yearly_apuppt_trends[k]["total_ltkt"] for k in sorted_yearly_keys_apuppt]
    yearly_trend_avg_sipesat_compliance = [yearly_apuppt_trends[k]["sipesat_uploaded_count"] / (yearly_apuppt_trends[k]["report_count"] * 4) * 100 if yearly_apuppt_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_apuppt]


    trend_chart_data_apuppt = {
        "yearly": {
            "labels": yearly_trend_labels_apuppt,
            "total_ltkt": yearly_trend_total_ltkt,
            "avg_sipesat_compliance": yearly_trend_avg_sipesat_compliance,
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_apuppt = get_top_bottom_analysis_apuppt(all_apuppt_reports)


    return render_template(
        'analisis_apuppt_global.html',
        all_apuppt_reports=all_apuppt_reports, # Pass the Python list directly
        unique_years=display_years, # Pass unique years for filter
        kpis_global=kpis_global_apuppt,
        trend_chart_data=trend_chart_data_apuppt, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_apuppt # Pass initial top/bottom data
    )

# NEW ROUTE: Export APUPPT Data to XLSX
@app.route('/export_apuppt_xlsx')
def export_apuppt_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_year = request.args.get('startYear')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_apuppt_global route
    all_apuppt_reports_raw = []
    for report in APUPPT_REPORTS_DATA:
        report_copy = report.copy()
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == report_copy.get("sandi_pjp")), None)
        report_copy["nama_pjp"] = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({report_copy.get('sandi_pjp')})"
        report_copy["sipesat_uploaded_count"] = (report_copy.get("lapor_sipesat_tw1", 0) + report_copy.get("lapor_sipesat_tw2", 0) +
                                                  report_copy.get("lapor_sipesat_tw3", 0) + report_copy.get("lapor_sipesat_tw4", 0))
        all_apuppt_reports_raw.append(report_copy)
    
    if start_year and end_year:
        start_year_int = int(start_year)
        end_year_int = int(end_year)
        
        filtered_reports = [
            report for report in all_apuppt_reports_raw
            if report.get("tahun_laporan", 0) >= start_year_int and \
               report.get("tahun_laporan", 0) <= end_year_int
        ]
    else:
        filtered_reports = all_apuppt_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan APUPPT Global"

    # Define headers (adjust based on your actual data fields)
    headers = [
        "Tahun Laporan", "Nama PJP", "Sandi PJP", "Nomor Surat", "Tanggal Surat",
        "Jumlah LTKT", "Jumlah LTKM", "Jumlah LTKL",
        "Lapor SIPESAT TW1", "Lapor SIPESAT TW2", "Lapor SIPESAT TW3", "Lapor SIPESAT TW4",
        "Jumlah Lapor Pemblokiran DTTOT", "Expected Lapor Pemblokiran DTTOT",
        "Jumlah Lapor Pemblokiran DPPSPM", "Expected Lapor Pemblokiran DPPSPM",
        "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            str(report.get("nomor_surat", "")),
            str(report.get("tanggal_surat", "")),
            report.get("jumlah_ltkt", 0),
            report.get("jumlah_ltkm", 0),
            report.get("jumlah_ltkl", 0),
            report.get("lapor_sipesat_tw1", False),
            report.get("lapor_sipesat_tw2", False),
            report.get("lapor_sipesat_tw3", False),
            report.get("lapor_sipesat_tw4", False),
            report.get("jumlah_lapor_pemblokiran_dttot", 0),
            report.get("expected_lapor_pemblokiran_dttot", 0),
            report.get("jumlah_lapor_pemblokiran_dppspm", 0),
            report.get("expected_lapor_pemblokiran_dppspm", 0),
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_APUPPT_Global"]
    if start_year and end_year:
        filename_parts.append(f"{start_year}_to_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# ... (kode app.py lainnya tetap sama)

# Helper function to get top/bottom reports and PJPs for Manajemen Reports
def get_top_bottom_analysis_manajemen(reports):
    pjp_aggregated = {} # { "PJP_Name": { "total_reports": N, "overall_compliance_sum": M } }

    for report in reports:
        pjp_name = report.get("nama_pjp", "Tidak Diketahui")
        
        # Calculate overall_compliance_percentage for each report dynamically
        status_fields = [
            "penilaian_perkembangan_industri_sp_status", "pengawasan_perkembangan_bisnis_pjp_status",
            "pengawasan_tata_kelola_pjp_status", "pandangan_perbaikan_pjp_status",
            "perubahan_komposisi_komisaris_status", "implementasi_kebijakan_direksi_status",
            "proses_pencapaian_kinerja_status", "perbandingan_target_dan_realisasi_status",
            "kendala_yang_dihadapi_pjp_status", "tata_kelola_pengawasan_direksi_komisaris_status",
            "tata_kelola_risiko_sdm_status", "tata_kelola_ketersediaan_prosedur_status",
            "tata_kelola_pengendalian_intern_status", "asesmen_kualitatif_sdm_status",
            "asesmen_kuantitatif_sdm_status", "asesmen_kepemilikan_domestik_asing_status",
            "asesmen_pengendalian_domestik_asing_status", "asesmen_pemenuhan_modal_status",
            "asesmen_manajemen_risiko_it_status", "pemantauan_kepatuhan_status"
        ]
        
        report_true_status_count = sum(1 for field in status_fields if report.get(field) is True)
        report_total_status_fields = sum(1 for field in status_fields if field in report)
        
        report_compliance_percentage = (report_true_status_count / report_total_status_fields * 100) if report_total_status_fields > 0 else 0
        
        # Aggregate by PJP
        if pjp_name not in pjp_aggregated:
            pjp_aggregated[pjp_name] = {
                "total_reports": 0,
                "overall_compliance_sum": 0,
                "avg_compliance": 0,
                "total_domestik_ownership": 0, # For averaging
                "total_asing_ownership": 0,   # For averaging
                "total_domestik_control": 0,  # For averaging
                "total_asing_control": 0,     # For averaging
            }
        pjp_aggregated[pjp_name]["total_reports"] += 1
        pjp_aggregated[pjp_name]["overall_compliance_sum"] += report_compliance_percentage
        
        pjp_aggregated[pjp_name]["total_domestik_ownership"] += report.get("persentase_kepemilikan_domestik", 0)
        pjp_aggregated[pjp_name]["total_asing_ownership"] += report.get("persentase_kepemilikan_asing", 0)
        pjp_aggregated[pjp_name]["total_domestik_control"] += report.get("pengendalian_domestik", 0)
        pjp_aggregated[pjp_name]["total_asing_control"] += report.get("pengendalian_asing", 0)

        # Update averages after summing
        pjp_aggregated[pjp_name]["avg_compliance"] = pjp_aggregated[pjp_name]["overall_compliance_sum"] / pjp_aggregated[pjp_name]["total_reports"]
        pjp_aggregated[pjp_name]["avg_domestik_ownership"] = pjp_aggregated[pjp_name]["total_domestik_ownership"] / pjp_aggregated[pjp_name]["total_reports"]
        pjp_aggregated[pjp_name]["avg_asing_ownership"] = pjp_aggregated[pjp_name]["total_asing_ownership"] / pjp_aggregated[pjp_name]["total_reports"]
        pjp_aggregated[pjp_name]["avg_domestik_control"] = pjp_aggregated[pjp_name]["total_domestik_control"] / pjp_aggregated[pjp_name]["total_reports"]
        pjp_aggregated[pjp_name]["avg_asing_control"] = pjp_aggregated[pjp_name]["total_asing_control"] / pjp_aggregated[pjp_name]["total_reports"]


    # Convert to list of (name, data_dict) tuples for PJPs
    pjps_list = list(pjp_aggregated.items())

    # Sort and slice for top/bottom 5 for PJPs by average compliance
    top_5_pjps_by_compliance = sorted(pjps_list, key=lambda item: item[1]["avg_compliance"], reverse=True)[:5]
    bottom_5_pjps_by_compliance = sorted(pjps_list, key=lambda item: item[1]["avg_compliance"])[:5]

    # Sort and slice for top/bottom 5 for PJPs by average domestic ownership
    top_5_pjps_by_domestik_ownership = sorted(pjps_list, key=lambda item: item[1]["avg_domestik_ownership"], reverse=True)[:5]
    bottom_5_pjps_by_domestik_ownership = sorted(pjps_list, key=lambda item: item[1]["avg_domestik_ownership"])[:5]


    # --- Top/Bottom Individual Reports ---
    # Make copies to sort without affecting the original filteredReports
    reports_by_compliance = list(reports)
    reports_by_domestik_ownership = list(reports)

    # Sort by overall_compliance_percentage (derived in route)
    top_5_reports_by_compliance = sorted(reports_by_compliance, key=lambda a: a.get("overall_compliance_percentage", 0), reverse=True)[:5]
    bottom_5_reports_by_compliance = sorted(reports_by_compliance, key=lambda a: a.get("overall_compliance_percentage", 0))[:5]

    # Sort by persentase_kepemilikan_domestik
    top_5_reports_by_domestik_ownership = sorted(reports_by_domestik_ownership, key=lambda a: a.get("persentase_kepemilikan_domestik", 0), reverse=True)[:5]
    bottom_5_reports_by_domestik_ownership = sorted(reports_by_domestik_ownership, key=lambda a: a.get("persentase_kepemilikan_domestik", 0))[:5]


    return {
        "top_5_reports_by_compliance": top_5_reports_by_compliance,
        "bottom_5_reports_by_compliance": bottom_5_reports_by_compliance,
        "top_5_reports_by_domestik_ownership": top_5_reports_by_domestik_ownership,
        "bottom_5_reports_by_domestik_ownership": bottom_5_reports_by_domestik_ownership,
        "top_5_pjps_by_compliance": top_5_pjps_by_compliance,
        "bottom_5_pjps_by_compliance": bottom_5_pjps_by_compliance,
        "top_5_pjps_by_domestik_ownership": top_5_pjps_by_domestik_ownership,
        "bottom_5_pjps_by_domestik_ownership": bottom_5_pjps_by_domestik_ownership,
    }

# NEW ROUTE: Global Manajemen Analysis Page
@app.route('/analisis_manajemen_global')
def analisis_manajemen_global():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    all_manajemen_reports = []
    # Iterate through each PJP's Management data
    for pjp_sandi, pjp_years_data in MANAGEMENT_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        # Iterate through each year for the current PJP
        # Ensure pjp_years_data is a dictionary before calling .items()
        if isinstance(pjp_years_data, dict):
            for year_str, report_data in pjp_years_data.items():
                report_copy = report_data.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)
                # For annual reports, periode_luler and periode_laporan_nama can just be the year
                report_copy["periode_luler"] = year_str 
                report_copy["periode_laporan_nama"] = year_str

                # Calculate overall_compliance_percentage for each report
                status_fields = [
                    "penilaian_perkembangan_industri_sp_status", "pengawasan_perkembangan_bisnis_pjp_status",
                    "pengawasan_tata_kelola_pjp_status", "pandangan_perbaikan_pjp_status",
                    "perubahan_komposisi_komisaris_status", "implementasi_kebijakan_direksi_status",
                    "proses_pencapaian_kinerja_status", "perbandingan_target_dan_realisasi_status",
                    "kendala_yang_dihadapi_pjp_status", "tata_kelola_pengawasan_direksi_komisaris_status",
                    "tata_kelola_risiko_sdm_status", "tata_kelola_ketersediaan_prosedur_status",
                    "tata_kelola_pengendalian_intern_status", "asesmen_kualitatif_sdm_status",
                    "asesmen_kuantitatif_sdm_status", "asesmen_kepemilikan_domestik_asing_status",
                    "asesmen_pengendalian_domestik_asing_status", "asesmen_pemenuhan_modal_status",
                    "asesmen_manajemen_risiko_it_status", "pemantauan_kepatuhan_status"
                ]
                
                report_true_status_count = sum(1 for field in status_fields if report_copy.get(field) is True)
                report_total_status_fields = sum(1 for field in status_fields if field in report_copy)
                
                report_copy["overall_compliance_percentage"] = (report_true_status_count / report_total_status_fields * 100) if report_total_status_fields > 0 else 0

                all_manajemen_reports.append(report_copy)

    # Sort all reports by year
    all_manajemen_reports.sort(key=lambda x: x.get("tahun_laporan", 0))

    # Get unique years for filtering
    unique_years = sorted(list(set([r.get('tahun_laporan', '') for r in all_manajemen_reports])))
    
    # Transform unique_years for display in dropdown
    display_years = [
        {"value": str(year), "text": str(year)}
        for year in unique_years
    ]

    # Calculate aggregated KPIs for the global view (initial load, for all data)
    total_reports_global = len(all_manajemen_reports)
    overall_compliance_sum_global = sum([report.get("overall_compliance_percentage", 0) for report in all_manajemen_reports])
    overall_compliance_percentage_global = overall_compliance_sum_global / total_reports_global if total_reports_global > 0 else 0
    
    avg_domestik_ownership_global = sum([report.get("persentase_kepemilikan_domestik", 0) for report in all_manajemen_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_asing_ownership_global = sum([report.get("persentase_kepemilikan_asing", 0) for report in all_manajemen_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_domestik_control_global = sum([report.get("pengendalian_domestik", 0) for report in all_manajemen_reports]) / total_reports_global if total_reports_global > 0 else 0
    avg_asing_control_global = sum([report.get("pengendalian_asing", 0) for report in all_manajemen_reports]) / total_reports_global if total_reports_global > 0 else 0


    kpis_global_manajemen = {
        "total_reports_global": total_reports_global,
        "overall_compliance_percentage_global": overall_compliance_percentage_global,
        "avg_domestik_ownership_global": avg_domestik_ownership_global,
        "avg_asing_ownership_global": avg_asing_ownership_global,
        "avg_domestik_control_global": avg_domestik_control_global,
        "avg_asing_control_global": avg_asing_control_global,
    }

    # Prepare data for trend charts (yearly only for annual reports)
    yearly_manajemen_trends = {}  # { "YYYY": { "overall_compliance": N, "domestik_ownership": M, ... } }

    for report in all_manajemen_reports:
        year = report.get("tahun_laporan", "")
        overall_compliance = report.get("overall_compliance_percentage", 0)
        domestik_ownership = report.get("persentase_kepemilikan_domestik", 0)
        asing_ownership = report.get("persentase_kepemilikan_asing", 0)
        domestik_control = report.get("pengendalian_domestik", 0)
        asing_control = report.get("pengendalian_asing", 0)

        # Yearly trends
        year_key = str(year)
        if year_key not in yearly_manajemen_trends:
            yearly_manajemen_trends[year_key] = {
                "overall_compliance_sum": 0,
                "domestik_ownership_sum": 0,
                "asing_ownership_sum": 0,
                "domestik_control_sum": 0,
                "asing_control_sum": 0,
                "report_count": 0
            }
        yearly_manajemen_trends[year_key]["overall_compliance_sum"] += overall_compliance
        yearly_manajemen_trends[year_key]["domestik_ownership_sum"] += domestik_ownership
        yearly_manajemen_trends[year_key]["asing_ownership_sum"] += asing_ownership
        yearly_manajemen_trends[year_key]["domestik_control_sum"] += domestik_control
        yearly_manajemen_trends[year_key]["asing_control_sum"] += asing_control
        yearly_manajemen_trends[year_key]["report_count"] += 1
    
    # Sort yearly trends for chart labels and calculate averages
    sorted_yearly_keys_manajemen = sorted(yearly_manajemen_trends.keys())
    yearly_trend_labels_manajemen = [k for k in sorted_yearly_keys_manajemen]
    yearly_trend_overall_compliance = [yearly_manajemen_trends[k]["overall_compliance_sum"] / yearly_manajemen_trends[k]["report_count"] if yearly_manajemen_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_manajemen]
    yearly_trend_domestik_ownership = [yearly_manajemen_trends[k]["domestik_ownership_sum"] / yearly_manajemen_trends[k]["report_count"] if yearly_manajemen_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_manajemen]
    yearly_trend_asing_ownership = [yearly_manajemen_trends[k]["asing_ownership_sum"] / yearly_manajemen_trends[k]["report_count"] if yearly_manajemen_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_manajemen]
    yearly_trend_domestik_control = [yearly_manajemen_trends[k]["domestik_control_sum"] / yearly_manajemen_trends[k]["report_count"] if yearly_manajemen_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_manajemen]
    yearly_trend_asing_control = [yearly_manajemen_trends[k]["asing_control_sum"] / yearly_manajemen_trends[k]["report_count"] if yearly_manajemen_trends[k]["report_count"] > 0 else 0 for k in sorted_yearly_keys_manajemen]


    trend_chart_data_manajemen = {
        "yearly": {
            "labels": yearly_trend_labels_manajemen,
            "overall_compliance": yearly_trend_overall_compliance,
            "domestik_ownership": yearly_trend_domestik_ownership,
            "asing_ownership": yearly_trend_asing_ownership,
            "domestik_control": yearly_trend_domestik_control,
            "asing_control": yearly_trend_asing_control,
        }
    }

    # Prepare data for top/bottom reports and PJPs (for initial load, based on all data)
    initial_top_bottom_data_manajemen = get_top_bottom_analysis_manajemen(all_manajemen_reports)


    return render_template(
        'analisis_manajemen_global.html',
        all_manajemen_reports=all_manajemen_reports, # Pass the Python list directly
        unique_years=display_years, # Pass unique years for filter
        kpis_global=kpis_global_manajemen,
        trend_chart_data=trend_chart_data_manajemen, # Pass the trend data
        initial_top_bottom_data=initial_top_bottom_data_manajemen # Pass initial top/bottom data
    )

# NEW ROUTE: Export Manajemen Data to XLSX
@app.route('/export_manajemen_xlsx')
def export_manajemen_xlsx():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    start_year = request.args.get('startYear')
    end_year = request.args.get('endYear')

    filtered_reports = []
    # Re-apply filtering logic from analisis_manajemen_global route
    all_manajemen_reports_raw = []
    for pjp_sandi, pjp_years_data in MANAGEMENT_REPORTS_DATA.items():
        pjp_info = next((pjp for pjp in PJP_COMPANIES_DATA if pjp["sandi"] == pjp_sandi), None)
        pjp_name = pjp_info["nama"] if pjp_info else f"PJP Tidak Dikenal ({pjp_sandi})"
        
        report_data = pjp_years_data
        if report_data:
            for year_str, report in report_data.items():
                report_copy = report.copy()
                report_copy["sandi_pjp"] = pjp_sandi
                report_copy["nama_pjp"] = pjp_name
                report_copy["tahun_laporan"] = int(year_str)

                status_fields = [
                    "penilaian_perkembangan_industri_sp_status", "pengawasan_perkembangan_bisnis_pjp_status",
                    "pengawasan_tata_kelola_pjp_status", "pandangan_perbaikan_pjp_status",
                    "perubahan_komposisi_komisaris_status", "implementasi_kebijakan_direksi_status",
                    "proses_pencapaian_kinerja_status", "perbandingan_target_dan_realisasi_status",
                    "kendala_yang_dihadapi_pjp_status", "tata_kelola_pengawasan_direksi_komisaris_status",
                    "tata_kelola_risiko_sdm_status", "tata_kelola_ketersediaan_prosedur_status",
                    "tata_kelola_pengendalian_intern_status", "asesmen_kualitatif_sdm_status",
                    "asesmen_kuantitatif_sdm_status", "asesmen_kepemilikan_domestik_asing_status",
                    "asesmen_pengendalian_domestik_asing_status", "asesmen_pemenuhan_modal_status",
                    "asesmen_manajemen_risiko_it_status", "pemantauan_kepatuhan_status"
                ]
                report_true_status_count = sum(1 for field in status_fields if report_copy.get(field) is True)
                report_total_status_fields = sum(1 for field in status_fields if field in report_copy)
                report_copy["overall_compliance_percentage"] = (report_true_status_count / report_total_status_fields * 100) if report_total_status_fields > 0 else 0

                all_manajemen_reports_raw.append(report_copy)
    
    if start_year and end_year:
        start_year_int = int(start_year)
        end_year_int = int(end_year)
        
        filtered_reports = [
            report for report in all_manajemen_reports_raw
            if report.get("tahun_laporan", 0) >= start_year_int and \
               report.get("tahun_laporan", 0) <= end_year_int
        ]
    else:
        filtered_reports = all_manajemen_reports_raw # If no filters, export all

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Manajemen Global"

    # Define headers (adjust based on your actual data fields)
    headers = [
        "Tahun Laporan", "Nama PJP", "Sandi PJP", "Nomor Surat", "Tanggal Surat",
        "Penilaian Perkembangan Industri SP Status", "Pengawasan Perkembangan Bisnis PJP Status",
        "Pengawasan Tata Kelola PJP Status", "Pandangan Perbaikan PJP Status",
        "Perubahan Komposisi Komisaris Status", "Implementasi Kebijakan Direksi Status",
        "Proses Pencapaian Kinerja Status", "Perbandingan Target dan Realisasi Status",
        "Kendala yang Dihadapi PJP Status", "Tata Kelola Pengawasan Direksi Komisaris Status",
        "Tata Kelola Risiko SDM Status", "Tata Kelola Ketersediaan Prosedur Status",
        "Tata Kelola Pengendalian Intern Status", "Asesmen Kualitatif SDM Status",
        "Asesmen Kuantitatif SDM Status", "Asesmen Kepemilikan Domestik Asing Status",
        "Asesmen Pengendalian Domestik Asing Status", "Asesmen Pemenuhan Modal Status",
        "Asesmen Manajemen Risiko IT Status", "Pemantauan Kepatuhan Status",
        "Persentase Kepemilikan Domestik", "Persentase Kepemilikan Asing",
        "Pengendalian Domestik", "Pengendalian Asing",
        "Overall Compliance Percentage", "Created At"
    ]
    sheet.append(headers)

    # Add data rows
    for report in filtered_reports:
        row_data = [
            str(report.get("tahun_laporan", "")),
            str(report.get("nama_pjp", "")),
            str(report.get("sandi_pjp", "")),
            str(report.get("nomor_surat", "")),
            str(report.get("tanggal_surat", "")),
            report.get("penilaian_perkembangan_industri_sp_status", False),
            report.get("pengawasan_perkembangan_bisnis_pjp_status", False),
            report.get("pengawasan_tata_kelola_pjp_status", False),
            report.get("pandangan_perbaikan_pjp_status", False),
            report.get("perubahan_komposisi_komisaris_status", False),
            report.get("implementasi_kebijakan_direksi_status", False),
            report.get("proses_pencapaian_kinerja_status", False),
            report.get("perbandingan_target_dan_realisasi_status", False),
            report.get("kendala_yang_dihadapi_pjp_status", False),
            report.get("tata_kelola_pengawasan_direksi_komisaris_status", False),
            report.get("tata_kelola_risiko_sdm_status", False),
            report.get("tata_kelola_ketersediaan_prosedur_status", False),
            report.get("tata_kelola_pengendalian_intern_status", False),
            report.get("asesmen_kualitatif_sdm_status", False),
            report.get("asesmen_kuantitatif_sdm_status", False),
            report.get("asesmen_kepemilikan_domestik_asing_status", False),
            report.get("asesmen_pengendalian_domestik_asing_status", False),
            report.get("asesmen_pemenuhan_modal_status", False),
            report.get("asesmen_manajemen_risiko_it_status", False),
            report.get("pemantauan_kepatuhan_status", False),
            report.get("persentase_kepemilikan_domestik", 0),
            report.get("persentase_kepemilikan_asing", 0),
            report.get("pengendalian_domestik", 0),
            report.get("pengendalian_asing", 0),
            report.get("overall_compliance_percentage", 0),
            str(report.get("created_at", ""))
        ]
        sheet.append(row_data)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Determine filename based on filters
    filename_parts = ["Laporan_Manajemen_Global"]
    if start_year and end_year:
        filename_parts.append(f"{start_year}_to_{end_year}")
    
    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Rute baru untuk menangani klik tombol "Laporan Kerjasama P2P" yang tidak menyertakan sandi PJP.
# Ini akan mengarahkan pengguna ke halaman analisis laporan utama.
@app.route('/kerjasamap2p_report')
def generic_kerjasamap2p_report_redirect():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('analisis_laporan'))


if __name__ == '__main__':
    app.run(debug=True)

